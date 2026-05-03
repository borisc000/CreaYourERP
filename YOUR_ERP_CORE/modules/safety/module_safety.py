"""
Safety module.
"""

from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Any, Dict, List, Optional, Set, Tuple
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, LongTable

from core.YOUR_ERP_core_framework import BaseModule, Request, Response, ValidationError
from core.YOUR_ERP_orm import BaseModel, Column, ColumnType, AuditMixin
from core.time_utils import utc_now, utc_now_iso, utc_today_iso
from modules.safety.miper_engine import (
    DEFAULT_MASTER_RISKS,
    DEFAULT_PROTOCOLS,
    action_from_vep,
    build_row_fingerprint,
    calculate_vep,
    dedupe_preserve_order,
    legal_requirements_for_headcount,
    merge_generated_rows,
    normalize_control_hierarchy,
    normalize_string_list,
    risk_level_from_vep,
    rule_blueprints,
    summarize_controls,
)
from modules.safety.risk_calculation_service import (
    calculate_compact_miper,
    calculate_risk,
    compact_miper_methodology_payload,
    methodology_payload,
    normalize_task_type_code,
)


PROFILE_RISK_LEVELS = ("low", "medium", "high", "critical")
FOLDER_STATUSES = ("draft", "ready", "in_progress", "closed")
TRAFFIC_LIGHTS = ("red", "yellow", "green")
DOCUMENT_STATUSES = ("draft", "pending_review", "approved", "obsolete", "expired")
DOCUMENT_TYPES = ("procedure", "diffusion", "startup", "record", "other")
MATRIX_STATUSES = ("draft", "pending_review", "approved", "archived")
MATRIX_CANONICAL_SOURCES = ("manual", "procedure", "process", "template", "blocks")
PPE_STATUSES = ("draft", "delivered", "replenishment")
CHECKLIST_RESULTS = ("pending", "ok", "critical")
IRL_STATUSES = ("draft", "issued", "acknowledged")
RULE_SCOPE_TYPES = (
    "transversal",
    "service_profile",
    "customer",
    "client_site",
    "client_area",
)
MATRIX_SOURCE_ORDER = {
    "procedure": 10,
    "cargo_profile": 20,
    "shared_place": 30,
    "equipment": 40,
    "transversal": 50,
    "service_profile": 60,
    "customer": 70,
    "client_site": 80,
    "client_area": 90,
    "fallback": 99,
    "manual": 120,
}
MATRIX_SOURCE_LABELS = {
    "procedure": "Procedimiento",
    "cargo_profile": "Cargo",
    "shared_place": "Sector",
    "equipment": "Equipo",
    "transversal": "Regla transversal",
    "service_profile": "Perfil de servicio",
    "customer": "Cliente",
    "client_site": "Instalacion",
    "client_area": "Area",
    "fallback": "Fallback",
    "manual": "Manual",
}
RESTRICTION_SEVERITIES = ("info", "warning", "blocking")
DEFAULT_PPE_CATALOG = [
    {"code": "CASCO", "name": "Casco de seguridad", "category": "cabeza"},
    {"code": "BARBIQUEJO", "name": "Casco con barbiquejo", "category": "cabeza"},
    {"code": "LENTES", "name": "Lentes de seguridad", "category": "ocular"},
    {"code": "FACIAL", "name": "Protector facial", "category": "facial"},
    {"code": "GUANTES", "name": "Guantes de seguridad", "category": "manos"},
    {"code": "RESPIRADOR", "name": "Respirador", "category": "respiratorio"},
    {"code": "AUDITIVO", "name": "Proteccion auditiva", "category": "auditivo"},
    {"code": "CHALECO", "name": "Chaleco reflectante", "category": "alta_visibilidad"},
    {"code": "CALZADO", "name": "Calzado de seguridad", "category": "pies"},
    {"code": "ARNES", "name": "Arnes de seguridad", "category": "alturas"},
]

CHECKLIST_LIBRARY = [
    {
        "code": "chk_andamios_armado",
        "name": "Checklist armado, uso y desarme de andamios",
        "type": "Andamios",
        "tags": ["andamios", "altura", "critico"],
        "items": [
            "Personal competente y autorizado para armado/desarme",
            "Componentes inspeccionados, compatibles y sin dano visible",
            "Base nivelada, firme y con placas de apoyo cuando corresponda",
            "Accesos, plataformas, barandas y rodapies completos",
            "Area inferior segregada y senalizada",
            "Tarjeta de andamio instalada y estado visible",
            "Herramientas aseguradas contra caida de objetos",
            "Checklist final firmado por supervisor o prevencion",
        ],
    },
    {
        "code": "chk_vehiculos_preuso",
        "name": "Checklist verificacion de vehiculos",
        "type": "Vehiculos",
        "tags": ["vehiculos", "conduccion", "preuso"],
        "items": [
            "Licencia y autorizacion del conductor vigentes",
            "Documentacion del vehiculo vigente y disponible",
            "Luces, frenos, neumaticos y direccion sin observaciones",
            "Extintor, botiquin, triangulos/conos y kit de emergencia disponibles",
            "Parabrisas, espejos y cinturones en buen estado",
            "Carga estibada y sin interferir visibilidad",
            "Ruta y zonas de estacionamiento definidas",
            "Novedades registradas antes de iniciar el traslado",
        ],
    },
    {
        "code": "chk_herramientas_control",
        "name": "Checklist control de herramientas",
        "type": "Herramientas",
        "tags": ["herramientas", "preuso", "control"],
        "items": [
            "Herramienta adecuada para la tarea y sin modificaciones",
            "Protecciones, mangos, carcasa y accesorios en buen estado",
            "Cables, enchufes y extensiones sin cortes ni empalmes inseguros",
            "Discos, brocas o consumibles compatibles y vigentes",
            "EPP definido segun energia, proyeccion, corte o ruido",
            "Herramientas en altura con sistema de amarre cuando aplique",
            "Herramientas defectuosas retiradas, rotuladas y registradas",
            "Area ordenada al cierre y herramientas devueltas",
        ],
    },
]

TALK_LIBRARY = [
    {
        "code": "talk_orden_aseo",
        "topic": "Orden y limpieza del area",
        "category": "general",
        "tags": ["general", "housekeeping"],
        "notes": "Revisar rutas despejadas, retiro de residuos, acopios seguros y responsabilidad diaria del equipo.",
    },
    {
        "code": "talk_epp",
        "topic": "Uso correcto de EPP",
        "category": "general",
        "tags": ["general", "epp"],
        "notes": "Confirmar EPP obligatorio, estado, ajuste, recambio y reporte de elementos deteriorados.",
    },
    {
        "code": "talk_linea_fuego",
        "topic": "Linea de fuego y energia peligrosa",
        "category": "general",
        "tags": ["general", "energia", "linea_fuego"],
        "notes": "Identificar trayectorias de energia, puntos de atrapamiento, cargas suspendidas y barreras.",
    },
    {
        "code": "talk_reporte_incidentes",
        "topic": "Reporte de incidentes y desviaciones",
        "category": "general",
        "tags": ["general", "reporte"],
        "notes": "Reforzar aviso temprano, registro de casi accidentes y cierre de acciones correctivas.",
    },
    {
        "code": "talk_fatiga",
        "topic": "Fatiga, pausas y autocuidado",
        "category": "general",
        "tags": ["general", "fatiga"],
        "notes": "Revisar hidratacion, pausas, alerta a somnolencia y comunicacion de condiciones personales.",
    },
    {
        "code": "talk_altura",
        "topic": "Trabajo en altura y control de caidas",
        "category": "tematica",
        "tags": ["altura", "spdc"],
        "notes": "Verificar autorizacion, puntos de anclaje, rescate, control de bordes y herramientas aseguradas.",
    },
    {
        "code": "talk_andamios",
        "topic": "Uso seguro de andamios",
        "category": "tematica",
        "tags": ["andamios", "altura"],
        "notes": "Revisar tarjeta, plataformas, barandas, acceso seguro, carga maxima y prohibicion de modificar.",
    },
    {
        "code": "talk_vehiculos",
        "topic": "Conduccion defensiva en faena",
        "category": "tematica",
        "tags": ["vehiculos", "conduccion"],
        "notes": "Controlar velocidad, peatones, puntos ciegos, estacionamiento, checklist y uso de cinturon.",
    },
    {
        "code": "talk_herramientas",
        "topic": "Herramientas manuales y electricas",
        "category": "tematica",
        "tags": ["herramientas"],
        "notes": "Inspeccionar protecciones, cables, discos/accesorios, postura de trabajo y retiro por falla.",
    },
]


def _fmt_dt(value: Any) -> Optional[str]:
    if value is None:
        return None
    if callable(value):
        return value().isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    return str(value)


def _safe_int(value: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        if value in (None, ""):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _normalize_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("1", "true", "yes", "on", "si")


def _normalize_str_list(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, str):
        parts = []
        for raw in value.replace("\r", "\n").replace(",", "\n").split("\n"):
            item = raw.strip()
            if item:
                parts.append(item)
        return parts
    return [str(value).strip()] if str(value).strip() else []


def _normalize_int_list(value: Any) -> List[int]:
    result: List[int] = []
    if isinstance(value, list):
        source = value
    else:
        source = _normalize_str_list(value)
    for item in source:
        item_id = _safe_int(item)
        if item_id and item_id not in result:
            result.append(item_id)
    return result


def _compact_miper_from_item(
    item: Dict[str, Any],
    *,
    routine_type: str = "",
    worker_count: Optional[int] = None,
    consequence: Optional[int] = None,
) -> Dict[str, Any]:
    task_type = normalize_task_type_code(
        item.get("task_type_code") or item.get("task_type") or item.get("tipo_tarea"),
        routine_type or str(item.get("routine_type") or ""),
    )
    pe = _safe_int(
        item.get("exposed_people_value")
        or item.get("pe")
        or item.get("personas_expuestas")
        or item.get("worker_count")
        or worker_count,
        1,
    ) or 1
    fe = _safe_int(
        item.get("exposure_frequency_value")
        or item.get("fe")
        or item.get("frecuencia_exposicion"),
        None,
    )
    if fe is None:
        fe = 2 if task_type == "R" else 1
    fo = _safe_int(
        item.get("occurrence_factor_value")
        or item.get("fo")
        or item.get("factor_ocurrencia"),
        1,
    ) or 1
    severity = _safe_int(
        item.get("severity_value")
        or item.get("severity")
        or item.get("severidad")
        or item.get("consequence_value")
        or item.get("consequence")
        or consequence,
        2,
    ) or 2
    compact = calculate_compact_miper(pe, fe, fo, severity)
    compact["task_type_code"] = task_type
    return compact


def _current_controls_text(item: Dict[str, Any], key: str, fallback: Any = "") -> str:
    explicit = item.get(key)
    if explicit not in (None, ""):
        return "; ".join(_normalize_str_list(explicit))
    return "; ".join(_normalize_str_list(fallback))


def _normalize_doc_blueprints(value: Any) -> List[Dict[str, Any]]:
    docs: List[Dict[str, Any]] = []
    if isinstance(value, list):
        raw_docs = value
    else:
        raw_docs = []

    for item in raw_docs:
        if isinstance(item, dict):
            title = str(item.get("title") or item.get("name") or "").strip()
            if not title:
                continue
            docs.append(
                {
                    "code": str(item.get("code") or title.lower().replace(" ", "_")).strip(),
                    "title": title,
                    "document_type": str(item.get("document_type") or "other").strip() or "other",
                    "is_critical": _normalize_bool(item.get("is_critical"), False),
                    "content": str(item.get("content") or "").strip(),
                }
            )
        elif str(item).strip():
            title = str(item).strip()
            docs.append(
                {
                    "code": title.lower().replace(" ", "_"),
                    "title": title,
                    "document_type": "other",
                    "is_critical": False,
                    "content": "",
                }
            )
    return docs


def _normalize_irl_materials(value: Any) -> List[Dict[str, str]]:
    materials: List[Dict[str, str]] = []
    if isinstance(value, list):
        raw_items = value
    else:
        raw_items = _normalize_str_list(value)

    for item in raw_items:
        if isinstance(item, dict):
            name = str(item.get("name") or "").strip()
            if not name:
                continue
            materials.append(
                {
                    "name": name,
                    "type": str(item.get("type") or "otro").strip() or "otro",
                    "location": str(item.get("location") or "").strip(),
                }
            )
            continue
        parts = [part.strip() for part in str(item).split("|")]
        name = parts[0] if parts else ""
        if not name:
            continue
        materials.append(
            {
                "name": name,
                "type": parts[1] if len(parts) > 1 and parts[1] else "otro",
                "location": parts[2] if len(parts) > 2 else "",
            }
        )
    return materials


def _normalize_irl_risk_items(value: Any) -> List[Dict[str, str]]:
    items: List[Dict[str, str]] = []
    if not isinstance(value, list):
        return items
    for item in value:
        if not isinstance(item, dict):
            continue
        risk = str(item.get("risk") or "").strip()
        preventive = str(item.get("preventive_measures") or item.get("preventive") or "").strip()
        method = str(item.get("work_method") or item.get("method") or "").strip()
        if not (risk or preventive or method):
            continue
        items.append(
            {
                "risk": risk,
                "preventive_measures": preventive,
                "work_method": method,
                "source_process": str(item.get("source_process") or "").strip(),
                "source_task": str(item.get("source_task") or "").strip(),
            }
        )
    return items


def _normalize_matrix_rows(value: Any) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not isinstance(value, list):
        return rows
    for item in value:
        if not isinstance(item, dict):
            continue
        process_name = str(item.get("process_name") or item.get("process") or "").strip()
        task_name = str(item.get("task_name") or item.get("task") or item.get("activity") or "").strip()
        hazard = str(item.get("hazard_factor") or item.get("hazard") or "").strip()
        risk = str(item.get("risk_name") or item.get("risk") or "").strip()
        controls = str(item.get("controls") or item.get("control") or "").strip()
        required_ppe = _normalize_str_list(item.get("required_ppe"))
        protocols = _normalize_str_list(item.get("protocol_codes"))
        owner_name = str(item.get("owner_name") or item.get("owner") or "").strip()
        position_name = str(item.get("position_name") or "").strip()
        place_name = str(item.get("place_name") or "").strip()
        master_risk_code = str(item.get("master_risk_code") or "").strip()
        origin_blocks = _normalize_str_list(item.get("origin_blocks"))
        sensitivity_tags = _normalize_str_list(item.get("sensitivity_tags"))
        restriction_alerts = _normalize_str_list(item.get("restriction_alerts"))
        source_group = str(item.get("source_group") or "").strip()
        source_title = str(item.get("source_title") or "").strip()
        source_id = _safe_int(item.get("source_id"), None)
        raw_source_sort = item.get("source_sort")
        source_sort = [
            _safe_int(value, 0) or 0
            for value in (raw_source_sort if isinstance(raw_source_sort, list) else [])
        ][:4]
        source_groups = _normalize_str_list(item.get("source_groups"))
        source_titles = _normalize_str_list(item.get("source_titles"))
        source_refs = item.get("source_refs") if isinstance(item.get("source_refs"), list) else []
        if not any(
            (
                process_name,
                task_name,
                hazard,
                risk,
                controls,
                required_ppe,
                owner_name,
                master_risk_code,
            )
        ):
            continue
        probability = _safe_int(item.get("probability"), 0) or 0
        consequence = _safe_int(item.get("consequence"), 0) or 0
        vep = _safe_int(item.get("vep"), None)
        if vep is None:
            vep = calculate_vep(probability, consequence)
        risk_result = calculate_risk(probability, consequence)
        control_hierarchy = normalize_control_hierarchy(item.get("control_hierarchy"))
        controls = controls or summarize_controls(control_hierarchy, "")
        compact = _compact_miper_from_item(
            item,
            routine_type=str(item.get("routine_type") or ""),
            worker_count=_safe_int(item.get("worker_count"), 1) or 1,
            consequence=consequence,
        )
        proposed_elimination = _normalize_str_list(
            item.get("proposed_elimination_controls") or control_hierarchy.get("elimination")
        )
        proposed_substitution = _normalize_str_list(
            item.get("proposed_substitution_controls") or control_hierarchy.get("substitution")
        )
        proposed_engineering = _normalize_str_list(
            item.get("proposed_engineering_controls") or control_hierarchy.get("engineering")
        )
        proposed_admin = _normalize_str_list(
            item.get("proposed_admin_controls") or control_hierarchy.get("administrative")
        )
        proposed_ppe = _normalize_str_list(
            item.get("proposed_ppe_controls") or control_hierarchy.get("ppe")
        )
        rows.append(
            {
                "activity": task_name or process_name,
                "process_name": process_name,
                "task_name": task_name,
                "position_name": position_name,
                "place_name": place_name,
                "hazard": hazard,
                "hazard_factor": hazard,
                "risk": risk,
                "risk_name": risk,
                "master_risk_id": _safe_int(item.get("master_risk_id"), None),
                "master_risk_code": master_risk_code,
                "risk_family": str(item.get("risk_family") or "").strip(),
                "controls": controls,
                "control_hierarchy": control_hierarchy,
                "required_ppe": required_ppe,
                "protocol_codes": protocols,
                "owner_name": owner_name,
                "probability": probability,
                "consequence": consequence,
                "probability_value": probability,
                "consequence_value": consequence,
                "vep": vep,
                "risk_value": risk_result["risk_level_value"],
                "risk_level_value": risk_result["risk_level_value"],
                "risk_level": str(item.get("risk_level") or risk_result["risk_level_label"]).strip(),
                "risk_level_label": str(item.get("risk_level_label") or risk_result["risk_level_label"]).strip(),
                "severity_color": str(item.get("severity_color") or risk_result["severity_color"]).strip(),
                "task_type_code": compact["task_type_code"],
                "exposed_people_value": compact["exposed_people_value"],
                "exposure_frequency_value": compact["exposure_frequency_value"],
                "occurrence_factor_value": compact["occurrence_factor_value"],
                "probability_score": compact["probability_score"],
                "severity_value": compact["severity_value"],
                "residual_risk_value": compact["residual_risk_value"],
                "residual_risk_label": compact["residual_risk_label"],
                "residual_risk_color": compact["residual_risk_color"],
                "compact_action_required": compact["compact_action_required"],
                "compact_approval_blocked": compact["compact_approval_blocked"],
                "compact_mitigation_required": compact["compact_mitigation_required"],
                "current_engineering_controls": _current_controls_text(
                    item,
                    "current_engineering_controls",
                    control_hierarchy.get("engineering"),
                ),
                "current_admin_controls": _current_controls_text(
                    item,
                    "current_admin_controls",
                    control_hierarchy.get("administrative"),
                ),
                "current_ppe_controls": _current_controls_text(
                    item,
                    "current_ppe_controls",
                    item.get("current_ppe") or required_ppe or control_hierarchy.get("ppe"),
                ),
                "proposed_elimination_controls": proposed_elimination,
                "proposed_substitution_controls": proposed_substitution,
                "proposed_engineering_controls": proposed_engineering,
                "proposed_admin_controls": proposed_admin,
                "proposed_ppe_controls": proposed_ppe,
                "safety_management_plan": str(item.get("safety_management_plan") or item.get("management_plan") or "").strip(),
                "action_required": str(item.get("action_required") or risk_result["action_required"]).strip(),
                "approval_blocked": _normalize_bool(
                    item.get("approval_blocked"),
                    bool(risk_result["approval_blocked"] or compact["compact_approval_blocked"]),
                ),
                "mitigation_required": _normalize_bool(
                    item.get("mitigation_required"),
                    bool(risk_result["mitigation_required"] or compact["compact_mitigation_required"]),
                ),
                "origin_blocks": origin_blocks,
                "origin_rule_ids": _normalize_int_list(item.get("origin_rule_ids")),
                "source_labels": _normalize_str_list(item.get("source_labels")),
                "source_group": source_group,
                "source_title": source_title,
                "source_id": source_id,
                "source_sort": source_sort,
                "source_groups": source_groups,
                "source_titles": source_titles,
                "source_refs": source_refs,
                "sensitivity_tags": sensitivity_tags,
                "restriction_alerts": restriction_alerts,
                "legal_reference": str(item.get("legal_reference") or "").strip(),
                "source_note": str(item.get("source_note") or "").strip(),
                "generated_at": str(item.get("generated_at") or ""),
                "is_blocking": _normalize_bool(
                    item.get("is_blocking"),
                    bool(risk_result["approval_blocked"] or compact["compact_approval_blocked"] or restriction_alerts),
                ),
                "row_fingerprint": str(item.get("row_fingerprint") or "").strip(),
            }
        )
    return rows


def _default_profile_payloads() -> List[Dict[str, Any]]:
    return [
        {
            "name": "Andamios",
            "risk_level": "high",
            "mandatory_documents": [
                {
                    "code": "pts_andamios",
                    "title": "Procedimiento de trabajo seguro para andamios",
                    "document_type": "procedure",
                    "is_critical": True,
                    "content": "Objetivo:\nEjecutar armado, uso y desarme de andamios de forma segura.\n\nSecuencia:\n1. Inspeccionar componentes.\n2. Delimitar area.\n3. Nivelar y anclar.\n4. Verificar plataforma, barandas y accesos.\n5. Autorizar uso solo con checklist conforme.",
                },
                {
                    "code": "difusion_andamios",
                    "title": "Difusion de seguridad del procedimiento de andamios",
                    "document_type": "diffusion",
                    "is_critical": True,
                    "content": "Tema:\nDifusion previa al inicio para caida de altura, caida de objetos y control del area inferior.",
                },
                {
                    "code": "carpeta_arranque",
                    "title": "Control de carpeta de arranque",
                    "document_type": "startup",
                    "is_critical": True,
                    "content": "Checklist base de alistamiento:\n- Personal asignado\n- Matriz aprobada\n- Procedimiento difundido\n- EPP entregado\n- Checklist disponible",
                },
                {
                    "code": "registro_epp",
                    "title": "Registro base de entrega de EPP",
                    "document_type": "record",
                    "is_critical": False,
                    "content": "Control para casco, arnes, cola, linea de vida, guantes y lentes.",
                },
            ],
            "mandatory_ppe": ["Casco", "Arnes de seguridad", "Cola de seguridad", "Linea de vida", "Guantes", "Lentes"],
            "mandatory_checklists": ["Inspeccion diaria de andamio", "Inspeccion de arnes y linea de vida"],
            "recommended_talks": ["Trabajo en altura", "Caida de objetos", "Orden y limpieza en plataforma"],
        },
        {
            "name": "Trabajo en Altura",
            "risk_level": "high",
            "mandatory_documents": [
                {
                    "code": "pts_altura",
                    "title": "Procedimiento de trabajo seguro en altura",
                    "document_type": "procedure",
                    "is_critical": True,
                    "content": "Objetivo:\nControlar riesgos de caida de distinto nivel en tareas sobre 1.8 metros.\n\nControles:\n- Permiso de trabajo\n- Arnes y sistema de anclaje\n- Area delimitada\n- Herramientas aseguradas",
                },
                {
                    "code": "difusion_altura",
                    "title": "Difusion de seguridad para trabajo en altura",
                    "document_type": "diffusion",
                    "is_critical": True,
                    "content": "Tema:\nAntes del inicio se difunden puntos de anclaje, rescate y exclusion de area.",
                },
                {
                    "code": "carpeta_arranque",
                    "title": "Control de carpeta de arranque",
                    "document_type": "startup",
                    "is_critical": True,
                    "content": "Revision documental previa al inicio del servicio.",
                },
            ],
            "mandatory_ppe": ["Casco con barbiquejo", "Arnes de seguridad", "Lente de seguridad", "Guantes"],
            "mandatory_checklists": ["Inspeccion de sistema de anclaje", "Checklist de acceso y trabajo en altura"],
            "recommended_talks": ["Lineas de fuego", "Rescate en altura"],
        },
        {
            "name": "Servicio General",
            "risk_level": "medium",
            "mandatory_documents": [
                {
                    "code": "pts_general",
                    "title": "Procedimiento de trabajo seguro general",
                    "document_type": "procedure",
                    "is_critical": True,
                    "content": "Procedimiento base para actividades operativas generales con control de riesgos y orden de trabajo.",
                },
                {
                    "code": "difusion_general",
                    "title": "Difusion general de seguridad",
                    "document_type": "diffusion",
                    "is_critical": False,
                    "content": "Difusion preventiva del servicio con controles basicos, roles y responsables.",
                },
                {
                    "code": "carpeta_arranque",
                    "title": "Control de carpeta de arranque",
                    "document_type": "startup",
                    "is_critical": True,
                    "content": "Checklist documental previo al inicio del servicio.",
                },
            ],
            "mandatory_ppe": ["Casco", "Guantes", "Lentes"],
            "mandatory_checklists": ["Checklist general de herramientas"],
            "recommended_talks": ["Orden y limpieza", "Observacion preventiva"],
        },
    ]


def _default_risk_rows(profile_name: str) -> List[Dict[str, Any]]:
    normalized = (profile_name or "").strip().lower()
    if "andam" in normalized:
        return [
            {
                "activity": "Armado de andamio",
                "hazard": "Caida de distinto nivel",
                "risk": "Lesion grave o fatalidad",
                "controls": "Arnes, anclaje, inspeccion de estructura, personal competente",
                "required_ppe": ["Casco", "Arnes de seguridad", "Guantes"],
                "owner_name": "Supervisor de terreno",
            },
            {
                "activity": "Uso de andamio",
                "hazard": "Caida de objetos",
                "risk": "Golpe a personal en nivel inferior",
                "controls": "Delimitar area, rodapies, asegurar herramientas",
                "required_ppe": ["Casco", "Lentes"],
                "owner_name": "Supervisor de terreno",
            },
        ]
    if "altura" in normalized:
        return [
            {
                "activity": "Trabajo sobre estructura elevada",
                "hazard": "Caida de altura",
                "risk": "Lesion grave o fatalidad",
                "controls": "Lineas de vida certificadas, permiso de trabajo y plan de rescate",
                "required_ppe": ["Casco con barbiquejo", "Arnes de seguridad"],
                "owner_name": "Prevencionista",
            }
        ]
    return [
        {
            "activity": "Ejecucion del servicio",
            "hazard": "Condicion insegura en area de trabajo",
            "risk": "Lesiones al personal y danos materiales",
            "controls": "Inspeccion previa, charla de inicio y control de EPP",
            "required_ppe": ["Casco", "Guantes", "Lentes"],
            "owner_name": "Supervisor",
        }
    ]


class SafetyServiceProfile(BaseModel, AuditMixin):
    __tablename__ = "safety_service_profiles"
    __displayname__ = "name"

    name = Column(ColumnType.STRING, required=True, label="Name")
    service_type_id = Column(ColumnType.INTEGER, label="Service Type")
    risk_level = Column(ColumnType.STRING, default="medium", label="Risk Level")
    mandatory_documents = Column(ColumnType.JSON, default=[], label="Mandatory Documents")
    mandatory_ppe = Column(ColumnType.JSON, default=[], label="Mandatory PPE")
    mandatory_checklists = Column(ColumnType.JSON, default=[], label="Mandatory Checklists")
    recommended_talks = Column(ColumnType.JSON, default=[], label="Recommended Talks")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.mandatory_documents = _normalize_doc_blueprints(self.mandatory_documents)
        self.mandatory_ppe = _normalize_str_list(self.mandatory_ppe)
        self.mandatory_checklists = _normalize_str_list(self.mandatory_checklists)
        self.recommended_talks = _normalize_str_list(self.recommended_talks)

    def before_save(self):
        self.mandatory_documents = _normalize_doc_blueprints(self.mandatory_documents)
        self.mandatory_ppe = _normalize_str_list(self.mandatory_ppe)
        self.mandatory_checklists = _normalize_str_list(self.mandatory_checklists)
        self.recommended_talks = _normalize_str_list(self.recommended_talks)

    def validate(self):
        super().validate()
        if not (self.name or "").strip():
            raise ValidationError("Service profile name is required")
        if self.risk_level not in PROFILE_RISK_LEVELS:
            raise ValidationError(
                f"Risk level must be one of: {', '.join(PROFILE_RISK_LEVELS)}"
            )

    def to_dict(self) -> Dict[str, Any]:
        service_type_name = None
        if self.service_type_id:
            try:
                from modules.crm.module_crm import ServiceType

                service_type = ServiceType.find_by_id(self.service_type_id)
                service_type_name = service_type.name if service_type else None
            except Exception:
                service_type_name = None
        return {
            "id": self.id,
            "name": self.name or "",
            "service_type_id": self.service_type_id,
            "service_type_name": service_type_name,
            "risk_level": self.risk_level or "medium",
            "mandatory_documents": _normalize_doc_blueprints(self.mandatory_documents),
            "mandatory_ppe": _normalize_str_list(self.mandatory_ppe),
            "mandatory_checklists": _normalize_str_list(self.mandatory_checklists),
            "recommended_talks": _normalize_str_list(self.recommended_talks),
            "active": bool(self.active),
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyFolder(BaseModel, AuditMixin):
    __tablename__ = "safety_folders"
    __displayname__ = "lead_id"

    lead_id = Column(ColumnType.INTEGER, required=True, label="Lead")
    service_profile_id = Column(ColumnType.INTEGER, label="Service Profile")
    procedure_id = Column(ColumnType.INTEGER, label="Procedure")
    client_site_id = Column(ColumnType.INTEGER, label="Client Site")
    client_area_id = Column(ColumnType.INTEGER, label="Client Area")
    procedure_ids = Column(ColumnType.JSON, default=[], label="Procedures")
    job_profile_ids = Column(ColumnType.JSON, default=[], label="Job Profiles")
    client_area_ids = Column(ColumnType.JSON, default=[], label="Client Areas")
    equipment_block_ids = Column(ColumnType.JSON, default=[], label="Equipment Blocks")
    status = Column(ColumnType.STRING, default="draft", label="Status")
    readiness_pct = Column(ColumnType.FLOAT, default=0.0, label="Readiness")
    traffic_light = Column(ColumnType.STRING, default="red", label="Traffic Light")
    planned_start_date = Column(ColumnType.STRING, label="Planned Start Date")
    assigned_employee_ids = Column(ColumnType.JSON, default=[], label="Assigned Employees")
    notes = Column(ColumnType.TEXT, label="Notes")
    miper_scope_notes = Column(ColumnType.TEXT, label="MIPER Scope Notes")
    approver_user_id = Column(ColumnType.INTEGER, label="Approved By")
    approved_at = Column(ColumnType.DATETIME, label="Approved At")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def _normalize_scope(self):
        self.procedure_ids = _normalize_int_list(self.procedure_ids)
        self.job_profile_ids = _normalize_int_list(self.job_profile_ids)
        self.client_area_ids = _normalize_int_list(self.client_area_ids)
        self.equipment_block_ids = _normalize_int_list(self.equipment_block_ids)
        if self.procedure_id and self.procedure_id not in self.procedure_ids:
            self.procedure_ids = [self.procedure_id] + self.procedure_ids
        if self.client_area_id and self.client_area_id not in self.client_area_ids:
            self.client_area_ids = [self.client_area_id] + self.client_area_ids
        if self.procedure_ids:
            self.procedure_id = self.procedure_ids[0]
        if self.client_area_ids:
            self.client_area_id = self.client_area_ids[0]

    def before_create(self):
        self.assigned_employee_ids = _normalize_int_list(self.assigned_employee_ids)
        self._normalize_scope()
        if not self.status:
            self.status = "draft"
        if not self.traffic_light:
            self.traffic_light = "red"

    def before_save(self):
        self.assigned_employee_ids = _normalize_int_list(self.assigned_employee_ids)
        self._normalize_scope()

    def validate(self):
        super().validate()
        if self.status not in FOLDER_STATUSES:
            raise ValidationError(f"Folder status must be one of: {', '.join(FOLDER_STATUSES)}")
        if self.traffic_light not in TRAFFIC_LIGHTS:
            raise ValidationError(
                f"Traffic light must be one of: {', '.join(TRAFFIC_LIGHTS)}"
            )

    def to_dict(self) -> Dict[str, Any]:
        profile = SafetyServiceProfile.find_by_id(self.service_profile_id) if self.service_profile_id else None
        procedure = None
        if self.procedure_id:
            try:
                from modules.safety_procedures.module_safety_procedures import SafetyProcedureTemplate

                procedure = SafetyProcedureTemplate.find_by_id(self.procedure_id)
            except Exception:
                procedure = None
        site = SafetyClientSite.find_by_id(self.client_site_id) if self.client_site_id else None
        area = SafetyClientArea.find_by_id(self.client_area_id) if self.client_area_id else None
        procedures: List[Dict[str, Any]] = []
        job_profiles: List[Dict[str, Any]] = []
        client_areas: List[Dict[str, Any]] = []
        equipment_blocks: List[Dict[str, Any]] = []
        try:
            from modules.safety_procedures.module_safety_procedures import SafetyProcedureTemplate

            for procedure_id in _normalize_int_list(self.procedure_ids):
                item = SafetyProcedureTemplate.find_by_id(procedure_id)
                if item:
                    procedures.append(item.to_dict())
        except Exception:
            procedures = []
        try:
            from modules.job_profiles.module_job_profiles import JobProfile

            for profile_id in _normalize_int_list(self.job_profile_ids):
                item = JobProfile.find_by_id(profile_id)
                if item:
                    job_profiles.append(item.to_dict())
        except Exception:
            job_profiles = []
        for area_id in _normalize_int_list(self.client_area_ids):
            item = SafetyClientArea.find_by_id(area_id)
            if item:
                client_areas.append(item.to_dict())
        for equipment_id in _normalize_int_list(self.equipment_block_ids):
            item = SafetyEquipmentBlock.find_by_id(equipment_id)
            if item:
                equipment_blocks.append(item.to_dict())
        lead_title = None
        project_code = None
        customer_id = None
        customer_name = None
        service_type_id = None
        service_type_name = None
        if self.lead_id:
            try:
                from modules.crm.module_crm import Lead, Customer, ServiceType

                lead = Lead.find_by_id(self.lead_id)
                if lead:
                    lead_title = lead.title
                    project_code = lead.project_code
                    if lead.customer_id:
                        customer_id = lead.customer_id
                        customer = Customer.find_by_id(lead.customer_id)
                        customer_name = customer.name if customer else None
                    if lead.service_type_id:
                        service_type_id = lead.service_type_id
                        service_type = ServiceType.find_by_id(lead.service_type_id)
                        service_type_name = service_type.name if service_type else None
            except Exception:
                pass
        return {
            "id": self.id,
            "lead_id": self.lead_id,
            "lead_title": lead_title,
            "project_code": project_code,
            "customer_id": customer_id,
            "customer_name": customer_name,
            "service_type_id": service_type_id,
            "service_type_name": service_type_name,
            "service_profile_id": self.service_profile_id,
            "service_profile_name": profile.name if profile else None,
            "procedure_id": self.procedure_id,
            "procedure_code": procedure.procedure_code if procedure else None,
            "procedure_name": procedure.name if procedure else None,
            "procedure_ids": _normalize_int_list(self.procedure_ids),
            "procedures": procedures,
            "procedure_names": [item.get("name") or item.get("procedure_code") for item in procedures],
            "procedure_count": len(procedures),
            "client_site_id": self.client_site_id,
            "client_site_name": site.name if site else None,
            "client_area_id": self.client_area_id,
            "client_area_name": area.name if area else None,
            "client_area_ids": _normalize_int_list(self.client_area_ids),
            "client_areas": client_areas,
            "client_area_names": [item.get("name") for item in client_areas],
            "client_area_count": len(client_areas),
            "job_profile_ids": _normalize_int_list(self.job_profile_ids),
            "job_profiles": job_profiles,
            "job_profile_names": [item.get("name") or item.get("code") for item in job_profiles],
            "job_profile_count": len(job_profiles),
            "equipment_block_ids": _normalize_int_list(self.equipment_block_ids),
            "equipment_blocks": equipment_blocks,
            "equipment_block_names": [item.get("name") or item.get("code") for item in equipment_blocks],
            "equipment_block_count": len(equipment_blocks),
            "status": self.status or "draft",
            "readiness_pct": round(float(self.readiness_pct or 0.0), 1),
            "traffic_light": self.traffic_light or "red",
            "planned_start_date": self.planned_start_date or "",
            "assigned_employee_ids": _normalize_int_list(self.assigned_employee_ids),
            "notes": self.notes or "",
            "miper_scope_notes": self.miper_scope_notes or "",
            "approver_user_id": self.approver_user_id,
            "approved_at": _fmt_dt(self._data.get("approved_at")),
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyFolderDocument(BaseModel, AuditMixin):
    __tablename__ = "safety_folder_documents"
    __displayname__ = "title"

    folder_id = Column(ColumnType.INTEGER, required=True, label="Folder")
    code = Column(ColumnType.STRING, required=True, label="Code")
    title = Column(ColumnType.STRING, required=True, label="Title")
    document_type = Column(ColumnType.STRING, default="other", label="Document Type")
    status = Column(ColumnType.STRING, default="draft", label="Status")
    version = Column(ColumnType.INTEGER, default=1, label="Version")
    is_critical = Column(ColumnType.BOOLEAN, default=False, label="Critical")
    owner_user_id = Column(ColumnType.INTEGER, label="Owner")
    due_date = Column(ColumnType.STRING, label="Due Date")
    content = Column(ColumnType.TEXT, label="Content")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def validate(self):
        super().validate()
        if self.document_type not in DOCUMENT_TYPES:
            raise ValidationError(
                f"Document type must be one of: {', '.join(DOCUMENT_TYPES)}"
            )
        if self.status not in DOCUMENT_STATUSES:
            raise ValidationError(
                f"Document status must be one of: {', '.join(DOCUMENT_STATUSES)}"
            )

    def to_dict(self) -> Dict[str, Any]:
        owner_name = None
        if self.owner_user_id:
            try:
                from modules.base.module_base import User

                owner = User.find_by_id(self.owner_user_id)
                owner_name = owner.name if owner else None
            except Exception:
                owner_name = None
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "code": self.code or "",
            "title": self.title or "",
            "document_type": self.document_type or "other",
            "status": self.status or "draft",
            "version": _safe_int(self.version, 1) or 1,
            "is_critical": bool(self.is_critical),
            "owner_user_id": self.owner_user_id,
            "owner_name": owner_name,
            "due_date": self.due_date or "",
            "content": self.content or "",
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyRiskMatrix(BaseModel, AuditMixin):
    __tablename__ = "safety_risk_matrices"
    __displayname__ = "title"

    folder_id = Column(ColumnType.INTEGER, label="Folder")
    code = Column(ColumnType.STRING, label="Code")
    title = Column(ColumnType.STRING, required=True, label="Title")
    status = Column(ColumnType.STRING, default="draft", label="Status")
    version = Column(ColumnType.INTEGER, default=1, label="Version")
    rows = Column(ColumnType.JSON, default=[], label="Rows")
    generation_summary = Column(ColumnType.JSON, default={}, label="Generation Summary")
    reviewed_by = Column(ColumnType.INTEGER, label="Reviewed By")
    reviewed_at = Column(ColumnType.DATETIME, label="Reviewed At")
    source_type = Column(ColumnType.STRING, default="manual", label="Source Type")
    source_id = Column(ColumnType.INTEGER, label="Source")
    process_name = Column(ColumnType.STRING, label="Process")
    work_center = Column(ColumnType.STRING, label="Work Center")
    elaboration_date = Column(ColumnType.STRING, label="Elaboration Date")
    last_update_date = Column(ColumnType.STRING, label="Last Update")
    review_due_date = Column(ColumnType.STRING, label="Review Due")
    methodology_config_id = Column(ColumnType.INTEGER, label="Methodology")
    project_id = Column(ColumnType.INTEGER, label="Project")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.code = str(self.code or "").strip().upper() or f"MIPER-{utc_today_iso().replace('-', '')}"
        self.source_type = str(self.source_type or "manual").strip() or "manual"
        self.elaboration_date = self.elaboration_date or utc_today_iso()
        self.last_update_date = self.last_update_date or self.elaboration_date
        self.rows = _normalize_matrix_rows(self.rows)
        self.generation_summary = self.generation_summary if isinstance(self.generation_summary, dict) else {}

    def before_save(self):
        self.before_create()
        self.generation_summary = self.generation_summary if isinstance(self.generation_summary, dict) else {}

    def validate(self):
        super().validate()
        if self.status not in MATRIX_STATUSES:
            raise ValidationError(f"Matrix status must be one of: {', '.join(MATRIX_STATUSES)}")

    def to_dict(self) -> Dict[str, Any]:
        row_records = [
            row.to_dict()
            for row in SafetyRiskMatrixRow.search([("risk_matrix_id", "=", self.id)])
            if row.active
        ] if self.id else []
        row_records.sort(key=lambda item: (item.get("sequence") or 0, item.get("id") or 0))
        rows = row_records or _normalize_matrix_rows(self.rows)
        intolerable_count = len(
            [
                row
                for row in rows
                if row.get("risk_level") == "Intolerable"
                or row.get("risk_level_label") == "Intolerable"
                or row.get("residual_risk_label") == "No aceptable"
            ]
        )
        important_count = len(
            [
                row
                for row in rows
                if row.get("risk_level") == "Importante"
                or row.get("risk_level_label") == "Importante"
                or row.get("residual_risk_label") == "Importante"
            ]
        )
        protocol_codes = dedupe_preserve_order(
            [code for row in rows for code in _normalize_str_list(row.get("protocol_codes"))]
        )
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "code": self.code or "",
            "title": self.title or "",
            "status": self.status or "draft",
            "version": _safe_int(self.version, 1) or 1,
            "rows": rows,
            "row_count": len(rows),
            "intolerable_count": intolerable_count,
            "important_count": important_count,
            "approval_blocked": intolerable_count > 0,
            "mitigation_required": bool(intolerable_count or important_count),
            "protocol_codes": protocol_codes,
            "generation_summary": self.generation_summary if isinstance(self.generation_summary, dict) else {},
            "reviewed_by": self.reviewed_by,
            "reviewed_at": _fmt_dt(self._data.get("reviewed_at")),
            "source_type": self.source_type or "manual",
            "source_id": self.source_id,
            "process_name": self.process_name or "",
            "work_center": self.work_center or "",
            "elaboration_date": self.elaboration_date or "",
            "last_update_date": self.last_update_date or "",
            "review_due_date": self.review_due_date or "",
            "methodology_config_id": self.methodology_config_id,
            "project_id": self.project_id,
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyRiskMatrixRow(BaseModel, AuditMixin):
    __tablename__ = "safety_risk_matrix_rows"
    __displayname__ = "activity_name"

    company_id = Column(ColumnType.INTEGER, required=True, label="Company")
    risk_matrix_id = Column(ColumnType.INTEGER, required=True, label="Risk Matrix")
    source_block_id = Column(ColumnType.INTEGER, label="Source Block")
    source_procedure_step_id = Column(ColumnType.INTEGER, label="Source Procedure Step")
    activity_name = Column(ColumnType.STRING, label="Activity")
    task_name = Column(ColumnType.STRING, label="Task")
    job_position = Column(ColumnType.STRING, label="Job Position")
    specific_workplace = Column(ColumnType.STRING, label="Specific Workplace")
    worker_count = Column(ColumnType.INTEGER, default=1, label="Workers")
    routine_type = Column(ColumnType.STRING, default="routine", label="Routine Type")
    hazard_name = Column(ColumnType.TEXT, label="Hazard")
    risk_name = Column(ColumnType.TEXT, label="Risk")
    probable_damage = Column(ColumnType.TEXT, label="Probable Damage")
    probability_value = Column(ColumnType.INTEGER, default=2, label="Probability")
    consequence_value = Column(ColumnType.INTEGER, default=2, label="Consequence")
    risk_value = Column(ColumnType.INTEGER, label="Risk Value")
    risk_level_label = Column(ColumnType.STRING, label="Risk Level")
    task_type_code = Column(ColumnType.STRING, default="R", label="Task Type")
    exposed_people_value = Column(ColumnType.INTEGER, default=1, label="Exposed People")
    exposure_frequency_value = Column(ColumnType.INTEGER, default=1, label="Exposure Frequency")
    occurrence_factor_value = Column(ColumnType.INTEGER, default=1, label="Occurrence Factor")
    probability_score = Column(ColumnType.INTEGER, label="MIPER Probability")
    severity_value = Column(ColumnType.INTEGER, default=2, label="Severity")
    residual_risk_value = Column(ColumnType.INTEGER, label="Residual Risk")
    residual_risk_label = Column(ColumnType.STRING, label="Residual Risk Label")
    current_engineering_controls = Column(ColumnType.TEXT, label="Current Engineering Controls")
    current_admin_controls = Column(ColumnType.TEXT, label="Current Admin Controls")
    current_ppe_controls = Column(ColumnType.TEXT, label="Current PPE Controls")
    proposed_elimination_controls = Column(ColumnType.JSON, default=[], label="Proposed Elimination")
    proposed_substitution_controls = Column(ColumnType.JSON, default=[], label="Proposed Substitution")
    proposed_engineering_controls = Column(ColumnType.JSON, default=[], label="Proposed Engineering")
    proposed_admin_controls = Column(ColumnType.JSON, default=[], label="Proposed Admin")
    proposed_ppe_controls = Column(ColumnType.JSON, default=[], label="Proposed PPE")
    safety_management_plan = Column(ColumnType.TEXT, label="S&SO Plan")
    existing_controls = Column(ColumnType.TEXT, label="Existing Controls")
    required_controls = Column(ColumnType.JSON, default={}, label="Required Controls")
    ppe_summary = Column(ColumnType.JSON, default=[], label="PPE")
    protocols_summary = Column(ColumnType.JSON, default=[], label="Protocols")
    legal_reference = Column(ColumnType.TEXT, label="Legal Reference")
    responsible = Column(ColumnType.STRING, label="Responsible")
    observations = Column(ColumnType.TEXT, label="Observations")
    sequence = Column(ColumnType.INTEGER, default=10, label="Sequence")
    origin_blocks = Column(ColumnType.JSON, default=[], label="Origin Blocks")
    source_labels = Column(ColumnType.JSON, default=[], label="Source Labels")
    source_group = Column(ColumnType.STRING, label="Source Group")
    source_title = Column(ColumnType.STRING, label="Source Title")
    source_ref_id = Column(ColumnType.INTEGER, label="Source Ref")
    severity_color = Column(ColumnType.STRING, label="Severity Color")
    approval_blocked = Column(ColumnType.BOOLEAN, default=False, label="Approval Blocked")
    mitigation_required = Column(ColumnType.BOOLEAN, default=False, label="Mitigation Required")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")

    def before_create(self):
        self.required_controls = normalize_control_hierarchy(self.required_controls)
        self.ppe_summary = _normalize_str_list(self.ppe_summary)
        self.protocols_summary = _normalize_str_list(self.protocols_summary)
        self.origin_blocks = _normalize_str_list(self.origin_blocks)
        self.source_labels = _normalize_str_list(self.source_labels)
        risk_result = calculate_risk(self.probability_value, self.consequence_value)
        self.task_type_code = normalize_task_type_code(self.task_type_code, self.routine_type or "")
        self.proposed_elimination_controls = _normalize_str_list(
            self.proposed_elimination_controls or self.required_controls.get("elimination")
        )
        self.proposed_substitution_controls = _normalize_str_list(
            self.proposed_substitution_controls or self.required_controls.get("substitution")
        )
        self.proposed_engineering_controls = _normalize_str_list(
            self.proposed_engineering_controls or self.required_controls.get("engineering")
        )
        self.proposed_admin_controls = _normalize_str_list(
            self.proposed_admin_controls or self.required_controls.get("administrative")
        )
        self.proposed_ppe_controls = _normalize_str_list(
            self.proposed_ppe_controls or self.required_controls.get("ppe")
        )
        compact = calculate_compact_miper(
            self.exposed_people_value,
            self.exposure_frequency_value,
            self.occurrence_factor_value,
            self.severity_value or self.consequence_value,
        )
        self.exposed_people_value = compact["exposed_people_value"]
        self.exposure_frequency_value = compact["exposure_frequency_value"]
        self.occurrence_factor_value = compact["occurrence_factor_value"]
        self.probability_score = compact["probability_score"]
        self.severity_value = compact["severity_value"]
        self.residual_risk_value = compact["residual_risk_value"]
        self.residual_risk_label = compact["residual_risk_label"]
        self.risk_value = compact["residual_risk_value"] or risk_result["risk_level_value"]
        self.risk_level_label = compact["residual_risk_label"] or risk_result["risk_level_label"]
        self.severity_color = compact["residual_risk_color"] or risk_result["severity_color"]
        self.approval_blocked = bool(risk_result["approval_blocked"] or compact["compact_approval_blocked"])
        self.mitigation_required = bool(risk_result["mitigation_required"] or compact["compact_mitigation_required"])

    def before_save(self):
        self.before_create()

    def to_dict(self) -> Dict[str, Any]:
        probability = _safe_int(self.probability_value, 2) or 2
        consequence = _safe_int(self.consequence_value, 2) or 2
        risk_result = calculate_risk(probability, consequence)
        compact = calculate_compact_miper(
            self.exposed_people_value,
            self.exposure_frequency_value,
            self.occurrence_factor_value,
            self.severity_value or consequence,
        )
        controls = normalize_control_hierarchy(self.required_controls)
        origin_blocks = _normalize_str_list(self.origin_blocks)
        if not origin_blocks and self.source_procedure_step_id:
            origin_blocks = ["procedure"]
        residual_value = compact["residual_risk_value"]
        residual_label = compact["residual_risk_label"]
        residual_color = compact["residual_risk_color"]
        return {
            "id": self.id,
            "company_id": self.company_id,
            "risk_matrix_id": self.risk_matrix_id,
            "source_block_id": self.source_block_id,
            "source_procedure_step_id": self.source_procedure_step_id,
            "activity": self.activity_name or "",
            "activity_name": self.activity_name or "",
            "task_name": self.task_name or "",
            "position_name": self.job_position or "",
            "job_position": self.job_position or "",
            "place_name": self.specific_workplace or "",
            "specific_workplace": self.specific_workplace or "",
            "worker_count": _safe_int(self.worker_count, 1) or 1,
            "routine_type": self.routine_type or "routine",
            "task_type_code": normalize_task_type_code(self.task_type_code, self.routine_type or ""),
            "hazard": self.hazard_name or "",
            "hazard_name": self.hazard_name or "",
            "risk": self.risk_name or "",
            "risk_name": self.risk_name or "",
            "probable_damage": self.probable_damage or "",
            "probability": probability,
            "probability_value": probability,
            "consequence": consequence,
            "consequence_value": consequence,
            "vep": residual_value,
            "risk_value": residual_value,
            "risk_level_value": residual_value,
            "risk_level": residual_label,
            "risk_level_label": residual_label,
            "severity_color": residual_color,
            "exposed_people_value": compact["exposed_people_value"],
            "exposure_frequency_value": compact["exposure_frequency_value"],
            "occurrence_factor_value": compact["occurrence_factor_value"],
            "probability_score": compact["probability_score"],
            "severity_value": compact["severity_value"],
            "residual_risk_value": residual_value,
            "residual_risk_label": residual_label,
            "residual_risk_color": residual_color,
            "compact_action_required": compact["compact_action_required"],
            "compact_approval_blocked": compact["compact_approval_blocked"],
            "compact_mitigation_required": compact["compact_mitigation_required"],
            "legacy_risk_value": risk_result["risk_level_value"],
            "legacy_risk_level_label": risk_result["risk_level_label"],
            "action_required": compact["compact_action_required"] or risk_result["action_required"],
            "approval_blocked": bool(risk_result["approval_blocked"] or compact["compact_approval_blocked"]),
            "mitigation_required": bool(risk_result["mitigation_required"] or compact["compact_mitigation_required"]),
            "controls": self.existing_controls or summarize_controls(controls, ""),
            "existing_controls": self.existing_controls or "",
            "control_hierarchy": controls,
            "required_controls": controls,
            "current_engineering_controls": self.current_engineering_controls or "",
            "current_admin_controls": self.current_admin_controls or "",
            "current_ppe_controls": self.current_ppe_controls or "",
            "proposed_elimination_controls": _normalize_str_list(self.proposed_elimination_controls),
            "proposed_substitution_controls": _normalize_str_list(self.proposed_substitution_controls),
            "proposed_engineering_controls": _normalize_str_list(self.proposed_engineering_controls),
            "proposed_admin_controls": _normalize_str_list(self.proposed_admin_controls),
            "proposed_ppe_controls": _normalize_str_list(self.proposed_ppe_controls),
            "safety_management_plan": self.safety_management_plan or "",
            "required_ppe": _normalize_str_list(self.ppe_summary),
            "ppe_summary": _normalize_str_list(self.ppe_summary),
            "protocol_codes": _normalize_str_list(self.protocols_summary),
            "protocols_summary": _normalize_str_list(self.protocols_summary),
            "legal_reference": self.legal_reference or "",
            "owner_name": self.responsible or "",
            "responsible": self.responsible or "",
            "observations": self.observations or "",
            "sequence": _safe_int(self.sequence, 10) or 10,
            "origin_blocks": origin_blocks,
            "source_labels": _normalize_str_list(self.source_labels),
            "source_group": self.source_group or ("procedure" if self.source_procedure_step_id else ""),
            "source_title": self.source_title or "",
            "source_id": self.source_ref_id,
            "active": bool(self.active),
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyRiskMethodology(BaseModel, AuditMixin):
    __tablename__ = "safety_risk_methodologies"
    __displayname__ = "name"

    company_id = Column(ColumnType.INTEGER, required=True, label="Company")
    name = Column(ColumnType.STRING, required=True, label="Name")
    probability_schema_json = Column(ColumnType.JSON, default=[], label="Probability Schema")
    consequence_schema_json = Column(ColumnType.JSON, default=[], label="Consequence Schema")
    risk_matrix_schema_json = Column(ColumnType.JSON, default=[], label="Risk Matrix Schema")
    default_flag = Column(ColumnType.BOOLEAN, default=False, label="Default")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")

    def before_create(self):
        payload = methodology_payload()
        self.probability_schema_json = self.probability_schema_json or payload.get("probability_schema") or []
        self.consequence_schema_json = self.consequence_schema_json or payload.get("consequence_schema") or []
        self.risk_matrix_schema_json = self.risk_matrix_schema_json or payload.get("risk_matrix_schema") or []

    def before_save(self):
        self.before_create()

    def to_methodology_dict(self) -> Dict[str, Any]:
        return {
            "name": self.name or "MIPER/IPER",
            "probability_schema": self.probability_schema_json or [],
            "consequence_schema": self.consequence_schema_json or [],
            "risk_matrix_schema": self.risk_matrix_schema_json or [],
            "default_flag": bool(self.default_flag),
        }

    def to_dict(self) -> Dict[str, Any]:
        payload = self.to_methodology_dict()
        payload.update(
            {
                "id": self.id,
                "company_id": self.company_id,
                "active": bool(self.active),
            }
        )
        return payload


class SafetyPPEDelivery(BaseModel, AuditMixin):
    __tablename__ = "safety_ppe_deliveries"
    __displayname__ = "employee_id"

    folder_id = Column(ColumnType.INTEGER, required=True, label="Folder")
    employee_id = Column(ColumnType.INTEGER, required=True, label="Employee")
    delivery_date = Column(ColumnType.STRING, required=True, label="Delivery Date")
    status = Column(ColumnType.STRING, default="draft", label="Status")
    items = Column(ColumnType.JSON, default=[], label="Items")
    notes = Column(ColumnType.TEXT, label="Notes")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.items = _normalize_str_list(self.items)

    def before_save(self):
        self.items = _normalize_str_list(self.items)

    def validate(self):
        super().validate()
        if self.status not in PPE_STATUSES:
            raise ValidationError(f"PPE status must be one of: {', '.join(PPE_STATUSES)}")

    def to_dict(self) -> Dict[str, Any]:
        employee_name = None
        employee_code = None
        if self.employee_id:
            try:
                from modules.hr.module_hr import EmployeeProfile

                employee = EmployeeProfile.find_by_id(self.employee_id)
                if employee:
                    employee_name = employee.full_name
                    employee_code = employee.employee_code
            except Exception:
                pass
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "employee_id": self.employee_id,
            "employee_name": employee_name,
            "employee_code": employee_code,
            "delivery_date": self.delivery_date or "",
            "status": self.status or "draft",
            "items": _normalize_str_list(self.items),
            "notes": self.notes or "",
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyTalk(BaseModel, AuditMixin):
    __tablename__ = "safety_talks"
    __displayname__ = "topic"

    folder_id = Column(ColumnType.INTEGER, required=True, label="Folder")
    talk_date = Column(ColumnType.STRING, required=True, label="Talk Date")
    topic = Column(ColumnType.STRING, required=True, label="Topic")
    speaker_user_id = Column(ColumnType.INTEGER, label="Speaker")
    attendee_ids = Column(ColumnType.JSON, default=[], label="Attendees")
    notes = Column(ColumnType.TEXT, label="Notes")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.attendee_ids = _normalize_int_list(self.attendee_ids)

    def before_save(self):
        self.attendee_ids = _normalize_int_list(self.attendee_ids)

    def validate(self):
        super().validate()
        if not (self.topic or "").strip():
            raise ValidationError("Talk topic is required")

    def to_dict(self) -> Dict[str, Any]:
        speaker_name = None
        if self.speaker_user_id:
            try:
                from modules.base.module_base import User

                speaker = User.find_by_id(self.speaker_user_id)
                speaker_name = speaker.name if speaker else None
            except Exception:
                pass
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "talk_date": self.talk_date or "",
            "topic": self.topic or "",
            "speaker_user_id": self.speaker_user_id,
            "speaker_name": speaker_name,
            "attendee_ids": _normalize_int_list(self.attendee_ids),
            "attendance_count": len(_normalize_int_list(self.attendee_ids)),
            "notes": self.notes or "",
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyChecklistRun(BaseModel, AuditMixin):
    __tablename__ = "safety_checklists"
    __displayname__ = "checklist_name"

    folder_id = Column(ColumnType.INTEGER, required=True, label="Folder")
    checklist_name = Column(ColumnType.STRING, required=True, label="Name")
    checklist_type = Column(ColumnType.STRING, label="Type")
    executed_at = Column(ColumnType.STRING, required=True, label="Executed At")
    executed_by = Column(ColumnType.INTEGER, label="Executed By")
    result = Column(ColumnType.STRING, default="pending", label="Result")
    items = Column(ColumnType.JSON, default=[], label="Items")
    findings = Column(ColumnType.TEXT, label="Findings")
    requires_action = Column(ColumnType.BOOLEAN, default=False, label="Requires Action")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.items = _normalize_str_list(self.items)

    def before_save(self):
        self.items = _normalize_str_list(self.items)

    def validate(self):
        super().validate()
        if self.result not in CHECKLIST_RESULTS:
            raise ValidationError(
                f"Checklist result must be one of: {', '.join(CHECKLIST_RESULTS)}"
            )

    def to_dict(self) -> Dict[str, Any]:
        executed_by_name = None
        if self.executed_by:
            try:
                from modules.base.module_base import User

                user = User.find_by_id(self.executed_by)
                executed_by_name = user.name if user else None
            except Exception:
                pass
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "checklist_name": self.checklist_name or "",
            "checklist_type": self.checklist_type or "",
            "executed_at": self.executed_at or "",
            "executed_by": self.executed_by,
            "executed_by_name": executed_by_name,
            "result": self.result or "pending",
            "items": _normalize_str_list(self.items),
            "findings": self.findings or "",
            "requires_action": bool(self.requires_action),
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyIRLRecord(BaseModel, AuditMixin):
    __tablename__ = "safety_irl_records"
    __displayname__ = "title"

    folder_id = Column(ColumnType.INTEGER, required=True, label="Folder")
    employee_id = Column(ColumnType.INTEGER, label="Employee")
    title = Column(ColumnType.STRING, required=True, label="Title")
    status = Column(ColumnType.STRING, default="draft", label="Status")
    version = Column(ColumnType.INTEGER, default=1, label="Version")
    worker_name = Column(ColumnType.STRING, label="Worker Name")
    worker_identifier = Column(ColumnType.STRING, label="Worker Identifier")
    position_title = Column(ColumnType.STRING, label="Position")
    place_name = Column(ColumnType.STRING, label="Place")
    activity_name = Column(ColumnType.STRING, label="Activity")
    activity_period = Column(ColumnType.STRING, label="Activity Period")
    modality = Column(ColumnType.STRING, default="Presencial", label="Modality")
    duration_hours = Column(ColumnType.STRING, default="08:00", label="Duration")
    executor_name = Column(ColumnType.STRING, label="Executor")
    relator_background = Column(ColumnType.STRING, label="Relator Background")
    target_group = Column(ColumnType.TEXT, label="Target Group")
    workspace_features = Column(ColumnType.TEXT, label="Workspace Features")
    environmental_conditions = Column(ColumnType.TEXT, label="Environmental Conditions")
    order_cleanliness = Column(ColumnType.TEXT, label="Order and Cleanliness")
    machines_tools = Column(ColumnType.TEXT, label="Machines and Tools")
    service_functions = Column(ColumnType.JSON, default=[], label="Service Functions")
    risk_items = Column(ColumnType.JSON, default=[], label="Risk Items")
    complement_materials = Column(ColumnType.JSON, default=[], label="Complement Materials")
    observations = Column(ColumnType.TEXT, label="Observations")
    intro_text = Column(ColumnType.TEXT, label="Intro Text")
    theme_color = Column(ColumnType.STRING, default="#0F4C81", label="Theme Color")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.service_functions = _normalize_str_list(self.service_functions)
        self.risk_items = _normalize_irl_risk_items(self.risk_items)
        self.complement_materials = _normalize_irl_materials(self.complement_materials)

    def before_save(self):
        self.service_functions = _normalize_str_list(self.service_functions)
        self.risk_items = _normalize_irl_risk_items(self.risk_items)
        self.complement_materials = _normalize_irl_materials(self.complement_materials)

    def validate(self):
        super().validate()
        if not (self.title or "").strip():
            raise ValidationError("IRL title is required")
        if self.status not in IRL_STATUSES:
            raise ValidationError(f"IRL status must be one of: {', '.join(IRL_STATUSES)}")

    def to_dict(self) -> Dict[str, Any]:
        employee_name = None
        employee_position = None
        if self.employee_id:
            try:
                from modules.hr.module_hr import EmployeeProfile

                employee = EmployeeProfile.find_by_id(self.employee_id)
                if employee:
                    employee_name = employee.full_name
                    employee_position = employee.position_title
            except Exception:
                pass
        return {
            "id": self.id,
            "folder_id": self.folder_id,
            "employee_id": self.employee_id,
            "employee_name": employee_name or self.worker_name or "",
            "employee_position": employee_position or self.position_title or "",
            "title": self.title or "",
            "status": self.status or "draft",
            "version": self.version or 1,
            "worker_name": self.worker_name or employee_name or "",
            "worker_identifier": self.worker_identifier or "",
            "position_title": self.position_title or employee_position or "",
            "place_name": self.place_name or "",
            "activity_name": self.activity_name or "",
            "activity_period": self.activity_period or "",
            "modality": self.modality or "Presencial",
            "duration_hours": self.duration_hours or "08:00",
            "executor_name": self.executor_name or "",
            "relator_background": self.relator_background or "",
            "target_group": self.target_group or "",
            "workspace_features": self.workspace_features or "",
            "environmental_conditions": self.environmental_conditions or "",
            "order_cleanliness": self.order_cleanliness or "",
            "machines_tools": self.machines_tools or "",
            "service_functions": _normalize_str_list(self.service_functions),
            "risk_items": _normalize_irl_risk_items(self.risk_items),
            "complement_materials": _normalize_irl_materials(self.complement_materials),
            "observations": self.observations or "",
            "intro_text": self.intro_text or "",
            "theme_color": self.theme_color or "#0F4C81",
            "company_id": self.company_id,
            "created_at": _fmt_dt(self._data.get("created_at")),
            "updated_at": _fmt_dt(self._data.get("updated_at")),
        }


class SafetyMasterRisk(BaseModel, AuditMixin):
    __tablename__ = "safety_master_risks"
    __displayname__ = "risk_name"

    isp_code = Column(ColumnType.STRING, required=True, label="ISP Code")
    family = Column(ColumnType.STRING, required=True, label="Family")
    risk_name = Column(ColumnType.STRING, required=True, label="Risk Name")
    official_definition = Column(ColumnType.TEXT, label="Definition")
    protocol_codes = Column(ColumnType.JSON, default=[], label="Protocols")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.protocol_codes = _normalize_str_list(self.protocol_codes)

    def before_save(self):
        self.protocol_codes = _normalize_str_list(self.protocol_codes)

    def validate(self):
        super().validate()
        if not (self.isp_code or "").strip():
            raise ValidationError("ISP code is required")
        if not (self.risk_name or "").strip():
            raise ValidationError("Master risk name is required")

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "isp_code": self.isp_code or "",
            "family": self.family or "",
            "risk_name": self.risk_name or "",
            "official_definition": self.official_definition or "",
            "protocol_codes": _normalize_str_list(self.protocol_codes),
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyProtocol(BaseModel, AuditMixin):
    __tablename__ = "safety_protocols"
    __displayname__ = "name"

    code = Column(ColumnType.STRING, required=True, label="Code")
    name = Column(ColumnType.STRING, required=True, label="Name")
    authority = Column(ColumnType.STRING, label="Authority")
    description = Column(ColumnType.TEXT, label="Description")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def validate(self):
        super().validate()
        if not (self.code or "").strip():
            raise ValidationError("Protocol code is required")
        if not (self.name or "").strip():
            raise ValidationError("Protocol name is required")

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "code": self.code or "",
            "name": self.name or "",
            "authority": self.authority or "",
            "description": self.description or "",
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyPPEItem(BaseModel, AuditMixin):
    __tablename__ = "safety_ppe_catalog"
    __displayname__ = "name"

    code = Column(ColumnType.STRING, required=True, label="Code")
    name = Column(ColumnType.STRING, required=True, label="Name")
    category = Column(ColumnType.STRING, label="Category")
    description = Column(ColumnType.TEXT, label="Description")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.code = str(self.code or "").strip().upper()

    def before_save(self):
        self.code = str(self.code or "").strip().upper()

    def validate(self):
        super().validate()
        if not (self.code or "").strip():
            raise ValidationError("PPE code is required")
        if not (self.name or "").strip():
            raise ValidationError("PPE name is required")

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "code": self.code or "",
            "name": self.name or "",
            "category": self.category or "",
            "description": self.description or "",
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyEquipmentBlock(BaseModel, AuditMixin):
    __tablename__ = "safety_equipment_blocks"
    __displayname__ = "name"

    code = Column(ColumnType.STRING, required=True, label="Code")
    name = Column(ColumnType.STRING, required=True, label="Name")
    description = Column(ColumnType.TEXT, label="Description")
    master_risk_ids = Column(ColumnType.JSON, default=[], label="Master Risks")
    control_hierarchy = Column(ColumnType.JSON, default={}, label="Control Hierarchy")
    controls_summary = Column(ColumnType.TEXT, label="Controls Summary")
    required_ppe = Column(ColumnType.JSON, default=[], label="Required PPE")
    protocol_codes = Column(ColumnType.JSON, default=[], label="Protocols")
    sensitivity_tags = Column(ColumnType.JSON, default=[], label="Sensitivity Tags")
    legal_reference = Column(ColumnType.STRING, label="Legal Reference")
    source_note = Column(ColumnType.TEXT, label="Source Note")
    probability = Column(ColumnType.INTEGER, default=2, label="Probability")
    consequence = Column(ColumnType.INTEGER, default=2, label="Consequence")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.code = str(self.code or "").strip().upper()
        self.master_risk_ids = _normalize_int_list(self.master_risk_ids)
        self.control_hierarchy = normalize_control_hierarchy(self.control_hierarchy)
        self.required_ppe = _normalize_str_list(self.required_ppe)
        self.protocol_codes = _normalize_str_list(self.protocol_codes)
        self.sensitivity_tags = _normalize_str_list(self.sensitivity_tags)

    def before_save(self):
        self.before_create()
        self.controls_summary = self.controls_summary or summarize_controls(
            self.control_hierarchy,
            "",
        )

    def validate(self):
        super().validate()
        if not (self.code or "").strip():
            raise ValidationError("Equipment code is required")
        if not (self.name or "").strip():
            raise ValidationError("Equipment name is required")

    def to_dict(self) -> Dict[str, Any]:
        risk_map = {
            risk.id: risk
            for risk in SafetyMasterRisk.search([("company_id", "=", self.company_id)])
            if risk.id and risk.active
        }
        probability = _safe_int(self.probability, 2) or 2
        consequence = _safe_int(self.consequence, 2) or 2
        vep = calculate_vep(probability, consequence)
        master_risks = [
            {
                "id": risk.id,
                "isp_code": risk.isp_code or "",
                "risk_name": risk.risk_name or "",
                "family": risk.family or "",
            }
            for risk_id in _normalize_int_list(self.master_risk_ids)
            for risk in [risk_map.get(risk_id)]
            if risk
        ]
        return {
            "id": self.id,
            "code": self.code or "",
            "name": self.name or "",
            "description": self.description or "",
            "master_risk_ids": _normalize_int_list(self.master_risk_ids),
            "master_risks": master_risks,
            "control_hierarchy": normalize_control_hierarchy(self.control_hierarchy),
            "controls_summary": self.controls_summary or summarize_controls(self.control_hierarchy, ""),
            "required_ppe": _normalize_str_list(self.required_ppe),
            "protocol_codes": _normalize_str_list(self.protocol_codes),
            "sensitivity_tags": _normalize_str_list(self.sensitivity_tags),
            "legal_reference": self.legal_reference or "",
            "source_note": self.source_note or "",
            "probability": probability,
            "consequence": consequence,
            "vep": vep,
            "risk_level": risk_level_from_vep(vep),
            "action_required": action_from_vep(vep),
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyClientSite(BaseModel, AuditMixin):
    __tablename__ = "safety_client_sites"
    __displayname__ = "name"

    customer_id = Column(ColumnType.INTEGER, required=True, label="Customer")
    name = Column(ColumnType.STRING, required=True, label="Name")
    address = Column(ColumnType.TEXT, label="Address")
    comuna = Column(ColumnType.STRING, label="Comuna")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def validate(self):
        super().validate()
        if not self.customer_id:
            raise ValidationError("customer_id is required")
        if not (self.name or "").strip():
            raise ValidationError("Site name is required")

    def to_dict(self) -> Dict[str, Any]:
        customer_name = None
        try:
            from modules.crm.module_crm import Customer

            customer = Customer.find_by_id(self.customer_id)
            customer_name = customer.name if customer else None
        except Exception:
            customer_name = None
        return {
            "id": self.id,
            "customer_id": self.customer_id,
            "customer_name": customer_name,
            "name": self.name or "",
            "address": self.address or "",
            "comuna": self.comuna or "",
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyClientArea(BaseModel, AuditMixin):
    __tablename__ = "safety_client_areas"
    __displayname__ = "name"

    site_id = Column(ColumnType.INTEGER, required=True, label="Site")
    parent_area_id = Column(ColumnType.INTEGER, label="Parent Area")
    name = Column(ColumnType.STRING, required=True, label="Name")
    risk_notes = Column(ColumnType.TEXT, label="Risk Notes")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def validate(self):
        super().validate()
        if not self.site_id:
            raise ValidationError("site_id is required")
        if not (self.name or "").strip():
            raise ValidationError("Area name is required")

    def to_dict(self) -> Dict[str, Any]:
        site = SafetyClientSite.find_by_id(self.site_id) if self.site_id else None
        return {
            "id": self.id,
            "site_id": self.site_id,
            "site_name": site.name if site else None,
            "customer_id": site.customer_id if site else None,
            "name": self.name or "",
            "parent_area_id": self.parent_area_id,
            "risk_notes": self.risk_notes or "",
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyWorkerRestriction(BaseModel, AuditMixin):
    __tablename__ = "safety_worker_restrictions"
    __displayname__ = "title"

    employee_id = Column(ColumnType.INTEGER, required=True, label="Employee")
    title = Column(ColumnType.STRING, required=True, label="Title")
    restriction_type = Column(ColumnType.STRING, label="Restriction Type")
    applies_to_tags = Column(ColumnType.JSON, default=[], label="Tags")
    severity = Column(ColumnType.STRING, default="warning", label="Severity")
    details = Column(ColumnType.TEXT, label="Details")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.applies_to_tags = _normalize_str_list(self.applies_to_tags)

    def before_save(self):
        self.applies_to_tags = _normalize_str_list(self.applies_to_tags)

    def validate(self):
        super().validate()
        if not self.employee_id:
            raise ValidationError("employee_id is required")
        if not (self.title or "").strip():
            raise ValidationError("Restriction title is required")
        if self.severity not in RESTRICTION_SEVERITIES:
            raise ValidationError(
                f"Restriction severity must be one of: {', '.join(RESTRICTION_SEVERITIES)}"
            )

    def to_dict(self) -> Dict[str, Any]:
        employee_name = None
        try:
            from modules.hr.module_hr import EmployeeProfile

            employee = EmployeeProfile.find_by_id(self.employee_id)
            employee_name = employee.full_name if employee else None
        except Exception:
            employee_name = None
        return {
            "id": self.id,
            "employee_id": self.employee_id,
            "employee_name": employee_name,
            "title": self.title or "",
            "restriction_type": self.restriction_type or "",
            "applies_to_tags": _normalize_str_list(self.applies_to_tags),
            "severity": self.severity or "warning",
            "details": self.details or "",
            "active": bool(self.active),
            "company_id": self.company_id,
        }


class SafetyGeneratorRule(BaseModel, AuditMixin):
    __tablename__ = "safety_generator_rules"
    __displayname__ = "name"

    name = Column(ColumnType.STRING, required=True, label="Name")
    scope_type = Column(ColumnType.STRING, required=True, label="Scope Type")
    scope_ref_id = Column(ColumnType.INTEGER, label="Scope Ref")
    process_name = Column(ColumnType.STRING, label="Process")
    task_name = Column(ColumnType.STRING, label="Task")
    position_name = Column(ColumnType.STRING, label="Position")
    hazard_factor = Column(ColumnType.STRING, required=True, label="Hazard")
    master_risk_id = Column(ColumnType.INTEGER, required=True, label="Master Risk")
    probability = Column(ColumnType.INTEGER, default=2, label="Probability")
    consequence = Column(ColumnType.INTEGER, default=2, label="Consequence")
    controls_summary = Column(ColumnType.TEXT, label="Controls")
    control_hierarchy = Column(ColumnType.JSON, default={}, label="Control Hierarchy")
    required_ppe = Column(ColumnType.JSON, default=[], label="Required PPE")
    protocol_codes = Column(ColumnType.JSON, default=[], label="Protocols")
    sensitivity_tags = Column(ColumnType.JSON, default=[], label="Sensitivity Tags")
    owner_name = Column(ColumnType.STRING, label="Owner")
    legal_reference = Column(ColumnType.STRING, label="Legal Reference")
    source_note = Column(ColumnType.TEXT, label="Source Note")
    trigger_config = Column(ColumnType.JSON, default={}, label="Trigger Config")
    suggested_controls = Column(ColumnType.JSON, default={}, label="Suggested Controls")
    suggested_ppe = Column(ColumnType.JSON, default=[], label="Suggested PPE")
    suggested_protocols = Column(ColumnType.JSON, default=[], label="Suggested Protocols")
    required_fields = Column(ColumnType.JSON, default=[], label="Required Fields")
    approval_policy = Column(ColumnType.JSON, default={}, label="Approval Policy")
    active = Column(ColumnType.BOOLEAN, default=True, label="Active")
    company_id = Column(ColumnType.INTEGER, required=True, label="Company")

    def before_create(self):
        self.required_ppe = _normalize_str_list(self.required_ppe)
        self.protocol_codes = _normalize_str_list(self.protocol_codes)
        self.sensitivity_tags = _normalize_str_list(self.sensitivity_tags)
        self.control_hierarchy = normalize_control_hierarchy(self.control_hierarchy)
        self.trigger_config = self.trigger_config if isinstance(self.trigger_config, dict) else {}
        self.suggested_controls = normalize_control_hierarchy(self.suggested_controls or self.control_hierarchy)
        self.suggested_ppe = _normalize_str_list(self.suggested_ppe or self.required_ppe)
        self.suggested_protocols = _normalize_str_list(self.suggested_protocols or self.protocol_codes)
        self.required_fields = _normalize_str_list(self.required_fields)
        self.approval_policy = self.approval_policy if isinstance(self.approval_policy, dict) else {}

    def before_save(self):
        self.required_ppe = _normalize_str_list(self.required_ppe)
        self.protocol_codes = _normalize_str_list(self.protocol_codes)
        self.sensitivity_tags = _normalize_str_list(self.sensitivity_tags)
        self.control_hierarchy = normalize_control_hierarchy(self.control_hierarchy)
        self.controls_summary = (self.controls_summary or "").strip() or summarize_controls(
            self.control_hierarchy, ""
        )

    def validate(self):
        super().validate()
        if not (self.name or "").strip():
            raise ValidationError("Generator rule name is required")
        if self.scope_type not in RULE_SCOPE_TYPES:
            raise ValidationError(
                f"Generator rule scope must be one of: {', '.join(RULE_SCOPE_TYPES)}"
            )
        if not (self.hazard_factor or "").strip():
            raise ValidationError("hazard_factor is required")
        if not self.master_risk_id:
            raise ValidationError("master_risk_id is required")
        if _safe_int(self.probability, 0) not in (1, 2, 4):
            raise ValidationError("probability must be one of 1, 2 or 4")
        if _safe_int(self.consequence, 0) not in (1, 2, 4):
            raise ValidationError("consequence must be one of 1, 2 or 4")

    def to_dict(self) -> Dict[str, Any]:
        risk = SafetyMasterRisk.find_by_id(self.master_risk_id) if self.master_risk_id else None
        return {
            "id": self.id,
            "name": self.name or "",
            "scope_type": self.scope_type or "",
            "scope_ref_id": self.scope_ref_id,
            "process_name": self.process_name or "",
            "task_name": self.task_name or "",
            "position_name": self.position_name or "",
            "hazard_factor": self.hazard_factor or "",
            "master_risk_id": self.master_risk_id,
            "master_risk_code": risk.isp_code if risk else None,
            "master_risk_name": risk.risk_name if risk else None,
            "risk_family": risk.family if risk else None,
            "probability": _safe_int(self.probability, 2) or 2,
            "consequence": _safe_int(self.consequence, 2) or 2,
            "controls_summary": self.controls_summary or summarize_controls(
                normalize_control_hierarchy(self.control_hierarchy), ""
            ),
            "control_hierarchy": normalize_control_hierarchy(self.control_hierarchy),
            "required_ppe": _normalize_str_list(self.required_ppe),
            "protocol_codes": _normalize_str_list(self.protocol_codes),
            "sensitivity_tags": _normalize_str_list(self.sensitivity_tags),
            "owner_name": self.owner_name or "",
            "legal_reference": self.legal_reference or "",
            "source_note": self.source_note or "",
            "trigger_config": self.trigger_config if isinstance(self.trigger_config, dict) else {},
            "suggested_controls": normalize_control_hierarchy(self.suggested_controls or self.control_hierarchy),
            "suggested_ppe": _normalize_str_list(self.suggested_ppe or self.required_ppe),
            "suggested_protocols": _normalize_str_list(self.suggested_protocols or self.protocol_codes),
            "required_fields": _normalize_str_list(self.required_fields),
            "approval_policy": self.approval_policy if isinstance(self.approval_policy, dict) else {},
            "active": bool(self.active),
            "company_id": self.company_id,
        }


def seed_default_profiles(company_id: int) -> List[SafetyServiceProfile]:
    existing = SafetyServiceProfile.search([("company_id", "=", company_id)])
    if existing:
        return existing
    created: List[SafetyServiceProfile] = []
    for payload in _default_profile_payloads():
        created.append(
            SafetyServiceProfile.create(
                {
                    **payload,
                    "service_type_id": None,
                    "active": True,
                    "company_id": company_id,
                }
            )
        )
    return created


def seed_default_miper_catalog(company_id: int) -> Dict[str, int]:
    protocols = SafetyProtocol.search([("company_id", "=", company_id)])
    existing_protocol_codes = {str(item.code or "").strip().upper() for item in protocols}
    protocols_created = 0
    for payload in DEFAULT_PROTOCOLS:
        code = str(payload.get("code") or "").strip().upper()
        if not code or code in existing_protocol_codes:
            continue
        SafetyProtocol.create(
            {
                "code": code,
                "name": payload.get("name") or code,
                "authority": payload.get("authority") or "MINSAL",
                "description": payload.get("description") or "",
                "active": True,
                "company_id": company_id,
            }
        )
        existing_protocol_codes.add(code)
        protocols_created += 1

    master_risks = SafetyMasterRisk.search([("company_id", "=", company_id)])
    existing_risk_codes = {str(item.isp_code or "").strip().upper() for item in master_risks}
    risks_created = 0
    for payload in DEFAULT_MASTER_RISKS:
        code = str(payload.get("isp_code") or "").strip().upper()
        if not code or code in existing_risk_codes:
            continue
        SafetyMasterRisk.create(
            {
                "isp_code": code,
                "family": payload.get("family") or "OTROS",
                "risk_name": payload.get("risk_name") or code,
                "official_definition": payload.get("official_definition") or "",
                "protocol_codes": payload.get("protocol_codes") or [],
                "active": True,
                "company_id": company_id,
            }
        )
        existing_risk_codes.add(code)
        risks_created += 1

    profiles = seed_default_profiles(company_id)
    profile_map = {profile.name: profile.id for profile in profiles if profile.id}
    risk_map = {
        str(risk.isp_code or "").strip().upper(): risk.id
        for risk in SafetyMasterRisk.search([("company_id", "=", company_id)])
        if risk.id
    }
    existing_rule_names = {
        str(rule.name or "").strip().lower()
        for rule in SafetyGeneratorRule.search([("company_id", "=", company_id)])
    }
    rules_created = 0
    for payload in rule_blueprints(profile_map):
        rule_name = str(payload.get("name") or "").strip()
        if not rule_name or rule_name.lower() in existing_rule_names:
            continue
        risk_id = risk_map.get(str(payload.get("master_risk_code") or "").strip().upper())
        if not risk_id:
            continue
        SafetyGeneratorRule.create(
            {
                "name": rule_name,
                "scope_type": payload.get("scope_type") or "transversal",
                "scope_ref_id": payload.get("scope_ref_id"),
                "process_name": payload.get("process_name") or "",
                "task_name": payload.get("task_name") or "",
                "position_name": payload.get("position_name") or "",
                "hazard_factor": payload.get("hazard_factor") or "",
                "master_risk_id": risk_id,
                "probability": _safe_int(payload.get("probability"), 2) or 2,
                "consequence": _safe_int(payload.get("consequence"), 2) or 2,
                "controls_summary": summarize_controls(
                    normalize_control_hierarchy(payload.get("control_hierarchy")),
                    "",
                ),
                "control_hierarchy": payload.get("control_hierarchy") or {},
                "required_ppe": payload.get("required_ppe") or [],
                "protocol_codes": payload.get("protocol_codes") or [],
                "sensitivity_tags": payload.get("sensitivity_tags") or [],
                "owner_name": payload.get("owner_name") or "",
                "legal_reference": payload.get("legal_reference") or "",
                "source_note": payload.get("source_note") or "",
                "trigger_config": payload.get("trigger_config") or {"scope_type": payload.get("scope_type") or "transversal"},
                "suggested_controls": payload.get("control_hierarchy") or {},
                "suggested_ppe": payload.get("required_ppe") or [],
                "suggested_protocols": payload.get("protocol_codes") or [],
                "required_fields": payload.get("required_fields") or [],
                "approval_policy": payload.get("approval_policy") or {},
                "active": True,
                "company_id": company_id,
            }
        )
        existing_rule_names.add(rule_name.lower())
        rules_created += 1

    smart_rules = [
        {
            "name": "BOT tag altura - EPP y permiso obligatorio",
            "risk_code": "A3",
            "hazard_factor": "Trabajo en altura detectado por etiqueta",
            "trigger_config": {"tags_any": ["altura"], "entity": "activity_block"},
            "controls": {
                "engineering": ["Protecciones colectivas, puntos de anclaje o linea de vida certificados"],
                "administrative": ["Permiso de trabajo en altura", "Plan de rescate", "Checklist SPDC"],
                "ppe": ["Arnes de seguridad", "Casco con barbiquejo"],
            },
            "ppe": ["Arnes de seguridad", "Casco con barbiquejo"],
            "protocols": [],
            "required_fields": ["plan_rescate", "permiso_altura", "punto_anclaje"],
            "approval_policy": {"criticality": "critical", "block_if_risk_level": "Intolerable"},
        },
        {
            "name": "Herramienta electrica - verificacion preventiva",
            "risk_code": "F1",
            "hazard_factor": "Uso de herramienta electrica",
            "trigger_config": {"tags_any": ["herramienta_electrica"], "resource_types_any": ["tool", "equipment"]},
            "controls": {
                "engineering": ["Protecciones, enchufes y aislacion en buen estado"],
                "administrative": ["Inspeccion preuso", "Verificacion electrica", "Retiro de equipos defectuosos"],
                "ppe": ["Guantes segun contexto", "Lentes de seguridad"],
            },
            "ppe": ["Guantes", "Lentes de seguridad"],
            "protocols": [],
            "required_fields": ["checklist_preuso"],
            "approval_policy": {"require_control_if_level_at_least": "Importante"},
        },
        {
            "name": "Andamios - riesgos criticos base",
            "risk_code": "A3",
            "hazard_factor": "Montaje, uso o modificacion de andamios",
            "trigger_config": {"tags_any": ["andamios"], "process_contains_any": ["andamio", "andamios"]},
            "controls": {
                "engineering": ["Estructura certificada, nivelada, arriostrada y con barandas completas"],
                "administrative": ["Armado por personal competente", "Tarjeta de andamio", "Inspeccion y habilitacion"],
                "ppe": ["Arnes de seguridad", "Casco con barbiquejo", "Guantes"],
            },
            "ppe": ["Arnes de seguridad", "Casco con barbiquejo", "Guantes"],
            "protocols": [],
            "required_fields": ["inspeccion_andamio", "tarjeta_andamio"],
            "approval_policy": {"criticality": "critical", "block_if_risk_level": "Intolerable"},
        },
        {
            "name": "BOT critico - aprobacion adicional",
            "risk_code": "N",
            "hazard_factor": "Bloque critico requiere aprobacion preventiva",
            "trigger_config": {"criticality_any": ["critical"], "entity": "activity_block"},
            "controls": {
                "administrative": ["Revision de supervisor y prevencion antes de aprobar", "Justificacion de controles criticos"],
            },
            "ppe": [],
            "protocols": [],
            "required_fields": ["aprobador_preventivo", "justificacion_controles"],
            "approval_policy": {"additional_approval_required": True},
        },
    ]
    for payload in smart_rules:
        rule_name = payload["name"]
        if rule_name.lower() in existing_rule_names:
            continue
        risk_id = risk_map.get(payload["risk_code"])
        if not risk_id:
            continue
        controls = normalize_control_hierarchy(payload.get("controls"))
        SafetyGeneratorRule.create(
            {
                "name": rule_name,
                "scope_type": "transversal",
                "scope_ref_id": None,
                "process_name": "Motor preventivo BOT",
                "task_name": rule_name,
                "position_name": "Supervisor / Prevencion",
                "hazard_factor": payload["hazard_factor"],
                "master_risk_id": risk_id,
                "probability": 2,
                "consequence": 4 if payload["risk_code"] == "A3" else 2,
                "controls_summary": summarize_controls(controls, ""),
                "control_hierarchy": controls,
                "required_ppe": payload.get("ppe") or [],
                "protocol_codes": payload.get("protocols") or [],
                "sensitivity_tags": payload.get("trigger_config", {}).get("tags_any") or [],
                "owner_name": "Supervisor de terreno",
                "legal_reference": "MIPER/IPER y jerarquia de controles",
                "source_note": "Regla parametrizable del motor preventivo BOT",
                "trigger_config": payload.get("trigger_config") or {},
                "suggested_controls": controls,
                "suggested_ppe": payload.get("ppe") or [],
                "suggested_protocols": payload.get("protocols") or [],
                "required_fields": payload.get("required_fields") or [],
                "approval_policy": payload.get("approval_policy") or {},
                "active": True,
                "company_id": company_id,
            }
        )
        existing_rule_names.add(rule_name.lower())
        rules_created += 1

    return {
        "protocols_created": protocols_created,
        "master_risks_created": risks_created,
        "rules_created": rules_created,
    }


def seed_default_ppe_catalog(company_id: Optional[int]) -> Dict[str, int]:
    if not company_id:
        return {"ppe_created": 0}
    existing = {
        str(item.code or "").strip().upper(): item
        for item in SafetyPPEItem.search([("company_id", "=", company_id)])
    }
    created = 0
    for payload in DEFAULT_PPE_CATALOG:
        code = str(payload.get("code") or "").strip().upper()
        if not code or code in existing:
            continue
        SafetyPPEItem.create(
            {
                "code": code,
                "name": payload.get("name") or code,
                "category": payload.get("category") or "",
                "description": payload.get("description") or "",
                "active": True,
                "company_id": company_id,
            }
        )
        existing[code] = True
        created += 1
    return {"ppe_created": created}


def seed_default_risk_methodology(company_id: Optional[int]) -> Dict[str, int]:
    if not company_id:
        return {"methodologies_created": 0}
    existing = [
        item
        for item in SafetyRiskMethodology.search([("company_id", "=", company_id)])
        if item.active
    ]
    if existing:
        return {"methodologies_created": 0}
    payload = methodology_payload()
    SafetyRiskMethodology.create(
        {
            "company_id": company_id,
            "name": payload.get("name") or "MIPER/IPER 1-2-4",
            "probability_schema_json": payload.get("probability_schema") or [],
            "consequence_schema_json": payload.get("consequence_schema") or [],
            "risk_matrix_schema_json": payload.get("risk_matrix_schema") or [],
            "default_flag": True,
            "active": True,
        }
    )
    return {"methodologies_created": 1}


def _matrix_row_payload_from_legacy(
    matrix: SafetyRiskMatrix,
    row: Dict[str, Any],
    sequence: int,
) -> Dict[str, Any]:
    controls = normalize_control_hierarchy(row.get("control_hierarchy"))
    routine_type = row.get("routine_type") or "routine"
    probability = _safe_int(row.get("probability_value") or row.get("probability"), 2) or 2
    consequence = _safe_int(row.get("consequence_value") or row.get("consequence"), 2) or 2
    compact = _compact_miper_from_item(
        row,
        routine_type=routine_type,
        worker_count=_safe_int(row.get("worker_count"), 1) or 1,
        consequence=consequence,
    )
    current_engineering = _current_controls_text(
        row,
        "current_engineering_controls",
        controls.get("engineering"),
    )
    current_admin = _current_controls_text(
        row,
        "current_admin_controls",
        controls.get("administrative"),
    )
    current_ppe = _current_controls_text(
        row,
        "current_ppe_controls",
        row.get("required_ppe") or row.get("ppe_summary") or controls.get("ppe"),
    )
    return {
        "company_id": matrix.company_id,
        "risk_matrix_id": matrix.id,
        "source_block_id": _safe_int(row.get("source_block_id"), None),
        "source_procedure_step_id": _safe_int(
            row.get("source_procedure_step_id") or row.get("source_procedure_block_id"),
            None,
        ),
        "activity_name": row.get("activity") or row.get("task_name") or "",
        "task_name": row.get("task_name") or row.get("activity") or "",
        "job_position": row.get("position_name") or row.get("job_position") or "",
        "specific_workplace": row.get("place_name") or row.get("specific_workplace") or "",
        "worker_count": _safe_int(row.get("worker_count"), 1) or 1,
        "routine_type": routine_type,
        "hazard_name": row.get("hazard") or row.get("hazard_factor") or row.get("hazard_name") or "",
        "risk_name": row.get("risk") or row.get("risk_name") or "",
        "probable_damage": row.get("probable_damage") or "",
        "probability_value": probability,
        "consequence_value": consequence,
        "task_type_code": compact["task_type_code"],
        "exposed_people_value": compact["exposed_people_value"],
        "exposure_frequency_value": compact["exposure_frequency_value"],
        "occurrence_factor_value": compact["occurrence_factor_value"],
        "probability_score": compact["probability_score"],
        "severity_value": compact["severity_value"],
        "residual_risk_value": compact["residual_risk_value"],
        "residual_risk_label": compact["residual_risk_label"],
        "current_engineering_controls": current_engineering,
        "current_admin_controls": current_admin,
        "current_ppe_controls": current_ppe,
        "proposed_elimination_controls": row.get("proposed_elimination_controls") or controls.get("elimination") or [],
        "proposed_substitution_controls": row.get("proposed_substitution_controls") or controls.get("substitution") or [],
        "proposed_engineering_controls": row.get("proposed_engineering_controls") or controls.get("engineering") or [],
        "proposed_admin_controls": row.get("proposed_admin_controls") or controls.get("administrative") or [],
        "proposed_ppe_controls": row.get("proposed_ppe_controls") or controls.get("ppe") or [],
        "safety_management_plan": row.get("safety_management_plan") or row.get("management_plan") or "",
        "existing_controls": row.get("controls") or row.get("existing_controls") or summarize_controls(controls, ""),
        "required_controls": controls,
        "ppe_summary": row.get("required_ppe") or row.get("ppe_summary") or [],
        "protocols_summary": row.get("protocol_codes") or row.get("protocols_summary") or [],
        "legal_reference": row.get("legal_reference") or "",
        "responsible": row.get("owner_name") or row.get("responsible") or "",
        "observations": row.get("observations") or row.get("source_note") or "",
        "sequence": sequence,
        "origin_blocks": row.get("origin_blocks") or [],
        "source_labels": row.get("source_labels") or [],
        "source_group": row.get("source_group") or "",
        "source_title": row.get("source_title") or "",
        "source_ref_id": _safe_int(row.get("source_id"), None),
        "active": True,
    }


def backfill_risk_matrix_rows(company_id: Optional[int]) -> Dict[str, int]:
    if not company_id:
        return {"matrix_rows_created": 0}
    created = 0
    matrices = SafetyRiskMatrix.search([("company_id", "=", company_id)])
    existing_by_matrix = {
        matrix.id: SafetyRiskMatrixRow.search([("risk_matrix_id", "=", matrix.id)])
        for matrix in matrices
        if matrix.id
    }
    for matrix in matrices:
        if existing_by_matrix.get(matrix.id):
            continue
        rows = _normalize_matrix_rows(matrix.rows)
        for index, row in enumerate(rows, start=1):
            SafetyRiskMatrixRow.create(_matrix_row_payload_from_legacy(matrix, row, index * 10))
            created += 1
    return {"matrix_rows_created": created}


class SafetyModule(BaseModule):
    name = "Safety"
    version = "1.0.0"
    author = "Your Company"
    description = "Prevention and safety folders linked to opportunities"
    depends = ["base", "crm", "hr"]

    def init_module(self):
        self.register_model("safety.service_profile", SafetyServiceProfile)
        self.register_model("safety.folder", SafetyFolder)
        self.register_model("safety.document", SafetyFolderDocument)
        self.register_model("safety.risk_matrix", SafetyRiskMatrix)
        self.register_model("safety.risk_matrix_row", SafetyRiskMatrixRow)
        self.register_model("safety.risk_methodology", SafetyRiskMethodology)
        self.register_model("safety.ppe_delivery", SafetyPPEDelivery)
        self.register_model("safety.talk", SafetyTalk)
        self.register_model("safety.checklist", SafetyChecklistRun)
        self.register_model("safety.irl_record", SafetyIRLRecord)
        self.register_model("safety.master_risk", SafetyMasterRisk)
        self.register_model("safety.protocol", SafetyProtocol)
        self.register_model("safety.ppe_catalog", SafetyPPEItem)
        self.register_model("safety.equipment_block", SafetyEquipmentBlock)
        self.register_model("safety.client_site", SafetyClientSite)
        self.register_model("safety.client_area", SafetyClientArea)
        self.register_model("safety.worker_restriction", SafetyWorkerRestriction)
        self.register_model("safety.generator_rule", SafetyGeneratorRule)

        self.register_route("/safety/lookups", self.get_lookups, methods=["GET"], auth_required=True)
        self.register_route("/safety/risk-matrices", self.list_risk_matrices, methods=["GET"], auth_required=True)
        self.register_route("/safety/risk-matrices", self.create_risk_matrix, methods=["POST"], auth_required=True)
        self.register_route("/safety/risk-matrices/generate", self.generate_canonical_risk_matrix, methods=["POST"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}", self.get_risk_matrix, methods=["GET"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}", self.update_risk_matrix, methods=["PUT"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}/export/miper.xlsx", self.export_risk_matrix_excel, methods=["GET"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}/export/miper.pdf", self.export_risk_matrix_pdf, methods=["GET"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}/rows", self.create_risk_matrix_row, methods=["POST"], auth_required=True)
        self.register_route("/safety/risk-matrices/{id}/rows/{row_id}", self.update_risk_matrix_row, methods=["PUT"], auth_required=True)
        self.register_route("/safety/master-risks", self.list_master_risks, methods=["GET"], auth_required=True)
        self.register_route("/safety/ppe-catalog", self.list_ppe_catalog, methods=["GET"], auth_required=True)
        self.register_route("/safety/ppe-catalog", self.create_ppe_catalog_item, methods=["POST"], auth_required=True)
        self.register_route("/safety/ppe-catalog/{id}", self.update_ppe_catalog_item, methods=["PUT"], auth_required=True)
        self.register_route("/safety/ppe-catalog/{id}", self.delete_ppe_catalog_item, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/equipment-blocks", self.list_equipment_blocks, methods=["GET"], auth_required=True)
        self.register_route("/safety/equipment-blocks", self.create_equipment_block, methods=["POST"], auth_required=True)
        self.register_route("/safety/equipment-blocks/{id}", self.update_equipment_block, methods=["PUT"], auth_required=True)
        self.register_route("/safety/equipment-blocks/{id}", self.delete_equipment_block, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/generator-rules", self.list_generator_rules, methods=["GET"], auth_required=True)
        self.register_route("/safety/generator-rules", self.create_generator_rule, methods=["POST"], auth_required=True)
        self.register_route("/safety/generator-rules/{id}", self.update_generator_rule, methods=["PUT"], auth_required=True)
        self.register_route("/safety/generator-rules/{id}", self.delete_generator_rule, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/bot-assistant/suggestions", self.bot_assistant_suggestions, methods=["POST"], auth_required=True)
        self.register_route("/safety/client-sites", self.list_client_sites, methods=["GET"], auth_required=True)
        self.register_route("/safety/client-sites", self.create_client_site, methods=["POST"], auth_required=True)
        self.register_route("/safety/client-sites/{id}", self.update_client_site, methods=["PUT"], auth_required=True)
        self.register_route("/safety/client-sites/{id}", self.delete_client_site, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/client-areas", self.list_client_areas, methods=["GET"], auth_required=True)
        self.register_route("/safety/client-areas", self.create_client_area, methods=["POST"], auth_required=True)
        self.register_route("/safety/client-areas/{id}", self.update_client_area, methods=["PUT"], auth_required=True)
        self.register_route("/safety/client-areas/{id}", self.delete_client_area, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/worker-restrictions", self.list_worker_restrictions, methods=["GET"], auth_required=True)
        self.register_route("/safety/worker-restrictions", self.create_worker_restriction, methods=["POST"], auth_required=True)

        self.register_route("/safety/service-profiles", self.list_service_profiles, methods=["GET"], auth_required=True)
        self.register_route("/safety/service-profiles", self.create_service_profile, methods=["POST"], auth_required=True)
        self.register_route("/safety/service-profiles/{id}", self.update_service_profile, methods=["PUT"], auth_required=True)
        self.register_route("/safety/service-profiles/{id}", self.delete_service_profile, methods=["DELETE"], auth_required=True)

        self.register_route("/safety/folders", self.list_folders, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders", self.create_folder, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/dossier", self.folder_dossier, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/generate-documents", self.generate_documents, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/generate-matrix", self.generate_matrix, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/export/miper.xlsx", self.export_matrix_excel, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/export/miper.pdf", self.export_matrix_pdf, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/documents", self.list_documents, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/documents", self.create_document, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/risk-matrix", self.upsert_risk_matrix, methods=["PUT"], auth_required=True)
        self.register_route("/safety/folders/{id}/irl-records", self.list_irl_records, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/irl-records/generate", self.generate_irl_record, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/ppe-deliveries", self.list_ppe_deliveries, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/ppe-deliveries", self.create_ppe_delivery, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/talks", self.list_talks, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/talks", self.create_talk, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}/checklists", self.list_checklists, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}/checklists", self.create_checklist, methods=["POST"], auth_required=True)
        self.register_route("/safety/folders/{id}", self.get_folder, methods=["GET"], auth_required=True)
        self.register_route("/safety/folders/{id}", self.update_folder, methods=["PUT"], auth_required=True)
        self.register_route("/safety/folders/{id}", self.delete_folder, methods=["DELETE"], auth_required=True)

        self.register_route("/safety/documents/{id}", self.update_document, methods=["PUT"], auth_required=True)
        self.register_route("/safety/documents/{id}", self.delete_document, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/irl-records/{id}", self.update_irl_record, methods=["PUT"], auth_required=True)
        self.register_route("/safety/irl-records/{id}", self.delete_irl_record, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/irl-records/{id}/export/pdf", self.export_irl_pdf, methods=["GET"], auth_required=True)
        self.register_route("/safety/ppe-deliveries/{id}", self.update_ppe_delivery, methods=["PUT"], auth_required=True)
        self.register_route("/safety/ppe-deliveries/{id}", self.delete_ppe_delivery, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/talks/{id}", self.update_talk, methods=["PUT"], auth_required=True)
        self.register_route("/safety/talks/{id}", self.delete_talk, methods=["DELETE"], auth_required=True)
        self.register_route("/safety/checklists/{id}", self.update_checklist, methods=["PUT"], auth_required=True)
        self.register_route("/safety/checklists/{id}", self.delete_checklist, methods=["DELETE"], auth_required=True)

        self.logger.info("Safety module initialized")

    def _company_id(self) -> Optional[int]:
        user = self.env.user
        return user.company_id if user else None

    def _tenant_filter(self) -> List[tuple]:
        user = self.env.user
        if user and user.role == "superadmin":
            return []
        return [("company_id", "=", self._company_id())]

    def _require_access(self) -> Optional[Response]:
        user = self.env.user
        if not user:
            return Response.unauthorized("Authentication required")
        if user.role in ("superadmin", "company_admin"):
            return None
        if "safety" not in (user.allowed_modules or []):
            return Response.forbidden("You do not have access to Safety")
        return None

    def _require_admin(self) -> Optional[Response]:
        err = self._require_access()
        if err:
            return err
        user = self.env.user
        if user.role not in ("superadmin", "company_admin"):
            return Response.forbidden("Only administrators can manage this resource")
        return None

    def _ensure_profiles(self):
        company_id = self._company_id()
        if company_id and not SafetyServiceProfile.search([("company_id", "=", company_id)]):
            seed_default_profiles(company_id)

    def _ensure_miper_catalog(self):
        company_id = self._company_id()
        if company_id:
            seed_default_miper_catalog(company_id)
            seed_default_ppe_catalog(company_id)
            seed_default_risk_methodology(company_id)
            backfill_risk_matrix_rows(company_id)

    def _profile_or_404(
        self, profile_id: Any
    ) -> Tuple[Optional[SafetyServiceProfile], Optional[Response]]:
        profile = SafetyServiceProfile.find_by_id(_safe_int(profile_id))
        if not profile or (
            self.env.user.role != "superadmin" and profile.company_id != self._company_id()
        ):
                return None, Response.not_found("Service profile not found")
        return profile, None

    def _procedure_or_404(self, procedure_id: Any) -> Tuple[Optional[Any], Optional[Response]]:
        try:
            from modules.safety_procedures.module_safety_procedures import SafetyProcedureTemplate

            procedure = SafetyProcedureTemplate.find_by_id(_safe_int(procedure_id))
        except Exception:
            procedure = None
        if not procedure or (
            self.env.user.role != "superadmin" and procedure.company_id != self._company_id()
        ):
            return None, Response.not_found("Procedure not found")
        return procedure, None

    def _folder_or_404(self, folder_id: Any) -> Tuple[Optional[SafetyFolder], Optional[Response]]:
        folder = SafetyFolder.find_by_id(_safe_int(folder_id))
        if not folder or (
            self.env.user.role != "superadmin" and folder.company_id != self._company_id()
        ):
            return None, Response.not_found("Safety folder not found")
        return folder, None

    def _risk_matrix_or_404(self, matrix_id: Any) -> Tuple[Optional[SafetyRiskMatrix], Optional[Response]]:
        matrix = SafetyRiskMatrix.find_by_id(_safe_int(matrix_id))
        if not matrix or (
            self.env.user.role != "superadmin" and matrix.company_id != self._company_id()
        ):
            return None, Response.not_found("Risk matrix not found")
        return matrix, None

    def _risk_matrix_row_or_404(
        self,
        matrix: SafetyRiskMatrix,
        row_id: Any,
    ) -> Tuple[Optional[SafetyRiskMatrixRow], Optional[Response]]:
        row = SafetyRiskMatrixRow.find_by_id(_safe_int(row_id))
        if not row or row.risk_matrix_id != matrix.id or row.company_id != matrix.company_id:
            return None, Response.not_found("Risk matrix row not found")
        return row, None

    def _document_or_404(
        self, document_id: Any
    ) -> Tuple[Optional[SafetyFolderDocument], Optional[Response]]:
        document = SafetyFolderDocument.find_by_id(_safe_int(document_id))
        if not document or (
            self.env.user.role != "superadmin" and document.company_id != self._company_id()
        ):
            return None, Response.not_found("Document not found")
        return document, None

    def _delivery_or_404(
        self, delivery_id: Any
    ) -> Tuple[Optional[SafetyPPEDelivery], Optional[Response]]:
        delivery = SafetyPPEDelivery.find_by_id(_safe_int(delivery_id))
        if not delivery or (
            self.env.user.role != "superadmin" and delivery.company_id != self._company_id()
        ):
            return None, Response.not_found("PPE delivery not found")
        return delivery, None

    def _talk_or_404(self, talk_id: Any) -> Tuple[Optional[SafetyTalk], Optional[Response]]:
        talk = SafetyTalk.find_by_id(_safe_int(talk_id))
        if not talk or (
            self.env.user.role != "superadmin" and talk.company_id != self._company_id()
        ):
            return None, Response.not_found("Talk not found")
        return talk, None

    def _checklist_or_404(
        self, checklist_id: Any
    ) -> Tuple[Optional[SafetyChecklistRun], Optional[Response]]:
        checklist = SafetyChecklistRun.find_by_id(_safe_int(checklist_id))
        if not checklist or (
            self.env.user.role != "superadmin" and checklist.company_id != self._company_id()
        ):
            return None, Response.not_found("Checklist not found")
        return checklist, None

    def _irl_or_404(
        self, irl_id: Any
    ) -> Tuple[Optional[SafetyIRLRecord], Optional[Response]]:
        irl = SafetyIRLRecord.find_by_id(_safe_int(irl_id))
        if not irl or (
            self.env.user.role != "superadmin" and irl.company_id != self._company_id()
        ):
            return None, Response.not_found("IRL record not found")
        return irl, None

    def _generator_rule_or_404(
        self, rule_id: Any
    ) -> Tuple[Optional[SafetyGeneratorRule], Optional[Response]]:
        rule = SafetyGeneratorRule.find_by_id(_safe_int(rule_id))
        if not rule or (
            self.env.user.role != "superadmin" and rule.company_id != self._company_id()
        ):
            return None, Response.not_found("Generator rule not found")
        return rule, None

    def _site_or_404(self, site_id: Any) -> Tuple[Optional[SafetyClientSite], Optional[Response]]:
        site = SafetyClientSite.find_by_id(_safe_int(site_id))
        if not site or (
            self.env.user.role != "superadmin" and site.company_id != self._company_id()
        ):
            return None, Response.not_found("Client site not found")
        return site, None

    def _area_or_404(self, area_id: Any) -> Tuple[Optional[SafetyClientArea], Optional[Response]]:
        area = SafetyClientArea.find_by_id(_safe_int(area_id))
        if not area or (
            self.env.user.role != "superadmin" and area.company_id != self._company_id()
        ):
            return None, Response.not_found("Client area not found")
        return area, None

    def _restriction_or_404(
        self, restriction_id: Any
    ) -> Tuple[Optional[SafetyWorkerRestriction], Optional[Response]]:
        restriction = SafetyWorkerRestriction.find_by_id(_safe_int(restriction_id))
        if not restriction or (
            self.env.user.role != "superadmin" and restriction.company_id != self._company_id()
        ):
            return None, Response.not_found("Worker restriction not found")
        return restriction, None

    def _lead_or_404(self, lead_id: Any):
        try:
            from modules.crm.module_crm import Lead

            lead = Lead.find_by_id(_safe_int(lead_id))
            if not lead or (
                self.env.user.role != "superadmin" and lead.company_id != self._company_id()
            ):
                return None, Response.not_found("Lead not found")
            return lead, None
        except Exception:
            return None, Response.not_found("Lead not found")

    def _employee_ids_for_company(self) -> Set[int]:
        try:
            from modules.hr.module_hr import EmployeeProfile

            employees = EmployeeProfile.search(self._tenant_filter())
            return {employee.id for employee in employees if employee.id}
        except Exception:
            return set()

    def _validate_employee_assignments(self, employee_ids: List[int]) -> Optional[Response]:
        available = self._employee_ids_for_company()
        if not employee_ids:
            return None
        invalid = [emp_id for emp_id in employee_ids if emp_id not in available]
        if invalid:
            return Response.bad_request("One or more employees do not belong to your company")
        return None

    def _folder_procedure_ids(self, folder: SafetyFolder) -> List[int]:
        ids = _normalize_int_list(getattr(folder, "procedure_ids", []))
        if not ids and folder.procedure_id:
            ids = [folder.procedure_id]
        return ids

    def _folder_job_profile_ids(self, folder: SafetyFolder) -> List[int]:
        return _normalize_int_list(getattr(folder, "job_profile_ids", []))

    def _folder_client_area_ids(self, folder: SafetyFolder) -> List[int]:
        ids = _normalize_int_list(getattr(folder, "client_area_ids", []))
        if not ids and folder.client_area_id:
            ids = [folder.client_area_id]
        return ids

    def _folder_equipment_block_ids(self, folder: SafetyFolder) -> List[int]:
        return _normalize_int_list(getattr(folder, "equipment_block_ids", []))

    def _validate_procedure_ids(self, procedure_ids: List[int]) -> Tuple[List[Any], Optional[Response]]:
        procedures = []
        for procedure_id in procedure_ids:
            procedure, error = self._procedure_or_404(procedure_id)
            if error:
                return [], error
            if procedure and getattr(procedure, "active", True):
                procedures.append(procedure)
        return procedures, None

    def _validate_job_profile_ids(self, job_profile_ids: List[int]) -> Tuple[List[Any], Optional[Response]]:
        if not job_profile_ids:
            return [], None
        try:
            from modules.job_profiles.module_job_profiles import JobProfile
        except Exception:
            return [], Response.bad_request("Job profiles are not available")
        profiles = []
        for profile_id in job_profile_ids:
            profile = JobProfile.find_by_id(_safe_int(profile_id))
            if not profile or (
                self.env.user.role != "superadmin" and profile.company_id != self._company_id()
            ):
                return [], Response.not_found("Job profile not found")
            if getattr(profile, "active", True):
                profiles.append(profile)
        return profiles, None

    def _validate_equipment_block_ids(
        self, equipment_block_ids: List[int]
    ) -> Tuple[List[SafetyEquipmentBlock], Optional[Response]]:
        blocks = []
        for equipment_id in equipment_block_ids:
            block = SafetyEquipmentBlock.find_by_id(_safe_int(equipment_id))
            if not block or (
                self.env.user.role != "superadmin" and block.company_id != self._company_id()
            ):
                return [], Response.not_found("Equipment block not found")
            if block.active:
                blocks.append(block)
        return blocks, None

    def _validate_client_area_ids(
        self, client_area_ids: List[int], lead: Any = None, site_id: Optional[int] = None
    ) -> Tuple[List[SafetyClientArea], Optional[Response]]:
        areas = []
        for area_id in client_area_ids:
            area, error = self._area_or_404(area_id)
            if error:
                return [], error
            site = SafetyClientSite.find_by_id(area.site_id) if area else None
            if site_id and area and area.site_id != site_id:
                return [], Response.bad_request("One or more areas do not belong to the chosen site")
            if lead and getattr(lead, "customer_id", None) and site and site.customer_id != lead.customer_id:
                return [], Response.bad_request("One or more areas do not belong to the lead customer")
            if area and area.active:
                areas.append(area)
        return areas, None

    def _sync_folder_scope(self, folder: SafetyFolder) -> None:
        folder.procedure_ids = self._folder_procedure_ids(folder)
        folder.client_area_ids = self._folder_client_area_ids(folder)
        folder.job_profile_ids = self._folder_job_profile_ids(folder)
        folder.equipment_block_ids = self._folder_equipment_block_ids(folder)
        folder.procedure_id = folder.procedure_ids[0] if folder.procedure_ids else None
        folder.client_area_id = folder.client_area_ids[0] if folder.client_area_ids else None

    def _folder_scope_snapshot(self, folder: SafetyFolder) -> Dict[str, Any]:
        payload = folder.to_dict()
        sources = [
            ("procedure", "Procedimientos", payload.get("procedures") or [], "name"),
            ("cargo_profile", "Cargos", payload.get("job_profiles") or [], "name"),
            ("shared_place", "Sectores", payload.get("client_areas") or [], "name"),
            ("equipment", "Equipos", payload.get("equipment_blocks") or [], "name"),
        ]
        scope = []
        for key, label, items, name_key in sources:
            names = [
                str(item.get(name_key) or item.get("code") or item.get("procedure_code") or item.get("id") or "").strip()
                for item in items
            ]
            scope.append(
                {
                    "key": key,
                    "label": label,
                    "count": len([name for name in names if name]),
                    "names": [name for name in names if name],
                }
            )
        return {"sources": scope}

    def _tag_matrix_row(
        self,
        row: Dict[str, Any],
        source_group: str,
        source_title: str,
        source_id: Optional[int],
        index: int,
    ) -> Dict[str, Any]:
        tagged = dict(row)
        source_title = str(source_title or MATRIX_SOURCE_LABELS.get(source_group, source_group)).strip()
        source_id = _safe_int(source_id, None)
        base_label = MATRIX_SOURCE_LABELS.get(source_group, source_group)
        label = f"{base_label}: {source_title}" if source_title else base_label
        tagged["source_group"] = source_group
        tagged["source_title"] = source_title
        tagged["source_id"] = source_id
        tagged["source_sort"] = [MATRIX_SOURCE_ORDER.get(source_group, 99), index, source_id or 0]
        tagged["source_groups"] = dedupe_preserve_order(
            _normalize_str_list(tagged.get("source_groups")) + [source_group]
        )
        tagged["source_titles"] = dedupe_preserve_order(
            _normalize_str_list(tagged.get("source_titles")) + ([source_title] if source_title else [])
        )
        tagged["source_refs"] = list(tagged.get("source_refs") or [])
        tagged["source_refs"].append({"group": source_group, "id": source_id, "title": source_title})
        tagged["origin_blocks"] = dedupe_preserve_order(
            [source_group] + _normalize_str_list(tagged.get("origin_blocks"))
        )
        tagged["source_labels"] = dedupe_preserve_order(
            [label] + _normalize_str_list(tagged.get("source_labels"))
        )
        tagged["row_fingerprint"] = tagged.get("row_fingerprint") or build_row_fingerprint(tagged)
        return tagged

    def _merge_matrix_row(self, rows_map: Dict[str, Dict[str, Any]], row: Dict[str, Any]) -> None:
        fingerprint = row.get("row_fingerprint") or build_row_fingerprint(row)
        row["row_fingerprint"] = fingerprint
        if fingerprint not in rows_map:
            rows_map[fingerprint] = row
            return
        merged = merge_generated_rows(rows_map[fingerprint], row)
        merged["source_groups"] = dedupe_preserve_order(
            _normalize_str_list(rows_map[fingerprint].get("source_groups"))
            + _normalize_str_list(row.get("source_groups"))
            + _normalize_str_list([rows_map[fingerprint].get("source_group"), row.get("source_group")])
        )
        merged["source_titles"] = dedupe_preserve_order(
            _normalize_str_list(rows_map[fingerprint].get("source_titles"))
            + _normalize_str_list(row.get("source_titles"))
            + _normalize_str_list([rows_map[fingerprint].get("source_title"), row.get("source_title")])
        )
        merged["source_refs"] = (rows_map[fingerprint].get("source_refs") or []) + (row.get("source_refs") or [])
        sort_left = rows_map[fingerprint].get("source_sort") or [99, 0, 0]
        sort_right = row.get("source_sort") or [99, 0, 0]
        merged["source_sort"] = min(sort_left, sort_right)
        merged["source_group"] = (merged.get("source_groups") or [row.get("source_group") or ""])[0]
        merged["source_title"] = (merged.get("source_titles") or [row.get("source_title") or ""])[0]
        rows_map[fingerprint] = merged

    def _find_profile_for_lead(self, lead) -> Optional[SafetyServiceProfile]:
        profiles = SafetyServiceProfile.search(self._tenant_filter())
        if not profiles:
            return None
        if getattr(lead, "service_type_id", None):
            for profile in profiles:
                if profile.service_type_id == lead.service_type_id and profile.active:
                    return profile
        lead_name = " ".join(
            [
                str(getattr(lead, "title", "") or ""),
                str(getattr(lead, "description", "") or ""),
            ]
        ).lower()
        for profile in profiles:
            if profile.active and profile.name and profile.name.lower() in lead_name:
                return profile
        active_profiles = [profile for profile in profiles if profile.active]
        return active_profiles[0] if active_profiles else None

    def _find_procedure_for_profile(self, profile: Optional[SafetyServiceProfile]) -> Optional[Any]:
        if not profile:
            return None
        try:
            from modules.safety_procedures.module_safety_procedures import (
                SafetyProcedureTemplate,
                seed_default_procedures,
            )

            seed_default_procedures(self._company_id())
            procedures = [
                procedure
                for procedure in SafetyProcedureTemplate.search(self._tenant_filter())
                if procedure.active and procedure.status != "archived"
            ]
        except Exception:
            return None

        for procedure in procedures:
            if procedure.service_profile_id and procedure.service_profile_id == profile.id:
                return procedure
        profile_name = str(profile.name or "").strip().lower()
        for procedure in procedures:
            if profile_name and profile_name in str(procedure.name or "").strip().lower():
                return procedure
        return None

    def _documents_for_folder(self, folder_id: int) -> List[SafetyFolderDocument]:
        documents = SafetyFolderDocument.search([("folder_id", "=", folder_id)])
        documents.sort(key=lambda item: ((item.title or "").lower(), item.id or 0))
        return documents

    def _matrix_for_folder(self, folder_id: int) -> Optional[SafetyRiskMatrix]:
        matrices = SafetyRiskMatrix.search([("folder_id", "=", folder_id)])
        matrices.sort(key=lambda item: item.id or 0, reverse=True)
        return matrices[0] if matrices else None

    def _deliveries_for_folder(self, folder_id: int) -> List[SafetyPPEDelivery]:
        deliveries = SafetyPPEDelivery.search([("folder_id", "=", folder_id)])
        deliveries.sort(key=lambda item: (item.delivery_date or "", item.id or 0), reverse=True)
        return deliveries

    def _talks_for_folder(self, folder_id: int) -> List[SafetyTalk]:
        talks = SafetyTalk.search([("folder_id", "=", folder_id)])
        talks.sort(key=lambda item: (item.talk_date or "", item.id or 0), reverse=True)
        return talks

    def _checklists_for_folder(self, folder_id: int) -> List[SafetyChecklistRun]:
        checklists = SafetyChecklistRun.search([("folder_id", "=", folder_id)])
        checklists.sort(key=lambda item: (item.executed_at or "", item.id or 0), reverse=True)
        return checklists

    def _irl_records_for_folder(self, folder_id: int) -> List[SafetyIRLRecord]:
        records = SafetyIRLRecord.search([("folder_id", "=", folder_id)])
        records.sort(key=lambda item: (item.id or 0), reverse=True)
        return records

    def _employee_map(self) -> Dict[int, Any]:
        try:
            from modules.hr.module_hr import EmployeeProfile

            return {
                employee.id: employee
                for employee in EmployeeProfile.search(self._tenant_filter())
                if employee.id
            }
        except Exception:
            return {}

    def _employee_payload(self, employee_id: Optional[int]) -> Dict[str, Any]:
        employee = self._employee_map().get(employee_id or 0)
        if not employee:
            return {}
        return employee.to_dict()

    def _company_employee_count(self) -> int:
        try:
            from modules.hr.module_hr import EmployeeProfile

            employees = EmployeeProfile.search(self._tenant_filter())
            return len([employee for employee in employees if (employee.status or "") != "inactive"])
        except Exception:
            return 0

    def _legal_snapshot(self, folder: SafetyFolder) -> Dict[str, Any]:
        worker_count = self._company_employee_count()
        return {
            "worker_count": worker_count,
            "segment": (
                "<=25"
                if worker_count <= 25
                else ("26-100" if worker_count <= 100 else ">100")
            ),
            "requirements": legal_requirements_for_headcount(worker_count),
        }

    def _master_risk_map(self) -> Dict[int, SafetyMasterRisk]:
        self._ensure_miper_catalog()
        return {
            risk.id: risk
            for risk in SafetyMasterRisk.search(self._tenant_filter())
            if risk.id and risk.active
        }

    def _protocol_code_map(self) -> Dict[str, SafetyProtocol]:
        self._ensure_miper_catalog()
        return {
            str(protocol.code or "").strip().upper(): protocol
            for protocol in SafetyProtocol.search(self._tenant_filter())
            if protocol.active
        }

    def _ppe_catalog(self) -> List[SafetyPPEItem]:
        self._ensure_miper_catalog()
        items = SafetyPPEItem.search(self._tenant_filter())
        items = [item for item in items if item.active]
        items.sort(key=lambda item: ((item.category or "").lower(), (item.name or "").lower(), item.id or 0))
        return items

    def _equipment_block_payload(self, equipment: SafetyEquipmentBlock) -> Dict[str, Any]:
        return equipment.to_dict()

    def _equipment_rows_for_folder(
        self,
        folder: SafetyFolder,
        procedure_payload: Any,
        place_name: str,
    ) -> List[Dict[str, Any]]:
        procedure_payloads = procedure_payload if isinstance(procedure_payload, list) else [procedure_payload]
        tool_refs: List[str] = []
        procedure_names: List[str] = []
        for payload in procedure_payloads:
            procedure = (payload or {}).get("procedure") if isinstance(payload, dict) else {}
            if not isinstance(procedure, dict):
                continue
            procedure_names.append(str(procedure.get("name") or procedure.get("procedure_code") or "").strip())
            tool_refs.extend(_normalize_str_list(procedure.get("tools_and_equipment")))

        risk_map = self._master_risk_map()
        equipment_items = [item for item in SafetyEquipmentBlock.search(self._tenant_filter()) if item.active]
        by_code = {str(item.code or "").strip().upper(): item for item in equipment_items if item.code}
        by_name = {str(item.name or "").strip().upper(): item for item in equipment_items if item.name}

        matched_items: List[SafetyEquipmentBlock] = []
        selected_equipment_ids = set(self._folder_equipment_block_ids(folder))
        for candidate in equipment_items:
            if candidate.id in selected_equipment_ids:
                matched_items.append(candidate)
        for tool in tool_refs:
            key = str(tool or "").strip().upper()
            if not key:
                continue
            match = by_code.get(key) or by_name.get(key)
            if not match:
                for candidate in equipment_items:
                    haystack = f"{candidate.code or ''} {candidate.name or ''}".strip().upper()
                    if key in haystack or haystack in key:
                        match = candidate
                        break
            if match and all(existing.id != match.id for existing in matched_items):
                matched_items.append(match)

        rows: List[Dict[str, Any]] = []
        folder_payload = folder.to_dict()
        process_name = ", ".join([name for name in procedure_names if name]) or folder_payload.get("folder_name") or "Equipo y herramienta"
        for equipment in matched_items:
            payload = equipment.to_dict()
            probability = _safe_int(payload.get("probability"), 2) or 2
            consequence = _safe_int(payload.get("consequence"), 2) or 2
            vep = calculate_vep(probability, consequence)
            for risk_info in payload.get("master_risks") or []:
                risk = risk_map.get(_safe_int(risk_info.get("id")))
                if not risk:
                    continue
                row = {
                    "activity": payload.get("name") or payload.get("code") or "Equipo operativo",
                    "process_name": process_name,
                    "task_name": payload.get("name") or payload.get("code") or "Equipo operativo",
                    "position_name": folder_payload.get("responsible_name") or "Equipo / herramienta",
                    "place_name": place_name,
                    "hazard": payload.get("description") or payload.get("name") or "",
                    "hazard_factor": payload.get("description") or payload.get("name") or "",
                    "risk": risk.risk_name or "",
                    "risk_name": risk.risk_name or "",
                    "master_risk_id": risk.id,
                    "master_risk_code": risk.isp_code or "",
                    "risk_family": risk.family or "",
                    "controls": payload.get("controls_summary") or summarize_controls(payload.get("control_hierarchy"), ""),
                    "control_hierarchy": normalize_control_hierarchy(payload.get("control_hierarchy")),
                    "required_ppe": _normalize_str_list(payload.get("required_ppe")),
                    "protocol_codes": dedupe_preserve_order(
                        _normalize_str_list(payload.get("protocol_codes"))
                        + _normalize_str_list(risk.protocol_codes)
                    ),
                    "owner_name": "Equipo / herramienta",
                    "probability": probability,
                    "consequence": consequence,
                    "vep": vep,
                    "risk_level": risk_level_from_vep(vep),
                    "action_required": action_from_vep(vep),
                    "origin_blocks": ["equipment"],
                    "origin_rule_ids": [],
                    "source_labels": [f"Equipo: {payload.get('name') or payload.get('code') or equipment.id}"],
                    "legal_reference": payload.get("legal_reference") or "",
                    "sensitivity_tags": _normalize_str_list(payload.get("sensitivity_tags")),
                    "restriction_alerts": [],
                    "equipment_block_id": equipment.id,
                }
                row["row_fingerprint"] = build_row_fingerprint(row)
                rows.append(row)
        return rows

    def _rules_for_folder(
        self, folder: SafetyFolder
    ) -> Tuple[List[Tuple[str, SafetyGeneratorRule]], Dict[str, int]]:
        self._ensure_miper_catalog()
        lead = None
        try:
            from modules.crm.module_crm import Lead

            lead = Lead.find_by_id(folder.lead_id)
        except Exception:
            lead = None
        matches: List[Tuple[str, SafetyGeneratorRule]] = []
        summary = {
            "transversal": 0,
            "service_profile": 0,
            "customer": 0,
            "client_site": 0,
            "client_area": 0,
        }
        area_ids = set(self._folder_client_area_ids(folder))
        rules = [rule for rule in SafetyGeneratorRule.search(self._tenant_filter()) if rule.active]
        for rule in rules:
            block_name = str(rule.scope_type or "").strip()
            matched = False
            if block_name == "transversal":
                matched = True
            elif block_name == "service_profile" and folder.service_profile_id:
                matched = rule.scope_ref_id == folder.service_profile_id
            elif block_name == "customer" and lead and lead.customer_id:
                matched = rule.scope_ref_id == lead.customer_id
            elif block_name == "client_site" and folder.client_site_id:
                matched = rule.scope_ref_id == folder.client_site_id
            elif block_name == "client_area" and area_ids:
                matched = rule.scope_ref_id in area_ids
            if matched:
                matches.append((block_name, rule))
                summary[block_name] = summary.get(block_name, 0) + 1
        return matches, summary

    def _procedure_matrix_payload_for_folder(
        self,
        folder: SafetyFolder,
        place_name: str,
    ) -> Dict[str, Any]:
        if not folder.procedure_id:
            return {"procedure": None, "rows": [], "step_count": 0, "activity_block_ids": []}
        try:
            from modules.safety_procedures.module_safety_procedures import (
                build_procedure_matrix_payload,
                seed_default_procedures,
            )

            company_id = folder.company_id or self._company_id()
            seed_default_procedures(company_id)
            return build_procedure_matrix_payload(
                folder.procedure_id,
                company_id=company_id,
                place_name=place_name,
            )
        except Exception:
            return {"procedure": None, "rows": [], "step_count": 0, "activity_block_ids": []}

    def _procedure_matrix_payloads_for_folder(
        self,
        folder: SafetyFolder,
        place_name: str,
    ) -> List[Dict[str, Any]]:
        procedure_ids = self._folder_procedure_ids(folder)
        if not procedure_ids:
            payload = self._procedure_matrix_payload_for_folder(folder, place_name)
            return [payload] if payload.get("rows") else []
        payloads: List[Dict[str, Any]] = []
        original_id = folder.procedure_id
        try:
            for procedure_id in procedure_ids:
                folder.procedure_id = procedure_id
                payload = self._procedure_matrix_payload_for_folder(folder, place_name)
                if payload.get("rows") or payload.get("procedure"):
                    payloads.append(payload)
        finally:
            folder.procedure_id = original_id
        return payloads

    def _default_place_risk_score(self, risk_code: str) -> Tuple[int, int]:
        high_codes = {"A3", "F1", "F3", "I1", "J", "J2", "B2"}
        if str(risk_code or "").strip().upper() in high_codes:
            return 2, 4
        return 2, 2

    def _shared_place_risk_rows_for_folder(
        self,
        folder: SafetyFolder,
        place_name: str,
    ) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        folder_payload = folder.to_dict()
        customer_id = _safe_int(folder_payload.get("customer_id"), None)
        area_name = str(folder_payload.get("client_area_name") or "").strip().lower()
        selected_area_names = {
            str(item.get("name") or "").strip().lower()
            for item in (folder_payload.get("client_areas") or [])
            if str(item.get("name") or "").strip()
        }
        if not customer_id:
            return [], {"source": "reports", "risk_count": 0, "matched_areas": []}

        try:
            from modules.reports.module_reports import (
                AreaFaena,
                AreaRiskAssignment,
                SectorFaena,
                SectorRiskAssignment,
            )
        except Exception:
            return [], {"source": "reports", "risk_count": 0, "matched_areas": []}

        report_areas = AreaFaena.search(
            [
                ("company_id", "=", folder.company_id),
                ("customer_id", "=", customer_id),
                ("active", "=", True),
            ]
        )
        exact_areas = [
            area
            for area in report_areas
            if (
                selected_area_names
                and str(area.nombre or "").strip().lower() in selected_area_names
            )
            or (
                not selected_area_names
                and area_name
                and str(area.nombre or "").strip().lower() == area_name
            )
        ]
        if exact_areas:
            report_areas = exact_areas
        elif selected_area_names:
            report_areas = []

        risk_map = self._master_risk_map()
        rows_map: Dict[str, Dict[str, Any]] = {}
        matched_areas: List[str] = []
        for area in report_areas:
            if area.nombre and area.nombre not in matched_areas:
                matched_areas.append(area.nombre)
            sector_ids = [
                sector.id
                for sector in SectorFaena.search(
                    [
                        ("company_id", "=", folder.company_id),
                        ("area_id", "=", area.id),
                        ("active", "=", True),
                    ]
                )
                if sector.id
            ]
            risk_ids: List[int] = []
            for assignment in AreaRiskAssignment.search(
                [
                    ("company_id", "=", folder.company_id),
                    ("area_id", "=", area.id),
                    ("active", "=", True),
                ]
            ):
                risk_id = _safe_int(assignment.master_risk_id, None)
                if risk_id and risk_id not in risk_ids:
                    risk_ids.append(risk_id)
            for sector_id in sector_ids:
                for assignment in SectorRiskAssignment.search(
                    [
                        ("company_id", "=", folder.company_id),
                        ("sector_id", "=", sector_id),
                        ("active", "=", True),
                    ]
                ):
                    risk_id = _safe_int(assignment.master_risk_id, None)
                    if risk_id and risk_id not in risk_ids:
                        risk_ids.append(risk_id)

            area_label = (
                str(area.nombre or "").strip()
                or place_name
                or str(folder_payload.get("client_site_name") or "").strip()
                or str(folder_payload.get("customer_name") or "").strip()
                or "Area cliente"
            )
            for risk_id in risk_ids:
                risk = risk_map.get(risk_id)
                if not risk:
                    continue
                probability, consequence = self._default_place_risk_score(risk.isp_code or "")
                vep = calculate_vep(probability, consequence)
                row = {
                    "process_name": "Contexto del lugar",
                    "task_name": f"Exposicion propia del area o sector: {area_label}",
                    "activity": f"Exposicion propia del area o sector: {area_label}",
                    "position_name": "Todo el personal",
                    "place_name": place_name or area_label,
                    "hazard": risk.official_definition
                    or f"Peligro especifico del area o sector {area_label}",
                    "hazard_factor": risk.official_definition
                    or f"Peligro especifico del area o sector {area_label}",
                    "risk": risk.risk_name or "",
                    "risk_name": risk.risk_name or "",
                    "master_risk_id": risk.id,
                    "master_risk_code": risk.isp_code or "",
                    "risk_family": risk.family or "",
                    "controls": (
                        "Aplicar controles del cliente, segregacion del area, AST/ART local y "
                        f"verificacion en terreno para {risk.risk_name or 'riesgo especifico'}."
                    ),
                    "control_hierarchy": normalize_control_hierarchy(
                        {
                            "administrative": [
                                "Revisar riesgos declarados por cliente, area y sector",
                                "Incorporar controles especificos en AST/ART y charla previa",
                                "Detener trabajos ante condiciones no controladas del entorno",
                            ]
                        }
                    ),
                    "required_ppe": [],
                    "protocol_codes": _normalize_str_list(risk.protocol_codes),
                    "owner_name": "Supervisor de terreno",
                    "probability": probability,
                    "consequence": consequence,
                    "vep": vep,
                    "risk_level": risk_level_from_vep(vep),
                    "action_required": action_from_vep(vep),
                    "origin_blocks": ["shared_place", "reports_area_sector"],
                    "origin_rule_ids": [],
                    "source_labels": [
                        f"{folder_payload.get('customer_name') or 'Cliente'} / {area_label}"
                    ],
                    "sensitivity_tags": [],
                    "restriction_alerts": [],
                    "legal_reference": "Riesgos operacionales declarados en areas/sectores del cliente",
                    "source_note": "Fuente compartida desde modulo Reports",
                    "generated_at": utc_now_iso(),
                    "is_blocking": vep >= 16,
                }
                row["row_fingerprint"] = build_row_fingerprint(row)
                fingerprint = row["row_fingerprint"]
                if fingerprint in rows_map:
                    rows_map[fingerprint] = merge_generated_rows(rows_map[fingerprint], row)
                else:
                    rows_map[fingerprint] = row

        rows = list(rows_map.values())
        rows.sort(
            key=lambda item: (
                str(item.get("place_name") or "").lower(),
                str(item.get("task_name") or "").lower(),
                str(item.get("master_risk_code") or "").lower(),
            )
        )
        return rows, {
            "source": "reports",
            "risk_count": len(rows),
            "matched_areas": matched_areas,
            "source_customer_id": customer_id,
            "source_area_name": folder_payload.get("client_area_name") or "",
        }

    def _restriction_alerts_for_row(
        self, row: Dict[str, Any], folder: SafetyFolder
    ) -> List[str]:
        assigned_ids = set(_normalize_int_list(folder.assigned_employee_ids))
        if not assigned_ids:
            return []
        row_tags = {
            item.strip().lower()
            for item in (
                _normalize_str_list(row.get("sensitivity_tags"))
                + _normalize_str_list(row.get("protocol_codes"))
                + _normalize_str_list([row.get("master_risk_code")])
            )
            if item
        }
        alerts: List[str] = []
        for restriction in SafetyWorkerRestriction.search(self._tenant_filter()):
            if not restriction.active or restriction.employee_id not in assigned_ids:
                continue
            tags = {tag.strip().lower() for tag in _normalize_str_list(restriction.applies_to_tags)}
            if tags and row_tags.isdisjoint(tags):
                continue
            employee_name = restriction.to_dict().get("employee_name") or f"Trabajador {restriction.employee_id}"
            alerts.append(
                f"{employee_name}: {restriction.title} ({restriction.severity})"
            )
        return dedupe_preserve_order(alerts)

    def _build_generated_row(
        self,
        folder: SafetyFolder,
        block_name: str,
        rule: SafetyGeneratorRule,
        risk: Optional[SafetyMasterRisk],
        lead_customer_name: str,
        site_name: str,
        area_name: str,
    ) -> Dict[str, Any]:
        control_hierarchy = normalize_control_hierarchy(rule.control_hierarchy)
        risk_protocols = _normalize_str_list(risk.protocol_codes) if risk else []
        protocol_codes = dedupe_preserve_order(
            _normalize_str_list(rule.protocol_codes) + risk_protocols
        )
        row = {
            "process_name": rule.process_name or "Proceso del servicio",
            "task_name": rule.task_name or "Tarea operativa",
            "position_name": rule.position_name or "Personal expuesto",
            "place_name": area_name or site_name or lead_customer_name,
            "hazard": rule.hazard_factor or "",
            "hazard_factor": rule.hazard_factor or "",
            "risk": risk.risk_name if risk else "",
            "risk_name": risk.risk_name if risk else "",
            "master_risk_id": risk.id if risk else None,
            "master_risk_code": risk.isp_code if risk else "",
            "risk_family": risk.family if risk else "",
            "controls": rule.controls_summary or summarize_controls(control_hierarchy, ""),
            "control_hierarchy": control_hierarchy,
            "required_ppe": _normalize_str_list(rule.required_ppe),
            "protocol_codes": protocol_codes,
            "owner_name": rule.owner_name or "Supervisor de terreno",
            "probability": _safe_int(rule.probability, 2) or 2,
            "consequence": _safe_int(rule.consequence, 2) or 2,
            "origin_blocks": [block_name],
            "origin_rule_ids": [rule.id] if rule.id else [],
            "source_labels": [rule.name or block_name],
            "sensitivity_tags": _normalize_str_list(rule.sensitivity_tags),
            "restriction_alerts": [],
            "legal_reference": rule.legal_reference or "",
            "source_note": rule.source_note or "",
                "generated_at": utc_now_iso(),
        }
        row["vep"] = calculate_vep(row["probability"], row["consequence"])
        row["risk_level"] = risk_level_from_vep(row["vep"])
        row["action_required"] = action_from_vep(row["vep"])
        row["restriction_alerts"] = self._restriction_alerts_for_row(row, folder)
        row["is_blocking"] = bool(row["vep"] >= 16 or row["restriction_alerts"])
        row["row_fingerprint"] = build_row_fingerprint(row)
        row["activity"] = row["task_name"]
        return row

    def _fallback_generated_rows(self, folder: SafetyFolder) -> List[Dict[str, Any]]:
        profile = SafetyServiceProfile.find_by_id(folder.service_profile_id) if folder.service_profile_id else None
        rows: List[Dict[str, Any]] = []
        for item in _default_risk_rows(profile.name if profile else "Servicio General"):
            row = _normalize_matrix_rows([item])[0]
            row["process_name"] = row.get("process_name") or "Proceso del servicio"
            row["task_name"] = row.get("task_name") or row.get("activity") or "Tarea operativa"
            row["origin_blocks"] = ["fallback"]
            row["source_labels"] = ["Regla base de respaldo"]
            row["row_fingerprint"] = build_row_fingerprint(row)
            rows.append(row)
        return rows

    def _generate_matrix_for_folder(self, folder: SafetyFolder) -> Tuple[SafetyRiskMatrix, Dict[str, Any]]:
        self._ensure_miper_catalog()
        folder_payload = folder.to_dict()
        lead_customer_name = folder_payload.get("customer_name") or ""
        site_name = folder_payload.get("client_site_name") or ""
        area_names = _normalize_str_list(folder_payload.get("client_area_names"))
        area_name = ", ".join(area_names) or folder_payload.get("client_area_name") or ""
        place_name = area_name or site_name or lead_customer_name
        procedure_payloads = self._procedure_matrix_payloads_for_folder(folder, place_name)
        matched_rules: List[Tuple[str, SafetyGeneratorRule]] = []
        by_block = {
            "procedure": 0,
            "shared_place": 0,
            "cargo_profile": 0,
            "equipment": 0,
            "transversal": 0,
            "service_profile": 0,
            "customer": 0,
            "client_site": 0,
            "client_area": 0,
        }
        if not any(payload.get("rows") for payload in procedure_payloads):
            matched_rules, legacy_summary = self._rules_for_folder(folder)
            by_block.update(legacy_summary)
        risk_map = self._master_risk_map()
        rows_map: Dict[str, Dict[str, Any]] = {}

        procedure_row_count = 0
        for index, payload in enumerate(procedure_payloads, start=1):
            procedure = payload.get("procedure") or {}
            title = (
                procedure.get("name")
                or procedure.get("procedure_code")
                or f"Procedimiento {index}"
            )
            for row in payload.get("rows", []):
                tagged = self._tag_matrix_row(row, "procedure", title, _safe_int(procedure.get("id"), None), index)
                self._merge_matrix_row(rows_map, tagged)
                procedure_row_count += 1
        by_block["procedure"] = procedure_row_count

        for index, (block_name, rule) in enumerate(matched_rules, start=1):
            risk = risk_map.get(rule.master_risk_id)
            row = self._build_generated_row(
                folder,
                block_name,
                rule,
                risk,
                lead_customer_name,
                site_name,
                area_name,
            )
            tagged = self._tag_matrix_row(row, block_name, rule.name or block_name, rule.id, index)
            self._merge_matrix_row(rows_map, tagged)

        place_rows, shared_place_summary = self._shared_place_risk_rows_for_folder(
            folder,
            place_name,
        )
        for index, row in enumerate(place_rows, start=1):
            title = (
                str(row.get("place_name") or "").strip()
                or ", ".join(shared_place_summary.get("matched_areas") or [])
                or "Sector cliente"
            )
            tagged = self._tag_matrix_row(row, "shared_place", title, None, index)
            self._merge_matrix_row(rows_map, tagged)
        by_block["shared_place"] = len(place_rows)

        equipment_rows = self._equipment_rows_for_folder(folder, procedure_payloads, place_name)
        for index, row in enumerate(equipment_rows, start=1):
            title = str(row.get("task_name") or row.get("activity") or "Equipo").strip()
            tagged = self._tag_matrix_row(row, "equipment", title, _safe_int(row.get("equipment_block_id"), None), index)
            self._merge_matrix_row(rows_map, tagged)
        by_block["equipment"] = len(equipment_rows)

        job_profile_payload = {"rows": [], "matched_employees": [], "matched_profiles": []}
        try:
            from modules.job_profiles.module_job_profiles import (
                JobProfile,
                build_job_profile_matrix_rows,
                build_personalized_matrix_rows_for_employees,
            )

            job_profile_ids = self._folder_job_profile_ids(folder)
            if job_profile_ids:
                for index, profile_id in enumerate(job_profile_ids, start=1):
                    profile = JobProfile.find_by_id(profile_id)
                    if not profile or not getattr(profile, "active", True):
                        continue
                    profile_rows = build_job_profile_matrix_rows(profile)
                    job_profile_payload["matched_profiles"].append(
                        {
                            "job_profile_id": profile.id,
                            "job_profile_name": profile.name,
                            "row_count": len(profile_rows),
                        }
                    )
                    for row in profile_rows:
                        title = profile.name or profile.code or f"Cargo {profile.id}"
                        tagged = self._tag_matrix_row(row, "cargo_profile", title, profile.id, index)
                        self._merge_matrix_row(rows_map, tagged)
                        job_profile_payload["rows"].append(tagged)
            else:
                job_profile_payload = build_personalized_matrix_rows_for_employees(
                    _normalize_int_list(folder.assigned_employee_ids)
                )
                for index, row in enumerate(job_profile_payload.get("rows", []), start=1):
                    title = str(row.get("job_profile_name") or row.get("position_name") or "Cargo asignado").strip()
                    tagged = self._tag_matrix_row(row, "cargo_profile", title, _safe_int(row.get("job_profile_id"), None), index)
                    self._merge_matrix_row(rows_map, tagged)
        except Exception:
            job_profile_payload = {"rows": [], "matched_employees": [], "matched_profiles": []}
        by_block["cargo_profile"] = len(job_profile_payload.get("rows", []))

        rows = list(rows_map.values()) if rows_map else self._fallback_generated_rows(folder)
        if not rows_map:
            rows = [
                self._tag_matrix_row(row, "fallback", "Regla base de respaldo", None, index)
                for index, row in enumerate(rows, start=1)
            ]
        rows.sort(
            key=lambda item: (
                item.get("source_sort") or [99, 0, 0],
                -(_safe_int(item.get("vep"), 0) or 0),
                str(item.get("source_title") or "").lower(),
                str(item.get("task_name") or "").lower(),
                str(item.get("master_risk_code") or "").lower(),
            )
        )
        source_counts: Dict[str, int] = {}
        for row in rows:
            group = str(row.get("source_group") or (row.get("source_groups") or [""])[0] or "sin_origen")
            source_counts[group] = source_counts.get(group, 0) + 1
        protocol_codes = dedupe_preserve_order(
            [code for row in rows for code in _normalize_str_list(row.get("protocol_codes"))]
        )
        restriction_count = len(
            [row for row in rows if _normalize_str_list(row.get("restriction_alerts"))]
        )
        intolerable_count = len([row for row in rows if row.get("risk_level") == "Intolerable"])
        generation_summary = {
            "generated_at": utc_now_iso(),
            "matched_rule_count": len(matched_rules),
            "matched_by_block": by_block,
            "source_counts": source_counts,
            "scope": self._folder_scope_snapshot(folder),
            "procedure_ids": self._folder_procedure_ids(folder),
            "procedure_id": folder.procedure_id,
            "procedure_code": ((procedure_payloads[0].get("procedure") if procedure_payloads else {}) or {}).get("procedure_code"),
            "procedure_name": ((procedure_payloads[0].get("procedure") if procedure_payloads else {}) or {}).get("name"),
            "procedure_step_count": sum(_safe_int(payload.get("step_count"), 0) or 0 for payload in procedure_payloads),
            "procedure_activity_block_ids": dedupe_preserve_order(
                [
                    block_id
                    for payload in procedure_payloads
                    for block_id in _normalize_int_list(payload.get("activity_block_ids"))
                ]
            ),
            "job_profile_ids": self._folder_job_profile_ids(folder),
            "client_area_ids": self._folder_client_area_ids(folder),
            "equipment_block_ids": self._folder_equipment_block_ids(folder),
            "shared_place_summary": shared_place_summary,
            "equipment_row_count": len(equipment_rows),
            "row_count": len(rows),
            "job_profile_row_count": len(job_profile_payload.get("rows", [])),
            "job_profile_employee_matches": job_profile_payload.get("matched_employees", []),
            "job_profile_matches": job_profile_payload.get("matched_profiles", []),
            "intolerable_count": intolerable_count,
            "restriction_alert_count": restriction_count,
            "protocol_codes": protocol_codes,
            "legal_snapshot": self._legal_snapshot(folder),
        }
        matrix = self._matrix_for_folder(folder.id)
        title = f"MIPER - {folder.to_dict().get('project_code') or folder.id}"
        if matrix:
            matrix.title = title
            matrix.status = "draft"
            matrix.rows = rows
            matrix.generation_summary = generation_summary
            matrix.version = (_safe_int(matrix.version, 1) or 1) + 1
            matrix.source_type = "folder"
            matrix.source_id = folder.id
            matrix.process_name = folder.to_dict().get("service_profile_name") or ""
            matrix.work_center = place_name
            matrix.last_update_date = utc_today_iso()
            matrix.save()
        else:
            matrix = SafetyRiskMatrix.create(
                {
                    "folder_id": folder.id,
                    "code": f"MIPER-FOLDER-{folder.id}",
                    "title": title,
                    "status": "draft",
                    "version": 1,
                    "rows": rows,
                    "generation_summary": generation_summary,
                    "source_type": "manual",
                    "source_id": folder.id,
                    "process_name": folder.to_dict().get("service_profile_name") or "",
                    "work_center": place_name,
                    "elaboration_date": utc_today_iso(),
                    "last_update_date": utc_today_iso(),
                    "company_id": folder.company_id,
                }
            )
        for old_row in SafetyRiskMatrixRow.search([("risk_matrix_id", "=", matrix.id)]):
            old_row.active = False
            old_row.save()
        for index, row in enumerate(rows, start=1):
            SafetyRiskMatrixRow.create(_matrix_row_payload_from_legacy(matrix, row, index * 10))
        return matrix, generation_summary

    def _generate_documents_for_folder(self, folder: SafetyFolder) -> Dict[str, int]:
        profile = SafetyServiceProfile.find_by_id(folder.service_profile_id) if folder.service_profile_id else None
        procedure_blueprints: List[Dict[str, Any]] = []
        procedure_ids = self._folder_procedure_ids(folder)
        if procedure_ids:
            try:
                from modules.safety_procedures.module_safety_procedures import (
                    build_procedure_document_blueprint,
                    seed_default_procedures,
                )

                seed_default_procedures(folder.company_id or self._company_id())
                for procedure_id in procedure_ids:
                    blueprint = build_procedure_document_blueprint(procedure_id)
                    if blueprint:
                        procedure_blueprints.append(blueprint)
            except Exception:
                procedure_blueprints = []

        if not profile and not procedure_blueprints:
            return {"created": 0, "existing": 0}

        existing = self._documents_for_folder(folder.id)
        existing_codes = {str(document.code or "").strip().lower() for document in existing}
        created = 0
        reused = 0
        blueprints: List[Dict[str, Any]] = []
        blueprints.extend(procedure_blueprints)
        if profile:
            for item in _normalize_doc_blueprints(profile.mandatory_documents):
                if procedure_blueprints and item.get("document_type") == "procedure":
                    continue
                blueprints.append(item)

        for blueprint in blueprints:
            code = str(blueprint.get("code") or "").strip().lower()
            if code in existing_codes:
                reused += 1
                continue
            SafetyFolderDocument.create(
                {
                    "folder_id": folder.id,
                    "code": code,
                    "title": blueprint["title"],
                    "document_type": blueprint.get("document_type") or "other",
                    "status": "draft",
                    "version": 1,
                    "is_critical": _normalize_bool(blueprint.get("is_critical"), False),
                    "content": blueprint.get("content") or "",
                    "owner_user_id": self.env.user.id if self.env.user else None,
                    "company_id": folder.company_id,
                }
            )
            created += 1
            existing_codes.add(code)

        matrix = self._matrix_for_folder(folder.id)
        if not matrix:
            self._generate_matrix_for_folder(folder)
        return {"created": created, "existing": reused}

    def _build_summary(self, folder: SafetyFolder) -> Dict[str, Any]:
        documents = self._documents_for_folder(folder.id)
        matrix = self._matrix_for_folder(folder.id)
        deliveries = self._deliveries_for_folder(folder.id)
        talks = self._talks_for_folder(folder.id)
        checklists = self._checklists_for_folder(folder.id)

        critical_documents = [document for document in documents if document.is_critical]
        approved_documents = [document for document in critical_documents if document.status == "approved"]

        assigned_ids = _normalize_int_list(folder.assigned_employee_ids)
        delivered_ids = {
            delivery.employee_id
            for delivery in deliveries
            if delivery.status in ("delivered", "replenishment") and delivery.employee_id
        }
        assigned_with_ppe = len([emp_id for emp_id in assigned_ids if emp_id in delivered_ids])
        outstanding_ppe = len([emp_id for emp_id in assigned_ids if emp_id not in delivered_ids])

        checklist_ok = any(checklist.result == "ok" for checklist in checklists)
        matrix_rows = matrix.to_dict().get("rows", []) if matrix else []
        intolerable_rows = [
            row for row in matrix_rows if str(row.get("risk_level") or "") == "Intolerable"
        ]
        restricted_rows = [
            row for row in matrix_rows if _normalize_str_list(row.get("restriction_alerts"))
        ]
        matrix_ok = bool(
            matrix
            and matrix.status == "approved"
            and matrix.rows
            and not intolerable_rows
            and not restricted_rows
        )

        readiness = 0.0
        readiness += 10.0 if folder.planned_start_date else 0.0
        readiness += 10.0 if assigned_ids else 0.0
        readiness += 40.0 * (
            len(approved_documents) / len(critical_documents) if critical_documents else 1.0
        )
        readiness += 20.0 if matrix_ok else 0.0
        if assigned_ids:
            readiness += 10.0 * (assigned_with_ppe / len(assigned_ids))
        readiness += 10.0 if checklist_ok else 0.0

        blockers: List[str] = []
        if not folder.planned_start_date:
            blockers.append("Definir fecha objetivo de arranque")
        if not assigned_ids:
            blockers.append("Asignar personal a la carpeta")
        if critical_documents and len(approved_documents) < len(critical_documents):
            blockers.append("Aprobar documentos criticos")
        if not matrix_ok:
            blockers.append("Aprobar matriz de riesgo")
        if intolerable_rows:
            blockers.append("Existen riesgos intolerables (VEP 16) que bloquean el inicio")
        if restricted_rows:
            blockers.append("Existen alertas por restricciones medicas o sensibilidad especial")
        if assigned_ids and outstanding_ppe > 0:
            blockers.append("Completar entrega de EPP al personal asignado")
        if not checklist_ok:
            blockers.append("Registrar al menos un checklist conforme")

        if readiness >= 85 and not blockers:
            traffic_light = "green"
        elif readiness >= 50:
            traffic_light = "yellow"
        else:
            traffic_light = "red"

        return {
            "readiness_pct": round(readiness, 1),
            "traffic_light": traffic_light,
            "critical_documents_total": len(critical_documents),
            "critical_documents_approved": len(approved_documents),
            "documents_total": len(documents),
            "documents_approved": len([document for document in documents if document.status == "approved"]),
            "assigned_employee_count": len(assigned_ids),
            "employees_with_ppe": assigned_with_ppe,
            "talks_count": len(talks),
            "checklists_count": len(checklists),
            "has_matrix": bool(matrix),
            "matrix_status": matrix.status if matrix else "missing",
            "intolerable_rows": len(intolerable_rows),
            "restriction_alert_rows": len(restricted_rows),
            "legal_snapshot": self._legal_snapshot(folder),
            "critical_blockers": blockers,
        }

    def _refresh_folder_metrics(self, folder: SafetyFolder) -> Dict[str, Any]:
        summary = self._build_summary(folder)
        folder.readiness_pct = summary["readiness_pct"]
        folder.traffic_light = summary["traffic_light"]
        if folder.status == "ready" and summary["traffic_light"] != "green":
            folder.status = "draft"
            folder.approver_user_id = None
            folder.approved_at = None
        folder.save()
        return summary

    def _folder_payload(self, folder: SafetyFolder) -> Dict[str, Any]:
        summary = self._refresh_folder_metrics(folder)
        return {
            **folder.to_dict(),
            "summary": summary,
        }

    def _export_filename(self, folder: SafetyFolder, extension: str) -> str:
        base = folder.to_dict().get("project_code") or f"folder_{folder.id}"
        base = re.sub(r"[^A-Za-z0-9._-]+", "_", str(base)).strip("_") or f"folder_{folder.id}"
        return f"MIPER_{base}.{extension}"

    def _export_palette(self, traffic_light: str = "red") -> Dict[str, str]:
        traffic = {
            "red": {"fill": "#FEE2E2", "text": "#991B1B", "label": "Rojo"},
            "yellow": {"fill": "#FEF3C7", "text": "#92400E", "label": "Amarillo"},
            "green": {"fill": "#DCFCE7", "text": "#166534", "label": "Verde"},
        }.get(traffic_light or "red", {"fill": "#E2E8F0", "text": "#334155", "label": "Sin definir"})
        return {
            "brand": "#0F172A",
            "brand_alt": "#1D4ED8",
            "brand_soft": "#DBEAFE",
            "surface": "#F8FAFC",
            "surface_alt": "#E2E8F0",
            "ink": "#0F172A",
            "muted": "#475569",
            "line": "#CBD5E1",
            "success": "#DCFCE7",
            "warning": "#FEF3C7",
            "danger": "#FEE2E2",
            "traffic_fill": traffic["fill"],
            "traffic_text": traffic["text"],
            "traffic_label": traffic["label"],
        }

    def _format_export_timestamp(self, value: Optional[datetime] = None) -> str:
        stamp = value or utc_now()
        return stamp.strftime("%d-%m-%Y %H:%M UTC")

    def _build_export_insights(
        self,
        folder_payload: Dict[str, Any],
        matrix_payload: Dict[str, Any],
        rows: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        summary = folder_payload.get("summary") or {}
        generation = matrix_payload.get("generation_summary") or {}
        risk_counts = {
            "No aceptable": 0,
            "Intolerable": 0,
            "Importante": 0,
            "Moderado": 0,
            "Aceptable": 0,
            "Tolerable": 0,
        }
        protocols: Set[str] = set()
        origins: Dict[str, int] = {}
        blocking_rows: List[Dict[str, Any]] = []
        restriction_rows: List[Dict[str, Any]] = []

        for row in rows:
            residual_value = _safe_int(row.get("residual_risk_value") or row.get("vep"), 0) or 0
            level = str(
                row.get("residual_risk_label")
                or row.get("risk_level")
                or row.get("risk_level_label")
                or risk_level_from_vep(residual_value)
            ).strip() or "Aceptable"
            risk_counts[level] = risk_counts.get(level, 0) + 1
            for protocol in _normalize_str_list(row.get("protocol_codes") or row.get("protocols_summary")):
                protocols.add(protocol)
            origin_labels = (
                _normalize_str_list(row.get("source_labels"))
                or _normalize_str_list(row.get("source_titles"))
                or _normalize_str_list(row.get("origin_blocks"))
            )
            for origin in origin_labels:
                origins[origin] = origins.get(origin, 0) + 1

            alerts = _normalize_str_list(row.get("restriction_alerts"))
            is_blocking = (
                _normalize_bool(row.get("is_blocking"), False)
                or _normalize_bool(row.get("compact_approval_blocked"), False)
                or level in ("No aceptable", "Intolerable")
                or residual_value >= 28
            )
            if is_blocking or alerts:
                blocker_payload = {
                    "process_name": row.get("process_name") or "-",
                    "task_name": row.get("task_name") or row.get("activity") or "-",
                    "risk_name": row.get("risk_name") or row.get("risk") or "-",
                    "vep": residual_value,
                    "risk_level": level,
                    "action_required": row.get("compact_action_required") or row.get("action_required") or action_from_vep(residual_value),
                    "alerts": alerts,
                    "origins": origin_labels,
                }
                blocking_rows.append(blocker_payload)
                if alerts:
                    restriction_rows.append(blocker_payload)

        blocking_rows.sort(key=lambda item: (-int(item.get("vep") or 0), item.get("process_name") or ""))
        restriction_rows.sort(key=lambda item: (-int(item.get("vep") or 0), item.get("process_name") or ""))

        return {
            "generated_at": self._format_export_timestamp(),
            "row_count": len(rows),
            "protocols": sorted(protocols),
            "origins": origins,
            "risk_counts": risk_counts,
            "blocking_rows": blocking_rows,
            "restriction_rows": restriction_rows,
            "critical_blockers": summary.get("critical_blockers") or [],
            "legal_snapshot": summary.get("legal_snapshot") or generation.get("legal_snapshot") or {},
            "matched_rule_count": generation.get("matched_rule_count", 0),
            "matched_by_block": generation.get("matched_by_block") or {},
            "scope": generation.get("scope") or {},
            "source_counts": generation.get("source_counts") or {},
            "scope_note": folder_payload.get("miper_scope_notes") or "",
            "folder_note": folder_payload.get("notes") or "",
        }

    def _miper_export_context(self, folder: SafetyFolder) -> Dict[str, Any]:
        matrix = self._matrix_for_folder(folder.id)
        if not matrix:
            self._generate_matrix_for_folder(folder)
            matrix = self._matrix_for_folder(folder.id)
        payload = self._folder_payload(folder)
        if not matrix:
            return {"folder": payload, "matrix": None, "rows": []}
        matrix_payload = matrix.to_dict()
        rows = sorted(
            matrix_payload.get("rows", []),
            key=lambda row: (
                row.get("source_sort") or [99, 0, 0],
                -(_safe_int(row.get("residual_risk_value") or row.get("vep"), 0) or 0),
                str(row.get("source_title") or ""),
                str(row.get("process_name") or row.get("task_name") or ""),
                str(row.get("risk_name") or row.get("risk") or ""),
            ),
        )
        matrix_payload["row_count"] = len(rows)
        return {
            "folder": payload,
            "matrix": matrix_payload,
            "rows": rows,
            "insights": self._build_export_insights(payload, matrix_payload, rows),
        }

    def _row_color(self, vep: Any, level: str) -> str:
        try:
            vep_value = int(vep or 0)
        except (TypeError, ValueError):
            vep_value = 0
        if level == "Aceptable":
            return "#DCFCE7"
        if level == "Moderado":
            return "#FEF3C7"
        if level == "Importante":
            return "#FFEDD5"
        if level in ("No aceptable", "Intolerable") or vep_value >= 28:
            return "#FEE2E2"
        if level in ("Importante", "Alto") or vep_value >= 19:
            return "#FFEDD5"
        if level in ("Moderado", "Medium") or vep_value >= 10:
            return "#FEF3C7"
        if level in ("Tolerable", "Medium") or vep_value >= 4:
            return "#DBEAFE"
        return "#DCFCE7"

    def _matrix_export_rows(self, rows: List[Dict[str, Any]]) -> List[List[Any]]:
        table_rows: List[List[Any]] = []
        for row in rows:
            table_rows.append(
                [
                    row.get("process_name") or "",
                    row.get("task_name") or row.get("activity") or "",
                    row.get("position_name") or "",
                    row.get("place_name") or "",
                    row.get("hazard_factor") or row.get("hazard") or "",
                    " - ".join([str(value) for value in [row.get("master_risk_code"), row.get("risk_name") or row.get("risk")] if value]),
                    _safe_int(row.get("vep"), 0) or 0,
                    row.get("risk_level") or "",
                    row.get("controls") or "",
                    ", ".join(_normalize_str_list(row.get("required_ppe"))),
                    ", ".join(_normalize_str_list(row.get("protocol_codes"))),
                    ", ".join(
                        _normalize_str_list(row.get("source_labels"))
                        or _normalize_str_list(row.get("source_titles"))
                        or _normalize_str_list(row.get("origin_blocks"))
                    ),
                    ", ".join(_normalize_str_list(row.get("restriction_alerts"))),
                    row.get("owner_name") or "",
                    row.get("legal_reference") or "",
                ]
            )
        return table_rows

    def _compact_label_color(self, row: Dict[str, Any]) -> str:
        return self._row_color(
            row.get("residual_risk_value") or row.get("vep"),
            row.get("residual_risk_label") or row.get("risk_level_label") or row.get("risk_level") or "",
        )

    def _row_origin_label(self, row: Dict[str, Any]) -> str:
        return ", ".join(
            _normalize_str_list(row.get("source_labels"))
            or _normalize_str_list(row.get("source_titles"))
            or _normalize_str_list(row.get("origin_blocks"))
        )

    def _build_miper_excel_extensive(self, folder: SafetyFolder, context: Dict[str, Any]) -> bytes:
        matrix = context.get("matrix") or {}
        rows = context.get("rows") or []
        payload = context.get("folder") or {}
        insights = context.get("insights") or {}
        palette = self._export_palette(payload.get("traffic_light") or "red")
        workbook = Workbook()
        ws = workbook.active
        ws.title = "Matriz IPER"
        ws_criteria = workbook.create_sheet("Criterios")
        ws_catalog = workbook.create_sheet("Catalogo de peligros")
        ws_changes = workbook.create_sheet("Control de modificaciones")

        def solid(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color.lstrip("#"))

        border = Border(
            left=Side(style="thin", color=palette["line"].lstrip("#")),
            right=Side(style="thin", color=palette["line"].lstrip("#")),
            top=Side(style="thin", color=palette["line"].lstrip("#")),
            bottom=Side(style="thin", color=palette["line"].lstrip("#")),
        )
        title_fill = solid("#111827")
        group_fill = solid("#DDE7F2")
        header_fill = solid("#243244")
        light_fill = solid("#F8FAFC")
        muted_fill = solid("#E5E7EB")
        doc_brand = solid("#143A6D")
        doc_brand_mid = solid("#1E4B82")
        doc_logo_fill = solid("#F8FAFC")

        headers = [
            "N",
            "Actividad",
            "Tarea",
            "Tipo R/NR/E",
            "Puesto",
            "Lugar",
            "Peligro / factor",
            "Tipo peligro",
            "Riesgo / incidente",
            "Dano probable",
            "Ingenieria actual",
            "Senalizacion / administrativos actuales",
            "EPP actual",
            "PE",
            "FE",
            "FO",
            "P",
            "Severidad",
            "Valor riesgo",
            "Clasificacion residual",
            "Eliminacion",
            "Sustitucion",
            "Ingenieria propuesta",
            "Administrativos / formacion / senalizacion",
            "EPP propuesto",
            "Plan S&SO",
            "Responsable",
            "Ref. legal",
            "Protocolos",
            "Origen / observaciones",
        ]
        last_col = get_column_letter(len(headers))
        ws.sheet_view.showGridLines = False

        def first_value(*values: Any, default: str = "-") -> str:
            for value in values:
                if value is None:
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                return str(value)
            return default

        def reviewer_name(value: Any, fallback: str = "Pendiente") -> str:
            if not value:
                return fallback
            if isinstance(value, int) or str(value).isdigit():
                return f"Usuario #{value}"
            return str(value)

        service_title = matrix.get("title") or payload.get("lead_title") or "Servicio operacional"
        work_center = matrix.get("work_center") or payload.get("client_site_name") or "-"
        area_name = payload.get("client_area_name") or "-"
        origin_label = matrix.get("source_type") or payload.get("service_profile_name") or payload.get("service_type_name") or "-"
        review_value = matrix.get("review_due_date") or payload.get("planned_start_date") or "-"
        customer_name = first_value(payload.get("customer_name"), payload.get("company_name"), "Empresa")
        updated_by = first_value(payload.get("updated_by_name"), payload.get("owner_name"), "Prevencion de Riesgos")
        elaborated_by = first_value(payload.get("created_by_name"), updated_by, "Prevencion de Riesgos")
        reviewed_by = reviewer_name(matrix.get("reviewed_by") or matrix.get("reviewed_by_name"))
        approved_by = first_value(matrix.get("approved_by_name"), matrix.get("approved_by"), "Pendiente")
        updated_at = first_value(matrix.get("last_update_date"), matrix.get("reviewed_at"), utc_today_iso())
        version = first_value(matrix.get("version"), "1")
        status = first_value(matrix.get("status"), "draft")

        def style_doc_cell(row_idx: int, start_col: int, end_col: int, value: Any, *, label: bool = False) -> None:
            ws.merge_cells(start_row=row_idx, start_column=start_col, end_row=row_idx, end_column=end_col)
            cell = ws.cell(row=row_idx, column=start_col, value=value or "-")
            cell.fill = doc_brand if label else doc_brand_mid
            cell.font = Font(size=9, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for merged_col in range(start_col, end_col + 1):
                ws.cell(row=row_idx, column=merged_col).border = border
                ws.cell(row=row_idx, column=merged_col).fill = doc_brand if label else doc_brand_mid

        ws.merge_cells("A1:D3")
        logo_cell = ws["A1"]
        logo_cell.value = f"{customer_name}\nMatriz MIPER/IPER"
        logo_cell.fill = doc_logo_fill
        logo_cell.font = Font(size=13, bold=True, color="143A6D")
        logo_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_idx in range(1, 4):
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).border = border
                ws.cell(row=row_idx, column=col_idx).fill = doc_logo_fill

        ws.merge_cells(start_row=1, start_column=5, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=5, value="MATRIZ DE IDENTIFICACION DE PELIGROS, EVALUACION DE RIESGOS")
        ws.cell(row=1, column=5).fill = doc_brand
        ws.cell(row=1, column=5).font = Font(bold=True, size=14, color="FFFFFF")
        ws.cell(row=1, column=5).alignment = Alignment(horizontal="center", vertical="center")
        for col_idx in range(5, len(headers) + 1):
            ws.cell(row=1, column=col_idx).border = border
            ws.cell(row=1, column=col_idx).fill = doc_brand

        style_doc_cell(2, 5, 6, "DIVISION / CLIENTE", label=True)
        style_doc_cell(2, 7, 10, customer_name)
        style_doc_cell(2, 11, 12, "LOCALIDAD", label=True)
        style_doc_cell(2, 13, 15, work_center)
        style_doc_cell(2, 16, 17, "ELABORADO POR", label=True)
        style_doc_cell(2, 18, 21, elaborated_by)
        style_doc_cell(2, 22, 24, "FECHA ACTUALIZACION", label=True)
        style_doc_cell(2, 25, 27, updated_at)
        style_doc_cell(2, 28, 29, "VERSION N", label=True)
        style_doc_cell(2, 30, 30, version)

        style_doc_cell(3, 5, 6, "SERVICIO / AREA", label=True)
        style_doc_cell(3, 7, 12, f"{service_title} | {area_name}")
        style_doc_cell(3, 13, 14, "REVISADO POR", label=True)
        style_doc_cell(3, 15, 18, reviewed_by)
        style_doc_cell(3, 19, 20, "AUTORIZADO POR", label=True)
        style_doc_cell(3, 21, 24, approved_by)
        style_doc_cell(3, 25, 26, "ESTADO / FILAS", label=True)
        style_doc_cell(3, 27, 28, f"{status} / {len(rows)}")
        style_doc_cell(3, 29, 29, "CODIGO", label=True)
        style_doc_cell(3, 30, 30, matrix.get("code") or payload.get("project_code") or folder.id)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 30

        groups = [
            (1, 6, "I. Caracteristica de la actividad"),
            (7, 10, "II. Identificacion de peligros"),
            (11, 13, "III. Controles actuales"),
            (14, 20, "IV. Evaluacion del riesgo"),
            (21, 25, "V. Controles propuestos"),
            (26, 30, "VI. Plan de gestion S&SO"),
        ]
        for start, end, label in groups:
            ws.merge_cells(start_row=4, start_column=start, end_row=4, end_column=end)
            cell = ws.cell(row=4, column=start, value=label)
            cell.fill = group_fill
            cell.font = Font(bold=True, color="0F172A", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[4]:
            cell.border = border
            if cell.value is None:
                cell.fill = group_fill
        ws.row_dimensions[4].height = 22

        for column_idx, header in enumerate(headers, start=1):
            ws.cell(row=5, column=column_idx, value=header)
        for cell in ws[5]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[5].height = 34

        for index, row in enumerate(rows, start=1):
            controls = normalize_control_hierarchy(row.get("required_controls") or row.get("control_hierarchy"))
            ws.append(
                [
                    index,
                    row.get("process_name") or row.get("activity_name") or row.get("activity") or "",
                    row.get("task_name") or row.get("activity") or "",
                    row.get("task_type_code") or "R",
                    row.get("job_position") or row.get("position_name") or "",
                    row.get("specific_workplace") or row.get("place_name") or "",
                    row.get("hazard_name") or row.get("hazard_factor") or row.get("hazard") or "",
                    row.get("risk_family") or "",
                    row.get("risk_name") or row.get("risk") or "",
                    row.get("probable_damage") or "",
                    row.get("current_engineering_controls") or "",
                    row.get("current_admin_controls") or row.get("controls") or row.get("existing_controls") or "",
                    row.get("current_ppe_controls") or ", ".join(_normalize_str_list(row.get("required_ppe") or row.get("ppe_summary"))),
                    _safe_int(row.get("exposed_people_value"), 1) or 1,
                    _safe_int(row.get("exposure_frequency_value"), 1) or 1,
                    _safe_int(row.get("occurrence_factor_value"), 1) or 1,
                    _safe_int(row.get("probability_score"), 0) or 0,
                    _safe_int(row.get("severity_value"), 0) or 0,
                    _safe_int(row.get("residual_risk_value") or row.get("vep"), 0) or 0,
                    row.get("residual_risk_label") or row.get("risk_level_label") or row.get("risk_level") or "",
                    "; ".join(_normalize_str_list(row.get("proposed_elimination_controls") or controls.get("elimination"))),
                    "; ".join(_normalize_str_list(row.get("proposed_substitution_controls") or controls.get("substitution"))),
                    "; ".join(_normalize_str_list(row.get("proposed_engineering_controls") or controls.get("engineering"))),
                    "; ".join(_normalize_str_list(row.get("proposed_admin_controls") or controls.get("administrative"))),
                    "; ".join(_normalize_str_list(row.get("proposed_ppe_controls") or controls.get("ppe"))),
                    row.get("safety_management_plan") or row.get("compact_action_required") or row.get("action_required") or "",
                    row.get("responsible") or row.get("owner_name") or "",
                    row.get("legal_reference") or "",
                    ", ".join(_normalize_str_list(row.get("protocols_summary") or row.get("protocol_codes"))),
                    " | ".join(
                        [
                            value
                            for value in [
                                self._row_origin_label(row),
                                row.get("observations") or row.get("source_note") or "",
                            ]
                            if value
                        ]
                    ),
                ]
            )

        for row_idx in range(6, ws.max_row + 1):
            source_row = rows[row_idx - 6] if row_idx - 6 < len(rows) else {}
            residual_fill = solid(self._compact_label_color(source_row))
            for cell in ws[row_idx]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=8.5, color="111827")
                cell.fill = light_fill
            for col_idx in range(14, 20):
                ws.cell(row=row_idx, column=col_idx).fill = solid("#FEF7E6")
            ws.cell(row=row_idx, column=20).fill = residual_fill
            ws.cell(row=row_idx, column=20).font = Font(size=8, bold=True, color="111827")

        ws.freeze_panes = "A6"
        ws.auto_filter.ref = f"A5:{last_col}{ws.max_row}"
        widths = [6, 20, 24, 12, 18, 18, 26, 14, 26, 22, 24, 30, 20, 7, 7, 7, 7, 10, 11, 18, 20, 20, 23, 28, 20, 28, 18, 22, 20, 28]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.sheet_view.zoomScale = 75
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = "$1:$5"

        methodology = compact_miper_methodology_payload()
        ws_criteria.sheet_view.showGridLines = False
        ws_criteria.append(["Criterio", "Valor/Codigo", "Etiqueta", "Descripcion"])
        for schema_name, label in [
            ("task_type_schema", "Tipo tarea"),
            ("exposed_people_schema", "PE"),
            ("exposure_frequency_schema", "FE"),
            ("occurrence_factor_schema", "FO"),
            ("severity_schema", "Severidad"),
            ("residual_risk_schema", "Clasificacion residual"),
        ]:
            for item in methodology.get(schema_name, []):
                ws_criteria.append(
                    [
                        label,
                        item.get("code") or item.get("value") or f"{item.get('min')}-{item.get('max')}",
                        item.get("label") or "",
                        item.get("description") or item.get("action_required") or "",
                    ]
                )
        for row in ws_criteria.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = header_fill if cell.row == 1 else light_fill
                cell.font = Font(bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "111827", size=9)
        for idx, width in enumerate([24, 14, 22, 70], start=1):
            ws_criteria.column_dimensions[get_column_letter(idx)].width = width

        ws_catalog.sheet_view.showGridLines = False
        ws_catalog.append(["Peligro / factor", "Riesgo / incidente", "Dano probable", "Tipo peligro", "Protocolos"])
        seen_catalog: Set[Tuple[str, str]] = set()
        for row in rows:
            key = (
                str(row.get("hazard_name") or row.get("hazard_factor") or row.get("hazard") or ""),
                str(row.get("risk_name") or row.get("risk") or ""),
            )
            if key in seen_catalog:
                continue
            seen_catalog.add(key)
            ws_catalog.append(
                [
                    key[0],
                    key[1],
                    row.get("probable_damage") or "",
                    row.get("risk_family") or "",
                    ", ".join(_normalize_str_list(row.get("protocols_summary") or row.get("protocol_codes"))),
                ]
            )
        for row in ws_catalog.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = header_fill if cell.row == 1 else light_fill
                cell.font = Font(bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "111827", size=9)
        for idx, width in enumerate([34, 34, 34, 18, 24], start=1):
            ws_catalog.column_dimensions[get_column_letter(idx)].width = width

        ws_changes.sheet_view.showGridLines = False
        ws_changes.append(["Version", "Fecha", "Descripcion", "Responsable"])
        ws_changes.append([matrix.get("version") or 1, matrix.get("last_update_date") or utc_today_iso(), "Emision desde ERP con metodologia compacta PE+FE+FO x S.", matrix.get("reviewed_by") or ""])
        for row in ws_changes.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = header_fill if cell.row == 1 else muted_fill
                cell.font = Font(bold=cell.row == 1, color="FFFFFF" if cell.row == 1 else "111827", size=9)
        for idx, width in enumerate([12, 18, 70, 24], start=1):
            ws_changes.column_dimensions[get_column_letter(idx)].width = width

        bio = BytesIO()
        workbook.save(bio)
        return bio.getvalue()

    def _build_miper_pdf_compact(self, folder: SafetyFolder, context: Dict[str, Any]) -> bytes:
        matrix = context.get("matrix") or {}
        rows = context.get("rows") or []
        payload = context.get("folder") or {}
        insights = context.get("insights") or {}
        summary = payload.get("summary") or {}
        legal_snapshot = insights.get("legal_snapshot") or summary.get("legal_snapshot") or {}
        palette = self._export_palette(payload.get("traffic_light") or "red")
        brand = "#143A6D"
        brand_mid = "#1E4B82"
        brand_line = "#5D7EA8"
        table_line = "#AEBBCA"
        surface = "#F5F8FC"
        surface_alt = "#EAF1F8"
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A3),
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=16 * mm,
            bottomMargin=12 * mm,
        )

        def safe_html(value: Any) -> str:
            return (
                str(value or "")
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )

        def draw_page(canvas, pdf_doc):
            width, height = pdf_doc.pagesize
            canvas.saveState()
            canvas.setFillColor(colors.HexColor(brand))
            canvas.rect(0, height - 13 * mm, width, 13 * mm, fill=1, stroke=0)
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.drawString(pdf_doc.leftMargin, height - 8.2 * mm, f"MIPER/IPER | {matrix.get('code') or payload.get('project_code') or folder.id}")
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(
                width - pdf_doc.rightMargin,
                height - 8.2 * mm,
                f"{insights.get('generated_at') or self._format_export_timestamp()} | Pag. {canvas.getPageNumber()}",
            )
            canvas.setStrokeColor(colors.HexColor(palette["line"]))
            canvas.line(pdf_doc.leftMargin, 8 * mm, width - pdf_doc.rightMargin, 8 * mm)
            canvas.setFillColor(colors.HexColor(palette["muted"]))
            canvas.setFont("Helvetica", 7.4)
            canvas.drawString(
                pdf_doc.leftMargin,
                5 * mm,
                f"{payload.get('customer_name') or '-'} | {payload.get('service_profile_name') or payload.get('service_type_name') or matrix.get('source_type') or '-'}",
            )
            canvas.restoreState()

        def first_value(*values: Any, default: str = "-") -> str:
            for value in values:
                if value is None:
                    continue
                if isinstance(value, list):
                    text = ", ".join(_normalize_str_list(value))
                    if text:
                        return text
                    continue
                if isinstance(value, str) and not value.strip():
                    continue
                return str(value)
            return default

        def join_nonempty(values: List[Any], separator: str = " | ") -> str:
            normalized = []
            for value in values:
                if isinstance(value, list):
                    text = ", ".join(_normalize_str_list(value))
                else:
                    text = str(value or "").strip()
                if text:
                    normalized.append(text)
            return separator.join(normalized)

        def cell(value: Any, style_name: str = "FormalCell") -> Paragraph:
            return Paragraph(safe_html(value if value is not None else "-") or "-", styles[style_name])

        def compact_evaluation(row: Dict[str, Any]) -> str:
            pe = _safe_int(row.get("exposed_people_value"), 1) or 1
            fe = _safe_int(row.get("exposure_frequency_value"), 1) or 1
            fo = _safe_int(row.get("occurrence_factor_value"), 1) or 1
            probability = _safe_int(row.get("probability_score"), None)
            if probability is None:
                probability = pe + fe + fo
            severity = _safe_int(row.get("severity_value"), None)
            if severity is None:
                severity = _safe_int(row.get("consequence_value"), 1) or 1
            vr = _safe_int(row.get("residual_risk_value"), None)
            if vr is None:
                vr = _safe_int(row.get("vep") or row.get("risk_value"), 0) or 0
            return f"PE {pe} | FE {fe} | FO {fo} | P {probability} | S {severity} | VR {vr}"

        def proposed_controls(row: Dict[str, Any]) -> str:
            parts = [
                ("Elim.", row.get("proposed_elimination_controls") or row.get("elimination_controls")),
                ("Sust.", row.get("proposed_substitution_controls") or row.get("substitution_controls")),
                ("Ing.", row.get("proposed_engineering_controls") or row.get("engineering_controls")),
                ("Adm.", row.get("proposed_admin_controls") or row.get("administrative_controls")),
                ("EPP", row.get("proposed_ppe_controls") or row.get("ppe_controls")),
            ]
            normalized = []
            for label, value in parts:
                text = ", ".join(_normalize_str_list(value)) if isinstance(value, list) else str(value or "").strip()
                if text:
                    normalized.append(f"{label}: {text}")
            return " | ".join(normalized) or "-"

        def reviewer_name(value: Any, fallback: str = "Pendiente") -> str:
            if not value:
                return fallback
            if isinstance(value, int) or str(value).isdigit():
                return f"Usuario #{value}"
            return str(value)

        doc_code = first_value(matrix.get("code"), payload.get("project_code"), folder.id)
        doc_title = first_value(matrix.get("title"), payload.get("lead_title"), "Matriz MIPER/IPER")
        service_title = first_value(payload.get("service_profile_name"), payload.get("service_type_name"), matrix.get("source_type"))
        customer_name = first_value(payload.get("customer_name"), payload.get("company_name"), "Empresa")
        work_center = first_value(matrix.get("work_center"), payload.get("client_site_name"), payload.get("place_name"))
        area_name = first_value(payload.get("client_area_name"), matrix.get("process_name"))
        updated_by = first_value(payload.get("updated_by_name"), payload.get("owner_name"), "Prevencion de Riesgos")
        elaborated_by = first_value(payload.get("created_by_name"), updated_by, "Prevencion de Riesgos")
        reviewed_by = reviewer_name(matrix.get("reviewed_by") or matrix.get("reviewed_by_name"))
        approved_by = first_value(matrix.get("approved_by_name"), matrix.get("approved_by"), "Aprobacion registrada" if matrix.get("status") == "approved" else "Pendiente")
        updated_at = first_value(matrix.get("last_update_date"), matrix.get("reviewed_at"), utc_today_iso())
        version = first_value(matrix.get("version"), "1")
        status = first_value(matrix.get("status"), "draft")

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="HeaderTitle", fontName="Helvetica-Bold", fontSize=13, leading=15, textColor=colors.white, alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="HeaderCell", fontName="Helvetica-Bold", fontSize=8.2, leading=9.5, textColor=colors.white, alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="HeaderValue", fontName="Helvetica-Bold", fontSize=8.4, leading=9.8, textColor=colors.white, alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="BrandBlock", fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=colors.HexColor(brand), alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="BrandSmall", fontName="Helvetica", fontSize=7.4, leading=8.6, textColor=colors.HexColor("#475569"), alignment=TA_CENTER))
        styles.add(ParagraphStyle(name="FormalSection", fontName="Helvetica-Bold", fontSize=9.5, leading=11, textColor=colors.HexColor(brand)))
        styles.add(ParagraphStyle(name="FormalMeta", fontName="Helvetica", fontSize=8.2, leading=9.6, textColor=colors.HexColor("#111827")))
        styles.add(ParagraphStyle(name="FormalCell", fontName="Helvetica", fontSize=6.35, leading=7.35, textColor=colors.black))
        styles.add(ParagraphStyle(name="FormalCellBold", fontName="Helvetica-Bold", fontSize=6.45, leading=7.4, textColor=colors.black))
        styles.add(ParagraphStyle(name="FormalHead", fontName="Helvetica-Bold", fontSize=6.9, leading=7.8, textColor=colors.white, alignment=TA_CENTER))

        story: List[Any] = [Spacer(1, 1.5 * mm)]
        header_data = [
            [
                Paragraph("MATRIZ DE IDENTIFICACION DE PELIGROS, EVALUACION DE RIESGOS", styles["HeaderTitle"]),
                "",
                "",
                "",
                "",
                "",
            ],
            [
                Paragraph("DIVISION / CLIENTE", styles["HeaderCell"]),
                Paragraph(safe_html(customer_name), styles["HeaderValue"]),
                Paragraph("ACTUALIZADO POR", styles["HeaderCell"]),
                Paragraph(safe_html(updated_by), styles["HeaderValue"]),
                Paragraph("FECHA ACTUALIZACION", styles["HeaderCell"]),
                Paragraph(safe_html(updated_at), styles["HeaderValue"]),
            ],
            [
                Paragraph("LOCALIDAD / FAENA", styles["HeaderCell"]),
                Paragraph(safe_html(work_center), styles["HeaderValue"]),
                Paragraph("REVISADO POR", styles["HeaderCell"]),
                Paragraph(safe_html(reviewed_by), styles["HeaderValue"]),
                Paragraph("VERSION N", styles["HeaderCell"]),
                Paragraph(safe_html(version), styles["HeaderValue"]),
            ],
            [
                Paragraph("CODIGO MATRIZ", styles["HeaderCell"]),
                Paragraph(safe_html(doc_code), styles["HeaderValue"]),
                Paragraph("APROBADO POR", styles["HeaderCell"]),
                Paragraph(safe_html(approved_by), styles["HeaderValue"]),
                Paragraph("ESTADO", styles["HeaderCell"]),
                Paragraph(safe_html(status), styles["HeaderValue"]),
            ],
            [
                Paragraph("SERVICIO / AREA", styles["HeaderCell"]),
                Paragraph(safe_html(f"{service_title} | {area_name}"), styles["HeaderValue"]),
                "",
                "",
                Paragraph("FILAS", styles["HeaderCell"]),
                Paragraph(safe_html(len(rows)), styles["HeaderValue"]),
            ],
        ]
        header_table = Table(header_data, colWidths=[38 * mm, 58 * mm, 38 * mm, 72 * mm, 42 * mm, 98 * mm])
        header_table.setStyle(
            TableStyle(
                [
                    ("SPAN", (0, 0), (-1, 0)),
                    ("SPAN", (1, 4), (3, 4)),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(brand)),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(brand_mid)),
                    ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor(brand_line)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )

        brand_block = Table(
            [
                [Paragraph(safe_html(customer_name), styles["BrandBlock"])],
                [Paragraph("Logo corporativo / cliente", styles["BrandSmall"])],
                [Paragraph(safe_html(service_title), styles["BrandSmall"])],
            ]
        )
        brand_block.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.45, colors.HexColor(brand_line)),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 7),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )

        outer_header = Table([[brand_block, header_table]], colWidths=[58 * mm, 346 * mm])
        outer_header.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        story.append(outer_header)
        story.append(Spacer(1, 2 * mm))

        signature_data = [
            [
                Paragraph(f"<b>Elaborado por</b><br/>{safe_html(elaborated_by)}", styles["FormalMeta"]),
                Paragraph(f"<b>Revisado por</b><br/>{safe_html(reviewed_by)}", styles["FormalMeta"]),
                Paragraph(f"<b>Aprobado por</b><br/>{safe_html(approved_by)}", styles["FormalMeta"]),
                Paragraph(f"<b>Fecha</b><br/>{safe_html(updated_at)}", styles["FormalMeta"]),
                Paragraph(f"<b>Version documento</b><br/>{safe_html(version)}", styles["FormalMeta"]),
            ]
        ]
        signature_table = Table(signature_data, colWidths=[80 * mm, 80 * mm, 80 * mm, 72 * mm, 92 * mm])
        signature_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(surface_alt)),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(table_line)),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor(table_line)),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(signature_table)
        story.append(Spacer(1, 2 * mm))
        glossary_table = Table(
            [
                [
                    Paragraph(
                        "<b>Glosa de evaluacion:</b> PE = personas expuestas | FE = frecuencia de exposicion | "
                        "FO = factor de ocurrencia | P = PE + FE + FO | S = severidad | VR = P x S. "
                        "La clasificacion residual se obtiene desde el VR: Aceptable, Moderado, Importante o No aceptable.",
                        styles["FormalMeta"],
                    )
                ]
            ],
            colWidths=[404 * mm],
        )
        glossary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF7ED")),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#FDBA74")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(glossary_table)
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph(f"{safe_html(doc_title)} | Metodologia: P = PE + FE + FO, VR = P x S | Segmento legal: {safe_html((legal_snapshot or {}).get('segment') or '-')}", styles["FormalSection"]))
        story.append(Spacer(1, 1.5 * mm))

        headers = [
            "N",
            "Actividad",
            "Tarea",
            "Tipo R/NR/E",
            "Puesto",
            "Lugar",
            "Peligro / factor",
            "Riesgo / incidente",
            "Dano probable",
            "Ing. actual",
            "Admin. / senalizacion",
            "EPP actual / requerido",
            "Evaluacion",
            "Clasificacion",
            "Controles propuestos",
            "Plan S&SO",
            "Responsable",
            "Ref. legal / protocolos",
        ]
        table_data: List[List[Any]] = [[Paragraph(item, styles["FormalHead"]) for item in headers]]
        for index, row in enumerate(rows, start=1):
            ppe_text = first_value(
                row.get("current_ppe_controls"),
                ", ".join(_normalize_str_list(row.get("required_ppe") or row.get("ppe_summary"))),
                row.get("proposed_ppe_controls"),
            )
            reference_text = join_nonempty(
                [
                    row.get("legal_reference"),
                    ", ".join(_normalize_str_list(row.get("protocols") or row.get("protocols_summary"))),
                    row.get("origin"),
                ]
            )
            table_data.append(
                [
                    str(index),
                    cell(first_value(row.get("process_name"), row.get("activity_name"), row.get("activity"), default="")),
                    cell(row.get("task_name") or ""),
                    cell(row.get("task_type_code") or "R", "FormalCellBold"),
                    cell(first_value(row.get("job_position"), row.get("position_name"), default="")),
                    cell(first_value(row.get("specific_workplace"), row.get("place_name"), work_center, default="")),
                    cell(first_value(row.get("hazard_name"), row.get("hazard_factor"), row.get("hazard"), default="")),
                    cell(first_value(row.get("risk_name"), row.get("risk"), default="")),
                    cell(row.get("probable_damage") or ""),
                    cell(row.get("current_engineering_controls") or ""),
                    cell(first_value(row.get("current_admin_controls"), row.get("controls"), row.get("existing_controls"), default="")),
                    cell(ppe_text),
                    cell(compact_evaluation(row), "FormalCellBold"),
                    cell(row.get("residual_risk_label") or row.get("risk_level_label") or "", "FormalCellBold"),
                    cell(proposed_controls(row)),
                    cell(first_value(row.get("safety_management_plan"), row.get("compact_action_required"), row.get("action_required"), default="")),
                    cell(first_value(row.get("responsible"), row.get("owner_name"), default="")),
                    cell(reference_text),
                ]
            )

        table = LongTable(
            table_data,
            repeatRows=1,
            colWidths=[
                6 * mm,
                20 * mm,
                22 * mm,
                8 * mm,
                17 * mm,
                16 * mm,
                30 * mm,
                30 * mm,
                22 * mm,
                24 * mm,
                29 * mm,
                22 * mm,
                24 * mm,
                19 * mm,
                34 * mm,
                28 * mm,
                18 * mm,
                24 * mm,
            ],
        )
        style_cmds: List[Tuple[Any, ...]] = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(brand)),
            ("GRID", (0, 0), (-1, -1), 0.24, colors.HexColor(table_line)),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 2.2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2.2),
            ("TOPPADDING", (0, 0), (-1, -1), 2.4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.4),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (3, 0), (3, -1), "CENTER"),
            ("ALIGN", (12, 0), (13, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]
        for idx, row in enumerate(rows, start=1):
            fill = colors.HexColor(self._compact_label_color(row))
            band = colors.HexColor(surface if idx % 2 else surface_alt)
            style_cmds.append(("BACKGROUND", (0, idx), (-1, idx), band))
            style_cmds.append(("BACKGROUND", (13, idx), (13, idx), fill))
            style_cmds.append(("FONTNAME", (12, idx), (13, idx), "Helvetica-Bold"))
        table.setStyle(TableStyle(style_cmds))
        story.append(table)
        doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
        return buffer.getvalue()

    def _build_miper_excel(self, folder: SafetyFolder, context: Dict[str, Any]) -> bytes:
        return self._build_miper_excel_extensive(folder, context)
        matrix = context.get("matrix") or {}
        rows = context.get("rows") or []
        payload = context.get("folder") or {}
        insights = context.get("insights") or {}
        summary = payload.get("summary") or {}
        legal_snapshot = insights.get("legal_snapshot") or summary.get("legal_snapshot") or {}
        palette = self._export_palette(payload.get("traffic_light") or "red")
        workbook = Workbook()
        ws_cover = workbook.active
        ws_cover.title = "Portada"
        ws_matrix = workbook.create_sheet("MIPER")
        ws_alerts = workbook.create_sheet("Alertas")
        ws_trace = workbook.create_sheet("Trazabilidad")

        def solid(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color.lstrip("#"))

        title_fill = solid(palette["brand_alt"])
        header_fill = solid(palette["brand"])
        light_fill = solid(palette["surface"])
        soft_fill = solid(palette["brand_soft"])
        traffic_fill = solid(palette["traffic_fill"])
        border = Border(
            left=Side(style="thin", color=palette["line"].lstrip("#")),
            right=Side(style="thin", color=palette["line"].lstrip("#")),
            top=Side(style="thin", color=palette["line"].lstrip("#")),
            bottom=Side(style="thin", color=palette["line"].lstrip("#")),
        )

        def style_range(worksheet, cell_range: str, fill: Optional[PatternFill] = None, font: Optional[Font] = None, alignment: Optional[Alignment] = None) -> None:
            for row in worksheet[cell_range]:
                for cell in row:
                    cell.border = border
                    if fill:
                        cell.fill = fill
                    if font:
                        cell.font = font
                    if alignment:
                        cell.alignment = alignment

        def write_meta_row(worksheet, row_idx: int, left_label: str, left_value: Any, right_label: str, right_value: Any) -> None:
            worksheet[f"A{row_idx}"] = left_label
            worksheet[f"B{row_idx}"] = str(left_value or "-")
            worksheet[f"E{row_idx}"] = right_label
            worksheet[f"F{row_idx}"] = str(right_value or "-")
            for cell_id in (f"A{row_idx}", f"E{row_idx}"):
                worksheet[cell_id].font = Font(bold=True, color=palette["ink"].lstrip("#"))
                worksheet[cell_id].fill = light_fill
                worksheet[cell_id].border = border
            for cell_id in (f"B{row_idx}", f"F{row_idx}"):
                worksheet[cell_id].fill = light_fill
                worksheet[cell_id].border = border
                worksheet[cell_id].alignment = Alignment(vertical="top", wrap_text=True)

        ws_cover.sheet_view.showGridLines = False
        ws_cover["A1"] = "MIPER - Matriz de Identificacion de Peligros y Evaluacion de Riesgos"
        ws_cover.merge_cells("A1:H2")
        style_range(
            ws_cover,
            "A1:H2",
            fill=title_fill,
            font=Font(bold=True, size=16, color="FFFFFF"),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws_cover["A3"] = (
            f"{payload.get('project_code') or folder.id} | {payload.get('lead_title') or 'Sin oportunidad'} | "
            f"{payload.get('customer_name') or '-'}"
        )
        ws_cover.merge_cells("A3:H3")
        style_range(
            ws_cover,
            "A3:H3",
            fill=soft_fill,
            font=Font(size=10, bold=True, color=palette["ink"].lstrip("#")),
            alignment=Alignment(horizontal="center", vertical="center"),
        )

        kpi_cards = [
            ("A5:B6", "Readiness", f"{float(payload.get('readiness_pct') or 0):.1f}%"),
            ("C5:D6", "Semaforo", palette["traffic_label"]),
            ("E5:F6", "Filas MIPER", str(insights.get("row_count", len(rows)))),
            ("G5:H6", "Bloqueos", str(len(insights.get("blocking_rows") or []))),
            ("A7:B8", "Protocolos", str(len(insights.get("protocols") or []))),
            ("C7:D8", "Reglas heredadas", str(insights.get("matched_rule_count", 0))),
            ("E7:F8", "Segmento legal", str((legal_snapshot or {}).get("segment") or "-")),
            ("G7:H8", "Emitido", str(insights.get("generated_at") or self._format_export_timestamp())),
        ]
        for cell_range, label, value in kpi_cards:
            ws_cover[cell_range.split(":")[0]] = f"{label}\n{value}"
            ws_cover.merge_cells(cell_range)
            style_range(
                ws_cover,
                cell_range,
                fill=traffic_fill if label == "Semaforo" else light_fill,
                font=Font(bold=True, size=11, color=palette["ink"].lstrip("#")),
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            )

        ws_cover["A10"] = "Identificacion y alcance"
        ws_cover.merge_cells("A10:H10")
        style_range(
            ws_cover,
            "A10:H10",
            fill=header_fill,
            font=Font(bold=True, size=11, color="FFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        write_meta_row(ws_cover, 11, "Proyecto", payload.get("project_code") or folder.id, "Estado", payload.get("status") or "draft")
        write_meta_row(ws_cover, 12, "Cliente", payload.get("customer_name") or "-", "Servicio", payload.get("service_profile_name") or payload.get("service_type_name") or "-")
        write_meta_row(ws_cover, 13, "Instalacion", payload.get("client_site_name") or "-", "Area", payload.get("client_area_name") or "-")
        write_meta_row(ws_cover, 14, "Arranque", payload.get("planned_start_date") or "-", "Dotacion", str((legal_snapshot or {}).get("worker_count") or 0))

        ws_cover["A16"] = "Notas de alcance"
        ws_cover.merge_cells("A16:H16")
        style_range(
            ws_cover,
            "A16:H16",
            fill=header_fill,
            font=Font(bold=True, size=11, color="FFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        ws_cover["A17"] = insights.get("scope_note") or insights.get("folder_note") or "Sin notas adicionales registradas."
        ws_cover.merge_cells("A17:H18")
        style_range(
            ws_cover,
            "A17:H18",
            fill=light_fill,
            font=Font(size=10, color=palette["ink"].lstrip("#")),
            alignment=Alignment(vertical="top", wrap_text=True),
        )

        ws_cover["A20"] = "Bloqueos criticos"
        ws_cover.merge_cells("A20:H20")
        style_range(
            ws_cover,
            "A20:H20",
            fill=header_fill,
            font=Font(bold=True, size=11, color="FFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        cover_row = 21
        blockers = insights.get("critical_blockers") or []
        if blockers:
            for blocker in blockers:
                ws_cover[f"A{cover_row}"] = blocker
                ws_cover.merge_cells(f"A{cover_row}:H{cover_row}")
                style_range(
                    ws_cover,
                    f"A{cover_row}:H{cover_row}",
                    fill=solid(palette["danger"]),
                    font=Font(size=10, color=palette["ink"].lstrip("#")),
                    alignment=Alignment(vertical="top", wrap_text=True),
                )
                cover_row += 1
        else:
            ws_cover[f"A{cover_row}"] = "No se registran bloqueos criticos al momento de la emision."
            ws_cover.merge_cells(f"A{cover_row}:H{cover_row}")
            style_range(
                ws_cover,
                f"A{cover_row}:H{cover_row}",
                fill=solid(palette["success"]),
                font=Font(size=10, color=palette["ink"].lstrip("#")),
                alignment=Alignment(vertical="top", wrap_text=True),
            )
            cover_row += 1

        ws_cover[f"A{cover_row + 1}"] = "Exigencias legales activas"
        ws_cover.merge_cells(f"A{cover_row + 1}:H{cover_row + 1}")
        style_range(
            ws_cover,
            f"A{cover_row + 1}:H{cover_row + 1}",
            fill=header_fill,
            font=Font(bold=True, size=11, color="FFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        legal_row = cover_row + 2
        legal_requirements = (legal_snapshot or {}).get("requirements") or []
        for requirement in legal_requirements or ["Sin exigencias adicionales registradas."]:
            ws_cover[f"A{legal_row}"] = requirement
            ws_cover.merge_cells(f"A{legal_row}:H{legal_row}")
            style_range(
                ws_cover,
                f"A{legal_row}:H{legal_row}",
                fill=light_fill,
                font=Font(size=10, color=palette["ink"].lstrip("#")),
                alignment=Alignment(vertical="top", wrap_text=True),
            )
            legal_row += 1

        ws_cover[f"A{legal_row + 1}"] = "Trazabilidad de origen"
        ws_cover.merge_cells(f"A{legal_row + 1}:H{legal_row + 1}")
        style_range(
            ws_cover,
            f"A{legal_row + 1}:H{legal_row + 1}",
            fill=header_fill,
            font=Font(bold=True, size=11, color="FFFFFF"),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        origin_row = legal_row + 2
        for origin, count in (insights.get("origins") or {}).items():
            ws_cover[f"A{origin_row}"] = origin
            ws_cover[f"B{origin_row}"] = count
            ws_cover[f"A{origin_row}"].border = border
            ws_cover[f"B{origin_row}"].border = border
            ws_cover[f"A{origin_row}"].fill = light_fill
            ws_cover[f"B{origin_row}"].fill = light_fill
            origin_row += 1
        if origin_row == legal_row + 2:
            ws_cover[f"A{origin_row}"] = "Sin origenes registrados."
            ws_cover.merge_cells(f"A{origin_row}:H{origin_row}")
            style_range(
                ws_cover,
                f"A{origin_row}:H{origin_row}",
                fill=light_fill,
                font=Font(size=10, color=palette["ink"].lstrip("#")),
                alignment=Alignment(vertical="top", wrap_text=True),
            )

        for col, width in {"A": 18, "B": 26, "C": 18, "D": 18, "E": 18, "F": 26, "G": 18, "H": 18}.items():
            ws_cover.column_dimensions[col].width = width

        headers = [
            "Proceso",
            "Tarea",
            "Puesto",
            "Lugar",
            "Peligro / Factor",
            "Riesgo ISP",
            "VEP",
            "Nivel",
            "Controles",
            "EPP",
            "Protocolos",
            "Origen",
            "Alertas",
            "Responsable",
            "Ref. legal",
        ]
        ws_matrix.sheet_view.showGridLines = False
        ws_matrix["A1"] = "MIPER consolidada"
        ws_matrix.merge_cells("A1:O1")
        style_range(
            ws_matrix,
            "A1:O1",
            fill=title_fill,
            font=Font(bold=True, size=14, color="FFFFFF"),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws_matrix["A2"] = (
            f"Proyecto {payload.get('project_code') or folder.id} | Cliente {payload.get('customer_name') or '-'} | "
            f"Servicio {payload.get('service_profile_name') or payload.get('service_type_name') or '-'} | "
            f"Emitido {insights.get('generated_at') or self._format_export_timestamp()}"
        )
        ws_matrix.merge_cells("A2:O2")
        style_range(
            ws_matrix,
            "A2:O2",
            fill=soft_fill,
            font=Font(size=10, bold=True, color=palette["ink"].lstrip("#")),
            alignment=Alignment(horizontal="left", vertical="center"),
        )
        ws_matrix["A3"] = (
            f"Readiness {float(payload.get('readiness_pct') or 0):.1f}% | Segmento legal {(legal_snapshot or {}).get('segment') or '-'} | "
            f"Protocolos {', '.join(insights.get('protocols') or []) or '-'}"
        )
        ws_matrix.merge_cells("A3:O3")
        style_range(
            ws_matrix,
            "A3:O3",
            fill=light_fill,
            font=Font(size=9, color=palette["muted"].lstrip("#")),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
        )

        ws_matrix.append([])
        ws_matrix.append(headers)
        for cell in ws_matrix[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in self._matrix_export_rows(rows):
            ws_matrix.append(row)

        for idx, row in enumerate(ws_matrix.iter_rows(min_row=6, max_row=ws_matrix.max_row), start=6):
            raw = rows[idx - 6] if idx - 6 < len(rows) else {}
            fill = PatternFill("solid", fgColor=self._row_color(raw.get("vep"), raw.get("risk_level") or "").lstrip("#"))
            for cell in row:
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9)
            row[6].font = Font(size=10, bold=True, color=palette["ink"].lstrip("#"))
            row[7].font = Font(size=9, bold=True, color=palette["ink"].lstrip("#"))
        ws_matrix.freeze_panes = "A6"
        ws_matrix.auto_filter.ref = f"A5:O{ws_matrix.max_row}"
        widths = [18, 22, 16, 18, 22, 22, 10, 12, 28, 18, 18, 16, 24, 18, 16]
        for idx, width in enumerate(widths, start=1):
            ws_matrix.column_dimensions[get_column_letter(idx)].width = width
        ws_matrix.sheet_view.zoomScale = 85
        ws_matrix.page_setup.orientation = ws_matrix.ORIENTATION_LANDSCAPE
        ws_matrix.page_setup.fitToWidth = 1
        ws_matrix.page_setup.fitToHeight = 0
        ws_matrix.print_title_rows = "$1:$5"

        ws_alerts.sheet_view.showGridLines = False
        ws_alerts["A1"] = "Bloqueos y alertas priorizadas"
        ws_alerts.merge_cells("A1:H1")
        style_range(
            ws_alerts,
            "A1:H1",
            fill=title_fill,
            font=Font(bold=True, size=14, color="FFFFFF"),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws_alerts.append([])
        ws_alerts.append(["Proceso", "Tarea", "Riesgo", "VEP", "Nivel", "Accion", "Alertas", "Origen"])
        for cell in ws_alerts[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        alert_rows = insights.get("blocking_rows") or []
        if alert_rows:
            for item in alert_rows:
                ws_alerts.append(
                    [
                        item.get("process_name") or "-",
                        item.get("task_name") or "-",
                        item.get("risk_name") or "-",
                        item.get("vep") or 0,
                        item.get("risk_level") or "-",
                        item.get("action_required") or "-",
                        " | ".join(item.get("alerts") or []) or "-",
                        ", ".join(item.get("origins") or []) or "-",
                    ]
                )
        else:
            ws_alerts.append(["Sin bloqueos ni alertas prioritarias.", "", "", "", "", "", "", ""])
        for row in ws_alerts.iter_rows(min_row=4, max_row=ws_alerts.max_row):
            fill = solid(palette["danger"]) if str(row[4].value or "").lower() == "intolerable" else light_fill
            for cell in row:
                cell.border = border
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9, color=palette["ink"].lstrip("#"))
        for idx, width in enumerate([18, 22, 22, 10, 14, 26, 32, 24], start=1):
            ws_alerts.column_dimensions[get_column_letter(idx)].width = width

        ws_trace.sheet_view.showGridLines = False
        ws_trace["A1"] = "Trazabilidad de filas"
        ws_trace.merge_cells("A1:I1")
        style_range(
            ws_trace,
            "A1:I1",
            fill=title_fill,
            font=Font(bold=True, size=14, color="FFFFFF"),
            alignment=Alignment(horizontal="center", vertical="center"),
        )
        ws_trace.append([])
        ws_trace.append(["Proceso", "Tarea", "Peligro", "Riesgo", "Origen", "Protocolos", "Responsable", "Ref. legal", "Accion"])
        for cell in ws_trace[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in rows:
            ws_trace.append(
                [
                    row.get("process_name") or "",
                    row.get("task_name") or row.get("activity") or "",
                    row.get("hazard_factor") or row.get("hazard") or "",
                    " - ".join([str(value) for value in [row.get("master_risk_code"), row.get("risk_name") or row.get("risk")] if value]),
                    ", ".join(
                        _normalize_str_list(row.get("source_labels"))
                        or _normalize_str_list(row.get("source_titles"))
                        or _normalize_str_list(row.get("origin_blocks"))
                    ),
                    ", ".join(_normalize_str_list(row.get("protocol_codes"))),
                    row.get("owner_name") or "",
                    row.get("legal_reference") or "",
                    row.get("action_required") or "",
                ]
            )
        for row in ws_trace.iter_rows(min_row=4, max_row=ws_trace.max_row):
            for cell in row:
                cell.border = border
                cell.fill = light_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(size=9, color=palette["ink"].lstrip("#"))
        for idx, width in enumerate([18, 22, 22, 24, 18, 18, 18, 20, 26], start=1):
            ws_trace.column_dimensions[get_column_letter(idx)].width = width

        bio = BytesIO()
        workbook.save(bio)
        return bio.getvalue()

    def _build_miper_pdf(self, folder: SafetyFolder, context: Dict[str, Any]) -> bytes:
        return self._build_miper_pdf_compact(folder, context)
        matrix = context.get("matrix") or {}
        rows = context.get("rows") or []
        payload = context.get("folder") or {}
        insights = context.get("insights") or {}
        summary = payload.get("summary") or {}
        legal_snapshot = insights.get("legal_snapshot") or summary.get("legal_snapshot") or {}
        palette = self._export_palette(payload.get("traffic_light") or "red")
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=18 * mm,
            bottomMargin=14 * mm,
        )

        def safe_html(value: Any) -> str:
            return (
                str(value or "")
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )

        def level_fill(level: str, vep: Any):
            return colors.HexColor(self._row_color(vep, level))

        def draw_page(canvas, pdf_doc):
            width, height = pdf_doc.pagesize
            canvas.saveState()
            canvas.setFillColor(colors.HexColor(palette["brand"]))
            canvas.rect(0, height - 14 * mm, width, 14 * mm, fill=1, stroke=0)
            canvas.setFillColor(colors.white)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.drawString(pdf_doc.leftMargin, height - 8.8 * mm, f"MIPER | {payload.get('project_code') or folder.id}")
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(
                width - pdf_doc.rightMargin,
                height - 8.8 * mm,
                f"Emitido {insights.get('generated_at') or self._format_export_timestamp()} | Pag. {canvas.getPageNumber()}",
            )
            canvas.setStrokeColor(colors.HexColor(palette["line"]))
            canvas.line(pdf_doc.leftMargin, 10 * mm, width - pdf_doc.rightMargin, 10 * mm)
            canvas.setFillColor(colors.HexColor(palette["muted"]))
            canvas.setFont("Helvetica", 7.2)
            canvas.drawString(
                pdf_doc.leftMargin,
                6.5 * mm,
                f"{payload.get('customer_name') or '-'} | {payload.get('service_profile_name') or payload.get('service_type_name') or '-'}",
            )
            canvas.restoreState()

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="MiperTitle", fontName="Helvetica-Bold", fontSize=16, leading=18, textColor=colors.HexColor(palette["ink"])))
        styles.add(ParagraphStyle(name="MiperSub", fontName="Helvetica", fontSize=9, leading=11, textColor=colors.HexColor(palette["muted"])))
        styles.add(ParagraphStyle(name="MiperSection", fontName="Helvetica-Bold", fontSize=10.5, leading=12, textColor=colors.HexColor(palette["ink"]), spaceAfter=3))
        styles.add(ParagraphStyle(name="MiperBody", fontName="Helvetica", fontSize=8.4, leading=10.5, textColor=colors.HexColor(palette["ink"])))
        styles.add(ParagraphStyle(name="CellSmall", fontName="Helvetica", fontSize=7.5, leading=9, textColor=colors.black))
        styles.add(ParagraphStyle(name="CellSmallBold", fontName="Helvetica-Bold", fontSize=7.5, leading=9, textColor=colors.black))
        story: List[Any] = []
        story.append(Spacer(1, 3 * mm))
        story.append(Paragraph("MIPER - Matriz de Identificacion de Peligros y Evaluacion de Riesgos", styles["MiperTitle"]))
        story.append(Spacer(1, 4 * mm))
        story.append(
            Paragraph(
                f"{safe_html(payload.get('project_code') or folder.id)} | {safe_html(payload.get('lead_title') or 'Sin oportunidad')} | "
                f"{safe_html(payload.get('customer_name') or '-')} | {safe_html(payload.get('client_site_name') or '-')} | {safe_html(payload.get('client_area_name') or '-')}",
                styles["MiperSub"],
            )
        )
        story.append(
            Paragraph(
                f"Readiness: {float(payload.get('readiness_pct') or 0):.1f}% | Semaforo: {palette['traffic_label']} | "
                f"Segmento legal: {safe_html((legal_snapshot or {}).get('segment') or '-')} | Emitido: {safe_html(insights.get('generated_at') or self._format_export_timestamp())}",
                styles["MiperSub"],
            )
        )
        story.append(Spacer(1, 4 * mm))

        kpi_data = [
            [
                Paragraph(f"<b>Readiness</b><br/>{float(payload.get('readiness_pct') or 0):.1f}%", styles["MiperBody"]),
                Paragraph(f"<b>Filas MIPER</b><br/>{safe_html(insights.get('row_count') or len(rows))}", styles["MiperBody"]),
                Paragraph(f"<b>Bloqueos</b><br/>{safe_html(len(insights.get('blocking_rows') or []))}", styles["MiperBody"]),
                Paragraph(f"<b>Protocolos</b><br/>{safe_html(len(insights.get('protocols') or []))}", styles["MiperBody"]),
            ],
            [
                Paragraph(f"<b>Reglas heredadas</b><br/>{safe_html(insights.get('matched_rule_count') or 0)}", styles["MiperBody"]),
                Paragraph(f"<b>Servicio</b><br/>{safe_html(payload.get('service_profile_name') or payload.get('service_type_name') or '-')}", styles["MiperBody"]),
                Paragraph(f"<b>Instalacion</b><br/>{safe_html(payload.get('client_site_name') or '-')}", styles["MiperBody"]),
                Paragraph(f"<b>Area</b><br/>{safe_html(payload.get('client_area_name') or '-')}", styles["MiperBody"]),
            ],
        ]
        kpi_table = Table(kpi_data, colWidths=[64 * mm, 64 * mm, 64 * mm, 64 * mm])
        kpi_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["surface"])),
                    ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor(palette["line"])),
                    ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                    ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(palette["brand_soft"])),
                    ("BACKGROUND", (1, 0), (1, 0), colors.HexColor(palette["success"])),
                    ("BACKGROUND", (2, 0), (2, 0), colors.HexColor(palette["warning"])),
                    ("BACKGROUND", (3, 0), (3, 0), colors.HexColor(palette["surface_alt"])),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(kpi_table)
        story.append(Spacer(1, 5 * mm))

        scope_sources = ((insights.get("scope") or {}).get("sources") or [])
        if scope_sources:
            story.append(Paragraph("Alcance de la matriz", styles["MiperSection"]))
            scope_data = [[
                Paragraph("<b>Fuente</b>", styles["CellSmallBold"]),
                Paragraph("<b>Seleccion</b>", styles["CellSmallBold"]),
                Paragraph("<b>Cantidad</b>", styles["CellSmallBold"]),
            ]]
            for item in scope_sources:
                names = _normalize_str_list(item.get("names"))
                scope_data.append(
                    [
                        Paragraph(safe_html(item.get("label") or item.get("key") or "-"), styles["CellSmallBold"]),
                        Paragraph(safe_html(", ".join(names[:8]) or "Sin seleccion"), styles["CellSmall"]),
                        Paragraph(safe_html(item.get("count") or 0), styles["CellSmallBold"]),
                    ]
                )
            scope_table = Table(scope_data, colWidths=[42 * mm, 189 * mm, 30 * mm], repeatRows=1)
            scope_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["brand"])),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor(palette["surface"])),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(palette["line"])),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            story.append(scope_table)
            story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Resumen ejecutivo", styles["MiperSection"]))
        scope_note = insights.get("scope_note") or insights.get("folder_note") or "Sin notas adicionales registradas."
        executive_data = [
            [Paragraph("<b>Alcance</b>", styles["MiperBody"]), Paragraph(safe_html(scope_note), styles["MiperBody"])],
            [Paragraph("<b>Exigencias legales</b>", styles["MiperBody"]), Paragraph("<br/>".join([safe_html(item) for item in ((legal_snapshot or {}).get("requirements") or ["Sin exigencias adicionales registradas."])]), styles["MiperBody"])],
            [Paragraph("<b>Protocolos MINSAL</b>", styles["MiperBody"]), Paragraph(safe_html(", ".join(insights.get("protocols") or []) or "-"), styles["MiperBody"])],
        ]
        executive_table = Table(executive_data, colWidths=[42 * mm, 219 * mm])
        executive_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["surface"])),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor(palette["line"])),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(palette["surface_alt"])),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(executive_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Bloqueos y acciones prioritarias", styles["MiperSection"]))
        blocking_rows = insights.get("blocking_rows") or []
        if blocking_rows:
            blocker_data = [[
                Paragraph("<b>Proceso / tarea</b>", styles["CellSmallBold"]),
                Paragraph("<b>Riesgo</b>", styles["CellSmallBold"]),
                Paragraph("<b>VEP</b>", styles["CellSmallBold"]),
                Paragraph("<b>Accion requerida</b>", styles["CellSmallBold"]),
                Paragraph("<b>Alertas</b>", styles["CellSmallBold"]),
            ]]
            for item in blocking_rows[:10]:
                blocker_data.append(
                    [
                        Paragraph(f"{safe_html(item.get('process_name') or '-')}<br/>{safe_html(item.get('task_name') or '-')}", styles["CellSmall"]),
                        Paragraph(safe_html(item.get("risk_name") or "-"), styles["CellSmall"]),
                        Paragraph(safe_html(item.get("vep") or 0), styles["CellSmallBold"]),
                        Paragraph(safe_html(item.get("action_required") or "-"), styles["CellSmall"]),
                        Paragraph("<br/>".join([safe_html(alert) for alert in (item.get("alerts") or [])]) or "-", styles["CellSmall"]),
                    ]
                )
            blocker_table = Table(blocker_data, colWidths=[52 * mm, 58 * mm, 14 * mm, 68 * mm, 69 * mm], repeatRows=1)
            blocker_style = TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["brand"])),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(palette["line"])),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
            for idx, item in enumerate(blocking_rows[:10], start=1):
                blocker_style.add("BACKGROUND", (0, idx), (-1, idx), level_fill(item.get("risk_level") or "", item.get("vep")))
            blocker_table.setStyle(blocker_style)
            story.append(blocker_table)
        else:
            blocker_clear = Table([[Paragraph("No se registran bloqueos criticos al momento de la emision.", styles["MiperBody"])]], colWidths=[261 * mm])
            blocker_clear.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["success"])),
                        ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]
                )
            )
            story.append(blocker_clear)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Matriz consolidada", styles["MiperSection"]))
        table_data = [[
            Paragraph("<b>Proceso</b>", styles["CellSmallBold"]),
            Paragraph("<b>Tarea</b>", styles["CellSmallBold"]),
            Paragraph("<b>Peligro / Riesgo</b>", styles["CellSmallBold"]),
            Paragraph("<b>VEP</b>", styles["CellSmallBold"]),
            Paragraph("<b>Nivel</b>", styles["CellSmallBold"]),
            Paragraph("<b>Controles</b>", styles["CellSmallBold"]),
            Paragraph("<b>EPP / Protocolos</b>", styles["CellSmallBold"]),
            Paragraph("<b>Alertas / origen</b>", styles["CellSmallBold"]),
        ]]
        for row in rows:
            table_data.append(
                [
                    Paragraph(safe_html(row.get("process_name") or "-"), styles["CellSmall"]),
                    Paragraph(safe_html(row.get("task_name") or row.get("activity") or "-"), styles["CellSmall"]),
                    Paragraph(
                        "<br/>".join(
                            filter(
                                None,
                                [
                                    safe_html(row.get("hazard_factor") or row.get("hazard") or "-"),
                                    safe_html(" - ".join([v for v in [row.get("master_risk_code"), row.get("risk_name") or row.get("risk")] if v]) or "-"),
                                ],
                            )
                        ),
                        styles["CellSmall"],
                    ),
                    Paragraph(safe_html(row.get("vep") or 0), styles["CellSmallBold"]),
                    Paragraph(safe_html(row.get("risk_level") or "-"), styles["CellSmall"]),
                    Paragraph(safe_html(row.get("controls") or "-"), styles["CellSmall"]),
                    Paragraph(
                        "<br/>".join(
                            filter(
                                None,
                                [
                                    "EPP: " + safe_html(", ".join(_normalize_str_list(row.get("required_ppe")))) if _normalize_str_list(row.get("required_ppe")) else "",
                                    "Prot: " + safe_html(", ".join(_normalize_str_list(row.get("protocol_codes")))) if _normalize_str_list(row.get("protocol_codes")) else "",
                                ],
                            )
                        ) or "-",
                        styles["CellSmall"],
                    ),
                    Paragraph(
                        "<br/>".join(
                            filter(
                                None,
                                [
                                    "<br/>".join([safe_html(item) for item in _normalize_str_list(row.get("restriction_alerts"))]) if _normalize_str_list(row.get("restriction_alerts")) else "",
                                    "Origen: " + safe_html(", ".join(_normalize_str_list(row.get("source_labels")) or _normalize_str_list(row.get("source_titles")) or _normalize_str_list(row.get("origin_blocks")))),
                                    "Ref: " + safe_html(row.get("legal_reference") or "") if row.get("legal_reference") else "",
                                ],
                            )
                        ) or "-",
                        styles["CellSmall"],
                    ),
                ]
            )

        col_widths = [24 * mm, 28 * mm, 48 * mm, 10 * mm, 16 * mm, 45 * mm, 45 * mm, 45 * mm]
        matrix_table = LongTable(table_data, colWidths=col_widths, repeatRows=1)
        matrix_style = TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(palette["brand"])),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor(palette["line"])),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("LEADING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
        for idx, row in enumerate(rows, start=1):
            fill = level_fill(row.get("risk_level") or "", row.get("vep"))
            matrix_style.add("BACKGROUND", (0, idx), (-1, idx), fill)
        matrix_table.setStyle(matrix_style)
        story.append(matrix_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("Control y firmas", styles["MiperSection"]))
        signature_table = Table(
            [[
                Paragraph("<b>Elaborado por</b><br/><br/>____________________________", styles["MiperBody"]),
                Paragraph("<b>Revisado por</b><br/><br/>____________________________", styles["MiperBody"]),
                Paragraph("<b>Aprobado por cliente / mandante</b><br/><br/>____________________________", styles["MiperBody"]),
            ]],
            colWidths=[87 * mm, 87 * mm, 87 * mm],
        )
        signature_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(palette["surface"])),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor(palette["line"])),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor(palette["line"])),
                    ("TOPPADDING", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(signature_table)

        doc.build(story, onFirstPage=draw_page, onLaterPages=draw_page)
        return buffer.getvalue()

    async def export_matrix_excel(self, request: Request):
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        context = self._miper_export_context(folder)
        if not context.get("matrix"):
            return Response.bad_request("No hay matriz MIPER disponible para exportar")
        content = self._build_miper_excel(folder, context)
        from fastapi.responses import StreamingResponse

        filename = self._export_filename(folder, "xlsx")
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    async def export_matrix_pdf(self, request: Request):
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        context = self._miper_export_context(folder)
        if not context.get("matrix"):
            return Response.bad_request("No hay matriz MIPER disponible para exportar")
        content = self._build_miper_pdf(folder, context)
        from fastapi.responses import StreamingResponse

        filename = self._export_filename(folder, "pdf")
        return StreamingResponse(
            BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    def _miper_export_context_from_matrix(self, matrix: SafetyRiskMatrix) -> Tuple[Any, Dict[str, Any]]:
        if matrix.folder_id:
            folder = SafetyFolder.find_by_id(matrix.folder_id)
            if folder:
                return folder, self._miper_export_context(folder)

        class MatrixExportFolder:
            def __init__(self, matrix_id: Any):
                self.id = matrix_id

        matrix_payload = matrix.to_dict()
        rows = sorted(
            matrix_payload.get("rows", []),
            key=lambda row: (
                row.get("source_sort") or [99, 0, 0],
                -(_safe_int(row.get("residual_risk_value") or row.get("vep"), 0) or 0),
                str(row.get("source_title") or ""),
                str(row.get("process_name") or row.get("task_name") or ""),
                str(row.get("risk_name") or row.get("risk") or ""),
            ),
        )
        payload = {
            "id": matrix.id,
            "project_code": matrix.code or f"MIPER-{matrix.id}",
            "lead_title": matrix.title or "",
            "customer_name": "",
            "client_site_name": matrix.work_center or "",
            "client_area_name": "",
            "service_profile_name": "",
            "service_type_name": "",
            "traffic_light": "red" if matrix_payload.get("approval_blocked") else "green",
            "readiness_pct": 100 if matrix.status == "approved" else 75,
            "summary": {},
        }
        matrix_payload["row_count"] = len(rows)
        context = {
            "folder": payload,
            "matrix": matrix_payload,
            "rows": rows,
            "insights": self._build_export_insights(payload, matrix_payload, rows),
        }
        return MatrixExportFolder(matrix.id), context

    def _risk_matrix_export_filename(self, matrix: SafetyRiskMatrix, extension: str) -> str:
        base = matrix.code or f"risk_matrix_{matrix.id}"
        base = re.sub(r"[^A-Za-z0-9._-]+", "_", str(base)).strip("_") or f"risk_matrix_{matrix.id}"
        return f"MIPER_{base}.{extension}"

    async def export_risk_matrix_excel(self, request: Request):
        err = self._require_access()
        if err:
            return err
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        folder, context = self._miper_export_context_from_matrix(matrix)
        if not context.get("matrix"):
            return Response.bad_request("No hay matriz MIPER disponible para exportar")
        content = self._build_miper_excel(folder, context)
        from fastapi.responses import StreamingResponse

        filename = self._risk_matrix_export_filename(matrix, "xlsx")
        return StreamingResponse(
            BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    async def export_risk_matrix_pdf(self, request: Request):
        err = self._require_access()
        if err:
            return err
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        folder, context = self._miper_export_context_from_matrix(matrix)
        if not context.get("matrix"):
            return Response.bad_request("No hay matriz MIPER disponible para exportar")
        content = self._build_miper_pdf(folder, context)
        from fastapi.responses import StreamingResponse

        filename = self._risk_matrix_export_filename(matrix, "pdf")
        return StreamingResponse(
            BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    def _build_irl_draft_payload(
        self,
        folder: SafetyFolder,
        employee_payload: Dict[str, Any],
        overrides: Dict[str, Any],
    ) -> Dict[str, Any]:
        folder_payload = self._folder_payload(folder)
        matrix = self._matrix_for_folder(folder.id)
        if not matrix:
            self._generate_matrix_for_folder(folder)
            matrix = self._matrix_for_folder(folder.id)
        rows = (matrix.to_dict().get("rows", []) if matrix else [])[:]
        documents = self._documents_for_folder(folder.id)

        worker_name = str(overrides.get("worker_name") or employee_payload.get("full_name") or "").strip()
        position_title = str(overrides.get("position_title") or employee_payload.get("position_title") or "").strip()
        place_name = str(
            overrides.get("place_name")
            or folder_payload.get("client_area_name")
            or folder_payload.get("client_site_name")
            or folder_payload.get("customer_name")
            or "Lugar de trabajo"
        ).strip()
        service_name = (
            folder_payload.get("service_profile_name")
            or folder_payload.get("service_type_name")
            or "Servicio operativo"
        )

        normalized_position = position_title.lower()
        normalized_place = place_name.lower()
        relevant_rows = [
            row
            for row in rows
            if normalized_position
            and normalized_position in str(row.get("position_name") or "").lower()
        ]
        if not relevant_rows:
            relevant_rows = [
                row
                for row in rows
                if normalized_place
                and normalized_place in str(row.get("place_name") or "").lower()
            ]
        if not relevant_rows:
            relevant_rows = rows
        relevant_rows = relevant_rows[:8]

        service_functions = _normalize_str_list(overrides.get("service_functions"))
        if not service_functions:
            seen_functions: Set[str] = set()
            for row in relevant_rows:
                function_label = (
                    row.get("task_name")
                    or row.get("activity")
                    or row.get("process_name")
                    or ""
                )
                normalized = str(function_label).strip()
                if normalized and normalized not in seen_functions:
                    service_functions.append(normalized)
                    seen_functions.add(normalized)

        risk_items = _normalize_irl_risk_items(overrides.get("risk_items"))
        if not risk_items:
            for row in relevant_rows:
                work_method_parts = []
                task_name = str(row.get("task_name") or row.get("activity") or "").strip()
                if task_name:
                    work_method_parts.append(f"Ejecutar {task_name.lower()} segun procedimiento autorizado")
                if row.get("action_required"):
                    work_method_parts.append(str(row.get("action_required")))
                protocols = _normalize_str_list(row.get("protocol_codes"))
                if protocols:
                    work_method_parts.append(f"Aplicar protocolos {', '.join(protocols)}")
                risk_items.append(
                    {
                        "risk": " - ".join(
                            [
                                str(value)
                                for value in [
                                    row.get("hazard_factor") or row.get("hazard"),
                                    row.get("risk_name") or row.get("risk"),
                                ]
                                if value
                            ]
                        ) or "Riesgo operacional",
                        "preventive_measures": str(row.get("controls") or "Aplicar controles definidos por la matriz MIPER.").strip(),
                        "work_method": ". ".join(work_method_parts) or "Ejecutar la tarea conforme a procedimiento y supervision directa.",
                        "source_process": str(row.get("process_name") or "").strip(),
                        "source_task": task_name,
                    }
                )

        complement_materials = _normalize_irl_materials(overrides.get("complement_materials"))
        if not complement_materials:
            for document in documents[:5]:
                complement_materials.append(
                    {
                        "name": document.title or "Documento de apoyo",
                        "type": document.document_type or "otro",
                        "location": "Carpeta de seguridad / dossier",
                    }
                )

        default_intro = (
            "De acuerdo con lo establecido en el D.S. N° 44, este registro informa en forma oportuna "
            "los riesgos laborales asociados al cargo, lugar de trabajo y funciones del servicio, junto con "
            "sus medidas preventivas y metodos de trabajo correctos."
        )

        workspace_features = str(
            overrides.get("workspace_features")
            or f"Las labores de {service_name.lower()} se ejecutaran en {place_name}, manteniendo accesos seguros, "
               "superficies estables, delimitacion del area y condiciones compatibles con el procedimiento."
        ).strip()
        environmental_conditions = str(
            overrides.get("environmental_conditions")
            or f"Se deben controlar las condiciones ambientales del puesto, incluyendo iluminacion, ventilacion, ruido, "
               f"interferencias operacionales y exposicion propia de {place_name}."
        ).strip()
        order_cleanliness = str(
            overrides.get("order_cleanliness")
            or "Mantener orden y aseo permanente, rutas despejadas, materiales acopiados en forma segura y retiro de residuos al cierre de la tarea."
        ).strip()
        machines_tools = str(
            overrides.get("machines_tools")
            or "Utilizar solo maquinas, herramientas y equipos autorizados, inspeccionados y compatibles con el procedimiento y checklist del servicio."
        ).strip()

        return {
            "folder_id": folder.id,
            "employee_id": _safe_int(overrides.get("employee_id"), employee_payload.get("id")),
            "title": str(
                overrides.get("title")
                or f"IRL - {worker_name or position_title or service_name} - {folder_payload.get('project_code') or folder.id}"
            ).strip(),
            "status": str(overrides.get("status") or "draft").strip() or "draft",
            "worker_name": worker_name,
            "worker_identifier": str(
                overrides.get("worker_identifier")
                or employee_payload.get("employee_code")
                or ""
            ).strip(),
            "position_title": position_title or "Cargo por definir",
            "place_name": place_name,
            "activity_name": str(overrides.get("activity_name") or service_name).strip(),
            "activity_period": str(
                overrides.get("activity_period")
                or (f"Inicio planificado: {folder_payload.get('planned_start_date')}" if folder_payload.get("planned_start_date") else "")
            ).strip(),
            "modality": str(overrides.get("modality") or "Presencial").strip() or "Presencial",
            "duration_hours": str(overrides.get("duration_hours") or "08:00").strip() or "08:00",
            "executor_name": str(
                overrides.get("executor_name")
                or folder_payload.get("customer_name")
                or "Equipo de operaciones"
            ).strip(),
            "relator_background": str(
                overrides.get("relator_background")
                or f"Supervisor / prevencionista del servicio {service_name}"
            ).strip(),
            "target_group": str(
                overrides.get("target_group")
                or f"{worker_name or 'Persona trabajadora'} - cargo {position_title or 'asignado'}"
            ).strip(),
            "workspace_features": workspace_features,
            "environmental_conditions": environmental_conditions,
            "order_cleanliness": order_cleanliness,
            "machines_tools": machines_tools,
            "service_functions": service_functions,
            "risk_items": risk_items,
            "complement_materials": complement_materials,
            "observations": str(overrides.get("observations") or folder_payload.get("notes") or "").strip(),
            "intro_text": str(overrides.get("intro_text") or default_intro).strip(),
            "theme_color": str(overrides.get("theme_color") or "#0F4C81").strip() or "#0F4C81",
            "company_id": folder.company_id,
        }

    def _irl_filename(self, irl: SafetyIRLRecord, extension: str) -> str:
        folder = SafetyFolder.find_by_id(irl.folder_id)
        folder_code = folder.to_dict().get("project_code") if folder else f"folder_{irl.folder_id}"
        worker_label = irl.worker_name or irl.position_title or f"irl_{irl.id}"
        base = f"IRL_{folder_code}_{worker_label}"
        base = re.sub(r"[^A-Za-z0-9._-]+", "_", str(base)).strip("_") or f"irl_{irl.id}"
        return f"{base}.{extension}"

    def _build_irl_pdf(self, irl: SafetyIRLRecord) -> bytes:
        payload = irl.to_dict()
        color_value = payload.get("theme_color") or "#0F4C81"
        try:
            accent = colors.HexColor(color_value)
        except Exception:
            accent = colors.HexColor("#0F4C81")

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=14 * mm,
            rightMargin=14 * mm,
            topMargin=18 * mm,
            bottomMargin=14 * mm,
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="IRLTitle", fontName="Helvetica-Bold", fontSize=16, leading=18, textColor=accent))
        styles.add(ParagraphStyle(name="IRLSub", fontName="Helvetica", fontSize=9, leading=11, textColor=colors.HexColor("#475569")))
        styles.add(ParagraphStyle(name="IRLSection", fontName="Helvetica-Bold", fontSize=10.5, leading=12, textColor=colors.HexColor("#0f172a"), spaceAfter=4))
        styles.add(ParagraphStyle(name="IRLBody", fontName="Helvetica", fontSize=8.6, leading=10.8, textColor=colors.HexColor("#0f172a")))
        styles.add(ParagraphStyle(name="IRLCell", fontName="Helvetica", fontSize=7.8, leading=9.2, textColor=colors.black))
        styles.add(ParagraphStyle(name="IRLCellBold", fontName="Helvetica-Bold", fontSize=7.8, leading=9.2, textColor=colors.black))

        def safe_html(value: Any) -> str:
            return (
                str(value or "")
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
            )

        story: List[Any] = []
        story.append(Paragraph("Registro de Informacion de Riesgos Laborales", styles["IRLTitle"]))
        story.append(Spacer(1, 3 * mm))
        story.append(
            Paragraph(
                f"{safe_html(payload.get('title'))} | Cargo: {safe_html(payload.get('position_title'))} | Lugar: {safe_html(payload.get('place_name'))}",
                styles["IRLSub"],
            )
        )
        story.append(Spacer(1, 4 * mm))

        intro_table = Table(
            [[Paragraph(safe_html(payload.get("intro_text")), styles["IRLBody"])]],
            colWidths=[182 * mm],
        )
        intro_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                    ("BOX", (0, 0), (-1, -1), 0.4, accent),
                    ("TOPPADDING", (0, 0), (-1, -1), 8),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(intro_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("1. Informacion de la actividad", styles["IRLSection"]))
        activity_table = Table(
            [
                ["Actividad", payload.get("activity_name") or "-", "Periodo", payload.get("activity_period") or "-"],
                ["Modalidad", payload.get("modality") or "-", "Duracion", payload.get("duration_hours") or "-"],
                ["Ejecuta", payload.get("executor_name") or "-", "Relator", payload.get("relator_background") or "-"],
                ["Grupo objetivo", payload.get("target_group") or "-", "Version", str(payload.get("version") or 1)],
            ],
            colWidths=[32 * mm, 60 * mm, 24 * mm, 66 * mm],
        )
        activity_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (0, -1), accent),
                    ("BACKGROUND", (2, 0), (2, -1), accent),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                    ("TEXTCOLOR", (2, 0), (2, -1), colors.white),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(activity_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("2. Caracteristicas del lugar de trabajo", styles["IRLSection"]))
        place_table = Table(
            [
                ["Espacio de trabajo", payload.get("workspace_features") or "-"],
                ["Condiciones ambientales", payload.get("environmental_conditions") or "-"],
                ["Orden y aseo", payload.get("order_cleanliness") or "-"],
                ["Maquinas y herramientas", payload.get("machines_tools") or "-"],
            ],
            colWidths=[46 * mm, 136 * mm],
        )
        place_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E2E8F0")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                ]
            )
        )
        story.append(place_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("3. Informacion de los riesgos", styles["IRLSection"]))
        risk_table_data: List[List[Any]] = [[
            Paragraph("<b>Riesgos</b>", styles["IRLCellBold"]),
            Paragraph("<b>Medidas preventivas</b>", styles["IRLCellBold"]),
            Paragraph("<b>Metodo o procedimiento de trabajo</b>", styles["IRLCellBold"]),
        ]]
        for item in payload.get("risk_items") or []:
            risk_table_data.append(
                [
                    Paragraph(safe_html(item.get("risk") or "-"), styles["IRLCell"]),
                    Paragraph(safe_html(item.get("preventive_measures") or "-"), styles["IRLCell"]),
                    Paragraph(safe_html(item.get("work_method") or "-"), styles["IRLCell"]),
                ]
            )
        if len(risk_table_data) == 1:
            risk_table_data.append(
                [
                    Paragraph("Sin riesgos configurados.", styles["IRLCell"]),
                    Paragraph("-", styles["IRLCell"]),
                    Paragraph("-", styles["IRLCell"]),
                ]
            )
        risk_table = LongTable(risk_table_data, colWidths=[52 * mm, 66 * mm, 64 * mm], repeatRows=1)
        risk_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), accent),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(risk_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("4. Funciones del servicio y material complementario", styles["IRLSection"]))
        functions_text = "<br/>".join([f"- {safe_html(item)}" for item in (payload.get("service_functions") or [])]) or "- No informado"
        materials_text = "<br/>".join(
            [
                f"- {safe_html(item.get('name'))} ({safe_html(item.get('type'))}) - {safe_html(item.get('location'))}"
                for item in (payload.get("complement_materials") or [])
            ]
        ) or "- No informado"
        complement_table = Table(
            [
                [Paragraph("<b>Funciones asociadas</b>", styles["IRLBody"]), Paragraph(functions_text, styles["IRLBody"])],
                [Paragraph("<b>Material complementario</b>", styles["IRLBody"]), Paragraph(materials_text, styles["IRLBody"])],
                [Paragraph("<b>Observaciones</b>", styles["IRLBody"]), Paragraph(safe_html(payload.get("observations") or "-"), styles["IRLBody"])],
            ],
            colWidths=[46 * mm, 136 * mm],
        )
        complement_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#E2E8F0")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(complement_table)
        story.append(Spacer(1, 4 * mm))

        story.append(Paragraph("5. Informacion del participante", styles["IRLSection"]))
        worker_table = Table(
            [
                ["Nombre", payload.get("worker_name") or "-", "Identificador", payload.get("worker_identifier") or "-"],
                ["Cargo", payload.get("position_title") or "-", "Lugar", payload.get("place_name") or "-"],
                ["Firma", "_______________________________", "Fecha", "____ / ____ / ________"],
            ],
            colWidths=[24 * mm, 68 * mm, 28 * mm, 62 * mm],
        )
        worker_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                    ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("BACKGROUND", (0, 0), (0, -1), accent),
                    ("BACKGROUND", (2, 0), (2, -1), accent),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                    ("TEXTCOLOR", (2, 0), (2, -1), colors.white),
                ]
            )
        )
        story.append(worker_table)

        doc.build(story)
        return buffer.getvalue()

    def _log_on_lead(self, lead_id: Optional[int], action: str, details: str) -> None:
        if not lead_id:
            return
        try:
            from modules.crm.module_crm import ActivityLog, Lead

            lead = Lead.find_by_id(lead_id)
            if not lead:
                return
            ActivityLog.create(
                {
                    "lead_id": lead.id,
                    "user_id": self.env.user.id if self.env.user else None,
                    "company_id": lead.company_id,
                    "action": action,
                    "details": details,
                }
            )
        except Exception:
            pass

    def _lookups_payload(self) -> Dict[str, Any]:
        self._ensure_miper_catalog()

        leads_data: List[Dict[str, Any]] = []
        employees_data: List[Dict[str, Any]] = []
        customers_data: List[Dict[str, Any]] = []
        procedures_data: List[Dict[str, Any]] = []
        job_profiles_data: List[Dict[str, Any]] = []

        try:
            from modules.crm.module_crm import Lead, Customer, ServiceType

            customers = Customer.search(self._tenant_filter())
            customers.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
            customers_data = [
                {
                    "id": customer.id,
                    "name": customer.name or "",
                    "tax_id": customer.tax_id or "",
                }
                for customer in customers
            ]

            leads = Lead.search(self._tenant_filter())
            leads = [lead for lead in leads if lead.status != "lost"]
            leads.sort(key=lambda item: ((item.title or "").lower(), item.id or 0))
            for lead in leads:
                customer_name = None
                service_type_name = None
                if lead.customer_id:
                    customer = Customer.find_by_id(lead.customer_id)
                    customer_name = customer.name if customer else None
                if lead.service_type_id:
                    service_type = ServiceType.find_by_id(lead.service_type_id)
                    service_type_name = service_type.name if service_type else None
                leads_data.append(
                    {
                        "id": lead.id,
                        "title": lead.title or "",
                        "project_code": lead.project_code or "",
                        "status": lead.status or "open",
                        "customer_id": lead.customer_id,
                        "customer_name": customer_name,
                        "service_type_name": service_type_name,
                    }
                )
        except Exception:
            leads_data = []

        try:
            from modules.hr.module_hr import EmployeeProfile

            employees = EmployeeProfile.search(self._tenant_filter())
            employees = [employee for employee in employees if employee.status in ("active", "onboarding")]
            employees.sort(key=lambda item: ((item.full_name or "").lower(), item.id or 0))
            employees_data = [employee.to_dict() for employee in employees]
        except Exception:
            employees_data = []

        try:
            from modules.safety_procedures.module_safety_procedures import (
                SafetyProcedureTemplate,
                seed_default_procedures,
            )

            seed_default_procedures(self._company_id())
            procedures = SafetyProcedureTemplate.search(self._tenant_filter())
            procedures = [procedure for procedure in procedures if procedure.active]
            procedures.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
            procedures_data = [procedure.to_dict() for procedure in procedures]
        except Exception:
            procedures_data = []

        try:
            from modules.job_profiles.module_job_profiles import JobProfile

            job_profiles = JobProfile.search(self._tenant_filter())
            job_profiles = [profile for profile in job_profiles if profile.active]
            job_profiles.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
            job_profiles_data = [profile.to_dict() for profile in job_profiles]
        except Exception:
            job_profiles_data = []

        profiles = SafetyServiceProfile.search(self._tenant_filter())
        profiles.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        sites = SafetyClientSite.search(self._tenant_filter())
        sites.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        areas = SafetyClientArea.search(self._tenant_filter())
        areas.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        master_risks = SafetyMasterRisk.search(self._tenant_filter())
        master_risks.sort(
            key=lambda item: (
                (item.family or "").lower(),
                (item.isp_code or "").lower(),
                (item.risk_name or "").lower(),
            )
        )
        protocols = SafetyProtocol.search(self._tenant_filter())
        protocols.sort(key=lambda item: ((item.code or "").lower(), item.id or 0))
        ppe_catalog = self._ppe_catalog()
        equipment_blocks = [item for item in SafetyEquipmentBlock.search(self._tenant_filter()) if item.active]
        equipment_blocks.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        methodologies = [
            item.to_dict()
            for item in SafetyRiskMethodology.search(self._tenant_filter())
            if item.active
        ]
        headcount = self._company_employee_count()

        return {
            "leads": leads_data,
            "customers": customers_data,
            "employees": employees_data,
            "service_profiles": [profile.to_dict() for profile in profiles if profile.active],
            "procedures": procedures_data,
            "job_profiles": job_profiles_data,
            "client_sites": [site.to_dict() for site in sites if site.active],
            "client_areas": [area.to_dict() for area in areas if area.active],
            "master_risks": [risk.to_dict() for risk in master_risks if risk.active],
            "protocols": [protocol.to_dict() for protocol in protocols if protocol.active],
            "ppe_catalog": [item.to_dict() for item in ppe_catalog],
            "equipment_blocks": [self._equipment_block_payload(item) for item in equipment_blocks],
            "risk_methodologies": methodologies,
            "risk_methodology": methodology_payload(),
            "compact_miper_methodology": compact_miper_methodology_payload(),
            "risk_matrix_sources": [{"code": item, "label": item.title()} for item in MATRIX_CANONICAL_SOURCES],
            "headcount": headcount,
            "legal_requirements": legal_requirements_for_headcount(headcount),
        }

    async def get_lookups(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        return Response.ok(self._lookups_payload())

    async def list_risk_matrices(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        source_type = str(request.get_param("source_type") or "").strip()
        status = str(request.get_param("status") or "").strip()
        search = str(request.get_param("search") or "").strip().lower()
        matrices = SafetyRiskMatrix.search(self._tenant_filter())
        matrices.sort(key=lambda item: ((item.updated_at if hasattr(item, "updated_at") else "") or "", item.id or 0), reverse=True)
        payloads = [matrix.to_dict() for matrix in matrices]
        if source_type:
            payloads = [item for item in payloads if item.get("source_type") == source_type]
        if status:
            payloads = [item for item in payloads if item.get("status") == status]
        if search:
            payloads = [
                item
                for item in payloads
                if search
                in " ".join(
                    [
                        item.get("code") or "",
                        item.get("title") or "",
                        item.get("process_name") or "",
                        item.get("work_center") or "",
                    ]
                ).lower()
            ]
        return Response.ok({"count": len(payloads), "results": payloads})

    async def get_risk_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        return Response.ok(matrix.to_dict())

    async def create_risk_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        data = request.data or {}
        source_type = str(data.get("source_type") or "manual").strip()
        if source_type not in MATRIX_CANONICAL_SOURCES:
            return Response.bad_request("source_type invalido")
        try:
            matrix = SafetyRiskMatrix.create(
                {
                    "company_id": self._company_id(),
                    "folder_id": _safe_int(data.get("folder_id"), None),
                    "code": data.get("code") or "",
                    "title": data.get("title") or "Matriz MIPER/IPER",
                    "status": data.get("status") or "draft",
                    "version": _safe_int(data.get("version"), 1) or 1,
                    "rows": [],
                    "generation_summary": data.get("generation_summary") or {},
                    "source_type": source_type,
                    "source_id": _safe_int(data.get("source_id"), None),
                    "process_name": data.get("process_name") or "",
                    "work_center": data.get("work_center") or "",
                    "elaboration_date": data.get("elaboration_date") or utc_today_iso(),
                    "last_update_date": data.get("last_update_date") or utc_today_iso(),
                    "review_due_date": data.get("review_due_date") or "",
                    "methodology_config_id": _safe_int(data.get("methodology_config_id"), None),
                    "project_id": _safe_int(data.get("project_id"), None),
                }
            )
            for index, row in enumerate(_normalize_matrix_rows(data.get("rows") or []), start=1):
                SafetyRiskMatrixRow.create(_matrix_row_payload_from_legacy(matrix, row, index * 10))
            return Response.created(matrix.to_dict())
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def update_risk_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        if data.get("status") == "approved" and matrix.to_dict().get("approval_blocked"):
            return Response.bad_request("No se puede aprobar una matriz con riesgos intolerables")
        for field in (
            "code",
            "title",
            "status",
            "version",
            "generation_summary",
            "source_type",
            "source_id",
            "process_name",
            "work_center",
            "elaboration_date",
            "last_update_date",
            "review_due_date",
            "methodology_config_id",
            "project_id",
        ):
            if field in data:
                setattr(matrix, field, data.get(field))
        if "folder_id" in data:
            matrix.folder_id = _safe_int(data.get("folder_id"), None)
        try:
            matrix.save()
            return Response.ok(matrix.to_dict())
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def create_risk_matrix_row(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        try:
            row = SafetyRiskMatrixRow.create(
                _matrix_row_payload_from_legacy(
                    matrix,
                    {
                        **data,
                        "activity": data.get("activity_name") or data.get("activity") or "",
                        "hazard": data.get("hazard_name") or data.get("hazard") or "",
                        "risk": data.get("risk_name") or data.get("risk") or "",
                        "probability": data.get("probability_value") or data.get("probability"),
                        "consequence": data.get("consequence_value") or data.get("consequence"),
                        "control_hierarchy": data.get("required_controls") or data.get("control_hierarchy") or {},
                        "required_ppe": data.get("ppe_summary") or data.get("required_ppe") or [],
                        "protocol_codes": data.get("protocols_summary") or data.get("protocol_codes") or [],
                    },
                    _safe_int(data.get("sequence"), 10) or 10,
                )
            )
            matrix.last_update_date = utc_today_iso()
            matrix.save()
            return Response.created({"row": row.to_dict(), "risk_matrix": matrix.to_dict()})
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def update_risk_matrix_row(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        matrix, error = self._risk_matrix_or_404(request.params.get("id"))
        if error:
            return error
        row, row_error = self._risk_matrix_row_or_404(matrix, request.params.get("row_id"))
        if row_error:
            return row_error
        data = request.data or {}
        field_map = {
            "activity_name": "activity_name",
            "activity": "activity_name",
            "task_name": "task_name",
            "job_position": "job_position",
            "position_name": "job_position",
            "specific_workplace": "specific_workplace",
            "place_name": "specific_workplace",
            "worker_count": "worker_count",
            "routine_type": "routine_type",
            "hazard_name": "hazard_name",
            "hazard": "hazard_name",
            "risk_name": "risk_name",
            "risk": "risk_name",
            "probable_damage": "probable_damage",
            "probability_value": "probability_value",
            "probability": "probability_value",
            "consequence_value": "consequence_value",
            "consequence": "consequence_value",
            "task_type_code": "task_type_code",
            "task_type": "task_type_code",
            "exposed_people_value": "exposed_people_value",
            "pe": "exposed_people_value",
            "exposure_frequency_value": "exposure_frequency_value",
            "fe": "exposure_frequency_value",
            "occurrence_factor_value": "occurrence_factor_value",
            "fo": "occurrence_factor_value",
            "severity_value": "severity_value",
            "severity": "severity_value",
            "current_engineering_controls": "current_engineering_controls",
            "current_admin_controls": "current_admin_controls",
            "current_ppe_controls": "current_ppe_controls",
            "proposed_elimination_controls": "proposed_elimination_controls",
            "proposed_substitution_controls": "proposed_substitution_controls",
            "proposed_engineering_controls": "proposed_engineering_controls",
            "proposed_admin_controls": "proposed_admin_controls",
            "proposed_ppe_controls": "proposed_ppe_controls",
            "safety_management_plan": "safety_management_plan",
            "existing_controls": "existing_controls",
            "controls": "existing_controls",
            "required_controls": "required_controls",
            "control_hierarchy": "required_controls",
            "ppe_summary": "ppe_summary",
            "required_ppe": "ppe_summary",
            "protocols_summary": "protocols_summary",
            "protocol_codes": "protocols_summary",
            "legal_reference": "legal_reference",
            "responsible": "responsible",
            "owner_name": "responsible",
            "observations": "observations",
            "sequence": "sequence",
            "active": "active",
        }
        for incoming, target in field_map.items():
            if incoming in data:
                setattr(row, target, data.get(incoming))
        try:
            row.save()
            matrix.last_update_date = utc_today_iso()
            matrix.save()
            return Response.ok({"row": row.to_dict(), "risk_matrix": matrix.to_dict()})
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def generate_canonical_risk_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        data = request.data or {}
        source_type = str(data.get("source_type") or "manual").strip()
        title = data.get("title") or "Matriz MIPER/IPER generada"
        rows: List[Dict[str, Any]] = []
        source_id = _safe_int(data.get("source_id") or data.get("procedure_id"), None)
        if source_type == "procedure":
            if not source_id:
                return Response.bad_request("source_id/procedure_id requerido")
            payload = {}
            try:
                from modules.safety_procedures.module_safety_procedures import build_procedure_matrix_payload

                payload = build_procedure_matrix_payload(
                    source_id,
                    company_id=self._company_id(),
                    place_name=data.get("work_center") or data.get("place_name") or "",
                )
            except Exception as exc:
                return Response.bad_request(f"No fue posible generar desde procedimiento: {exc}")
            rows = payload.get("rows") or []
            procedure = payload.get("procedure") or {}
            title = data.get("title") or f"MIPER/IPER - {procedure.get('name') or procedure.get('procedure_code') or source_id}"
        elif source_type in ("blocks", "manual"):
            block_ids = _normalize_int_list(data.get("block_ids") or data.get("source_ids") or [])
            try:
                from modules.safety_activities.module_safety_activities import build_activity_block_matrix_rows

                for block_id in block_ids:
                    rows.extend(
                        build_activity_block_matrix_rows(
                            block_id,
                            process_name=data.get("process_name") or "",
                            place_name=data.get("work_center") or data.get("place_name") or "",
                        )
                    )
            except Exception as exc:
                return Response.bad_request(f"No fue posible generar desde BOT: {exc}")
            rows.extend(_normalize_matrix_rows(data.get("rows") or []))
            source_type = "blocks" if block_ids else "manual"
        else:
            return Response.bad_request("source_type debe ser procedure, blocks o manual")

        rows = _normalize_matrix_rows(rows)
        matrix = SafetyRiskMatrix.create(
            {
                "company_id": self._company_id(),
                "folder_id": _safe_int(data.get("folder_id"), None),
                "code": data.get("code") or "",
                "title": title,
                "status": "draft",
                "version": _safe_int(data.get("version"), 1) or 1,
                "rows": rows,
                "generation_summary": {
                    "source_type": source_type,
                    "source_id": source_id,
                    "row_count": len(rows),
                    "generated_at": utc_now_iso(),
                },
                "source_type": source_type,
                "source_id": source_id,
                "process_name": data.get("process_name") or "",
                "work_center": data.get("work_center") or data.get("place_name") or "",
                "elaboration_date": data.get("elaboration_date") or utc_today_iso(),
                "last_update_date": utc_today_iso(),
                "review_due_date": data.get("review_due_date") or "",
                "methodology_config_id": _safe_int(data.get("methodology_config_id"), None),
                "project_id": _safe_int(data.get("project_id"), None),
            }
        )
        for index, row in enumerate(rows, start=1):
            SafetyRiskMatrixRow.create(_matrix_row_payload_from_legacy(matrix, row, index * 10))
        return Response.created(matrix.to_dict())

    async def list_master_risks(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        risks = SafetyMasterRisk.search(self._tenant_filter())
        risks = [risk for risk in risks if risk.active]
        risks.sort(
            key=lambda item: (
                (item.family or "").lower(),
                (item.isp_code or "").lower(),
                (item.risk_name or "").lower(),
            )
        )
        return Response.ok({"count": len(risks), "results": [risk.to_dict() for risk in risks]})

    async def list_ppe_catalog(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        items = self._ppe_catalog()
        search = str(request.get_param("search") or "").strip().lower()
        if search:
            items = [
                item
                for item in items
                if search in " ".join([item.code or "", item.name or "", item.category or "", item.description or ""]).lower()
            ]
        return Response.ok({"count": len(items), "results": [item.to_dict() for item in items]})

    async def create_ppe_catalog_item(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        try:
            item = SafetyPPEItem.create(
                {
                    "code": data.get("code") or "",
                    "name": data.get("name") or "",
                    "category": data.get("category") or "",
                    "description": data.get("description") or "",
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created(item.to_dict())

    async def update_ppe_catalog_item(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        item = SafetyPPEItem.find_by_id(_safe_int(request.params.get("id")))
        if not item or (self.env.user.role != "superadmin" and item.company_id != self._company_id()):
            return Response.not_found("PPE item not found")
        data = request.data or {}
        for field in ("code", "name", "category", "description"):
            if field in data:
                setattr(item, field, data.get(field))
        if "active" in data:
            item.active = _normalize_bool(data.get("active"), True)
        try:
            item.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.ok(item.to_dict())

    async def delete_ppe_catalog_item(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        item = SafetyPPEItem.find_by_id(_safe_int(request.params.get("id")))
        if not item or (self.env.user.role != "superadmin" and item.company_id != self._company_id()):
            return Response.not_found("PPE item not found")
        item.active = False
        item.save()
        return Response.ok({"message": "PPE item archived", "item": item.to_dict()})

    async def list_equipment_blocks(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        items = [item for item in SafetyEquipmentBlock.search(self._tenant_filter()) if item.active]
        search = str(request.get_param("search") or "").strip().lower()
        if search:
            items = [
                item
                for item in items
                if search in " ".join([item.code or "", item.name or "", item.description or "", item.controls_summary or ""]).lower()
            ]
        items.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        return Response.ok({"count": len(items), "results": [self._equipment_block_payload(item) for item in items]})

    async def create_equipment_block(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        self._ensure_miper_catalog()
        data = request.data or {}
        try:
            item = SafetyEquipmentBlock.create(
                {
                    "code": data.get("code") or "",
                    "name": data.get("name") or "",
                    "description": data.get("description") or "",
                    "master_risk_ids": data.get("master_risk_ids") or [],
                    "control_hierarchy": data.get("control_hierarchy") or {},
                    "controls_summary": data.get("controls_summary") or data.get("controls") or "",
                    "required_ppe": data.get("required_ppe") or [],
                    "protocol_codes": data.get("protocol_codes") or [],
                    "sensitivity_tags": data.get("sensitivity_tags") or [],
                    "legal_reference": data.get("legal_reference") or "",
                    "source_note": data.get("source_note") or "",
                    "probability": _safe_int(data.get("probability"), 2) or 2,
                    "consequence": _safe_int(data.get("consequence"), 2) or 2,
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created(self._equipment_block_payload(item))

    async def update_equipment_block(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        item = SafetyEquipmentBlock.find_by_id(_safe_int(request.params.get("id")))
        if not item or (self.env.user.role != "superadmin" and item.company_id != self._company_id()):
            return Response.not_found("Equipment block not found")
        data = request.data or {}
        for field in (
            "code",
            "name",
            "description",
            "master_risk_ids",
            "control_hierarchy",
            "controls_summary",
            "required_ppe",
            "protocol_codes",
            "sensitivity_tags",
            "legal_reference",
            "source_note",
        ):
            if field in data:
                setattr(item, field, data.get(field))
        if "controls" in data and "controls_summary" not in data:
            item.controls_summary = data.get("controls") or ""
        if "probability" in data:
            item.probability = _safe_int(data.get("probability"), 2) or 2
        if "consequence" in data:
            item.consequence = _safe_int(data.get("consequence"), 2) or 2
        if "active" in data:
            item.active = _normalize_bool(data.get("active"), True)
        try:
            item.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.ok(self._equipment_block_payload(item))

    async def delete_equipment_block(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        item = SafetyEquipmentBlock.find_by_id(_safe_int(request.params.get("id")))
        if not item or (self.env.user.role != "superadmin" and item.company_id != self._company_id()):
            return Response.not_found("Equipment block not found")
        item.active = False
        item.save()
        return Response.ok({"message": "Equipment block archived", "equipment": self._equipment_block_payload(item)})

    async def list_generator_rules(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        scope_type = (request.get_param("scope_type") or "").strip().lower()
        rules = SafetyGeneratorRule.search(self._tenant_filter())
        if scope_type:
            rules = [rule for rule in rules if (rule.scope_type or "") == scope_type]
        rules.sort(key=lambda item: ((item.scope_type or "").lower(), (item.name or "").lower(), item.id or 0))
        return Response.ok({"count": len(rules), "results": [rule.to_dict() for rule in rules]})

    async def create_generator_rule(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        risk_id = _safe_int(data.get("master_risk_id"), None)
        if not risk_id and data.get("master_risk_code"):
            risk_code = str(data.get("master_risk_code") or "").strip().upper()
            for risk in SafetyMasterRisk.search(self._tenant_filter()):
                if (risk.isp_code or "").strip().upper() == risk_code:
                    risk_id = risk.id
                    break
        try:
            rule = SafetyGeneratorRule.create(
                {
                    "name": (data.get("name") or "").strip(),
                    "scope_type": data.get("scope_type") or "transversal",
                    "scope_ref_id": _safe_int(data.get("scope_ref_id"), None),
                    "process_name": data.get("process_name") or "",
                    "task_name": data.get("task_name") or "",
                    "position_name": data.get("position_name") or "",
                    "hazard_factor": data.get("hazard_factor") or data.get("hazard") or "",
                    "master_risk_id": risk_id,
                    "probability": _safe_int(data.get("probability"), 2) or 2,
                    "consequence": _safe_int(data.get("consequence"), 2) or 2,
                    "controls_summary": data.get("controls_summary") or data.get("controls") or "",
                    "control_hierarchy": data.get("control_hierarchy") or {},
                    "required_ppe": data.get("required_ppe") or [],
                    "protocol_codes": data.get("protocol_codes") or [],
                    "sensitivity_tags": data.get("sensitivity_tags") or [],
                    "owner_name": data.get("owner_name") or "",
                    "legal_reference": data.get("legal_reference") or "",
                    "source_note": data.get("source_note") or "",
                    "trigger_config": data.get("trigger_config") or {},
                    "suggested_controls": data.get("suggested_controls") or data.get("control_hierarchy") or {},
                    "suggested_ppe": data.get("suggested_ppe") or data.get("required_ppe") or [],
                    "suggested_protocols": data.get("suggested_protocols") or data.get("protocol_codes") or [],
                    "required_fields": data.get("required_fields") or [],
                    "approval_policy": data.get("approval_policy") or {},
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created({"rule": rule.to_dict()})

    async def update_generator_rule(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        rule, error = self._generator_rule_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        editable_fields = (
            "name",
            "scope_type",
            "process_name",
            "task_name",
            "position_name",
            "hazard_factor",
            "controls_summary",
            "control_hierarchy",
            "required_ppe",
            "protocol_codes",
            "sensitivity_tags",
            "owner_name",
            "legal_reference",
            "source_note",
            "trigger_config",
            "suggested_controls",
            "suggested_ppe",
            "suggested_protocols",
            "required_fields",
            "approval_policy",
        )
        for field in editable_fields:
            if field in data:
                setattr(rule, field, data.get(field))
        if "scope_ref_id" in data:
            rule.scope_ref_id = _safe_int(data.get("scope_ref_id"), None)
        if "master_risk_id" in data:
            rule.master_risk_id = _safe_int(data.get("master_risk_id"), None)
        if "probability" in data:
            rule.probability = _safe_int(data.get("probability"), 2) or 2
        if "consequence" in data:
            rule.consequence = _safe_int(data.get("consequence"), 2) or 2
        if "active" in data:
            rule.active = _normalize_bool(data.get("active"), True)
        try:
            rule.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.ok({"rule": rule.to_dict()})

    async def delete_generator_rule(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        rule, error = self._generator_rule_or_404(request.params.get("id"))
        if error:
            return error
        rule.active = False
        rule.save()
        return Response.ok({"message": "Generator rule archived", "rule": rule.to_dict()})

    async def bot_assistant_suggestions(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_miper_catalog()
        payload = request.data or {}
        context = self._bot_assistant_context(payload)
        suggestions: List[Dict[str, Any]] = []
        for rule in SafetyGeneratorRule.search(self._tenant_filter()):
            if not rule.active:
                continue
            score, reasons = self._score_bot_rule(rule, context)
            if score <= 0:
                continue
            rule_payload = rule.to_dict()
            controls = normalize_control_hierarchy(
                rule_payload.get("suggested_controls") or rule_payload.get("control_hierarchy")
            )
            risk = calculate_risk(rule_payload.get("probability"), rule_payload.get("consequence"))
            suggestions.append(
                {
                    "rule_id": rule.id,
                    "name": rule.name or "Sugerencia preventiva",
                    "score": score,
                    "reasons": reasons,
                    "hazard": {
                        "master_risk_id": rule.master_risk_id,
                        "master_risk_code": rule_payload.get("master_risk_code"),
                        "master_risk_name": rule_payload.get("master_risk_name"),
                        "hazard_factor": rule.hazard_factor or "",
                        "probability": _safe_int(rule.probability, 2) or 2,
                        "consequence": _safe_int(rule.consequence, 2) or 2,
                        "control_hierarchy": controls,
                        "controls_summary": rule.controls_summary or summarize_controls(controls, ""),
                        "required_ppe": _normalize_str_list(rule_payload.get("suggested_ppe") or rule.required_ppe),
                        "protocol_codes": _normalize_str_list(rule_payload.get("suggested_protocols") or rule.protocol_codes),
                        "sensitivity_tags": _normalize_str_list(rule.sensitivity_tags),
                        "legal_reference": rule.legal_reference or "",
                        "source_note": f"Asistente BOT: {rule.name or 'regla preventiva'}",
                        **risk,
                    },
                    "required_fields": _normalize_str_list(rule.required_fields),
                    "approval_policy": rule.approval_policy if isinstance(rule.approval_policy, dict) else {},
                }
            )
        suggestions.sort(key=lambda item: (-int(item.get("score") or 0), str(item.get("name") or "").lower()))
        return Response.ok(
            {
                "count": len(suggestions),
                "results": suggestions[:8],
                "assistant_mode": "embedded",
                "message": "Sugerencias calculadas para el editor BOT.",
            }
        )

    def _bot_assistant_context(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        text_parts = [
            payload.get("code"),
            payload.get("name"),
            payload.get("description"),
            payload.get("base_process"),
            payload.get("default_process_name"),
            payload.get("base_task"),
            payload.get("default_task_name"),
            payload.get("objective"),
            payload.get("context"),
        ]
        resources = payload.get("resources") or []
        resource_types: List[str] = []
        if isinstance(resources, list):
            for resource in resources:
                if isinstance(resource, dict):
                    resource_types.append(str(resource.get("resource_type") or "").strip().lower())
                    text_parts.extend([resource.get("name"), resource.get("description")])
                else:
                    text_parts.append(resource)
        raw_text = " ".join(str(item or "") for item in text_parts).lower()
        tags = set(_normalize_str_list(payload.get("tags") or payload.get("sensitivity_tags")))
        inferred = {
            "altura": ["altura", "arnes", "arnés", "spdc", "plataforma", "borde"],
            "andamios": ["andamio", "andamios", "baranda", "rodapie", "rodapié"],
            "herramienta_electrica": ["herramienta electrica", "herramienta eléctrica", "esmeril", "taladro", "extension electrica", "extensión eléctrica"],
        }
        for tag, needles in inferred.items():
            if any(needle in raw_text for needle in needles):
                tags.add(tag)
        return {
            "text": raw_text,
            "tags": {str(tag).strip().lower() for tag in tags if str(tag).strip()},
            "criticality": str(payload.get("criticality") or "").strip().lower(),
            "block_type": str(payload.get("block_type") or payload.get("type") or "").strip().lower(),
            "resource_types": {item for item in resource_types if item},
        }

    def _score_bot_rule(self, rule: SafetyGeneratorRule, context: Dict[str, Any]) -> Tuple[int, List[str]]:
        trigger = rule.trigger_config if isinstance(rule.trigger_config, dict) else {}
        score = 0
        reasons: List[str] = []
        tags = context.get("tags") or set()
        text = context.get("text") or ""
        for tag in _normalize_str_list(trigger.get("tags_any")):
            if tag.lower() in tags:
                score += 40
                reasons.append(f"Etiqueta/contexto: {tag}")
        for needle in _normalize_str_list(trigger.get("process_contains_any")):
            if needle.lower() and needle.lower() in text:
                score += 35
                reasons.append(f"Proceso detectado: {needle}")
        for criticality in _normalize_str_list(trigger.get("criticality_any")):
            if criticality.lower() == context.get("criticality"):
                score += 35
                reasons.append(f"Criticidad: {criticality}")
        resource_types = context.get("resource_types") or set()
        for resource_type in _normalize_str_list(trigger.get("resource_types_any")):
            if resource_type.lower() in resource_types:
                score += 25
                reasons.append(f"Recurso: {resource_type}")
        if "herramienta" in text and "electr" in text and "herramienta_electrica" in _normalize_str_list(trigger.get("tags_any")):
            score += 30
            reasons.append("Herramienta electrica detectada")
        if not trigger and rule.hazard_factor and rule.hazard_factor.lower() in text:
            score += 20
            reasons.append("Coincidencia por descripcion")
        return score, dedupe_preserve_order(reasons)

    async def list_client_sites(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        customer_id = _safe_int(request.get_param("customer_id"), None)
        sites = SafetyClientSite.search(self._tenant_filter())
        if customer_id:
            sites = [site for site in sites if site.customer_id == customer_id]
        sites.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        return Response.ok({"count": len(sites), "results": [site.to_dict() for site in sites]})

    async def create_client_site(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        try:
            site = SafetyClientSite.create(
                {
                    "customer_id": _safe_int(data.get("customer_id"), None),
                    "name": data.get("name") or "",
                    "address": data.get("address") or "",
                    "comuna": data.get("comuna") or "",
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created({"site": site.to_dict()})

    async def update_client_site(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        site, error = self._site_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        if "customer_id" in data:
            customer_id = _safe_int(data.get("customer_id"), None)
            if not customer_id:
                return Response.bad_request("customer_id is required")
            site.customer_id = customer_id
        for field in ("name", "address", "comuna"):
            if field in data:
                setattr(site, field, data.get(field))
        if "active" in data:
            site.active = _normalize_bool(data.get("active"), True)
        try:
            site.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.ok({"site": site.to_dict()})

    async def delete_client_site(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        site, error = self._site_or_404(request.params.get("id"))
        if error:
            return error
        site.active = False
        site.save()
        return Response.ok({"message": "Client site archived", "site": site.to_dict()})

    async def list_client_areas(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        site_id = _safe_int(request.get_param("site_id"), None)
        areas = SafetyClientArea.search(self._tenant_filter())
        if site_id:
            areas = [area for area in areas if area.site_id == site_id]
        areas.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        return Response.ok({"count": len(areas), "results": [area.to_dict() for area in areas]})

    async def create_client_area(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        try:
            area = SafetyClientArea.create(
                {
                    "site_id": _safe_int(data.get("site_id"), None),
                    "parent_area_id": _safe_int(data.get("parent_area_id"), None),
                    "name": data.get("name") or "",
                    "risk_notes": data.get("risk_notes") or "",
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created({"area": area.to_dict()})

    async def update_client_area(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        area, error = self._area_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        if "site_id" in data:
            site_id = _safe_int(data.get("site_id"), None)
            if not site_id:
                return Response.bad_request("site_id is required")
            site, site_error = self._site_or_404(site_id)
            if site_error:
                return site_error
            area.site_id = site.id
        if "parent_area_id" in data:
            area.parent_area_id = _safe_int(data.get("parent_area_id"), None)
        for field in ("name", "risk_notes"):
            if field in data:
                setattr(area, field, data.get(field))
        if "active" in data:
            area.active = _normalize_bool(data.get("active"), True)
        try:
            area.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.ok({"area": area.to_dict()})

    async def delete_client_area(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        area, error = self._area_or_404(request.params.get("id"))
        if error:
            return error
        area.active = False
        area.save()
        return Response.ok({"message": "Client area archived", "area": area.to_dict()})

    async def list_worker_restrictions(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        employee_id = _safe_int(request.get_param("employee_id"), None)
        restrictions = SafetyWorkerRestriction.search(self._tenant_filter())
        if employee_id:
            restrictions = [item for item in restrictions if item.employee_id == employee_id]
        restrictions.sort(key=lambda item: ((item.title or "").lower(), item.id or 0))
        return Response.ok({"count": len(restrictions), "results": [item.to_dict() for item in restrictions]})

    async def create_worker_restriction(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        employee_id = _safe_int(data.get("employee_id"), None)
        employee_error = self._validate_employee_assignments([employee_id] if employee_id else [])
        if employee_error:
            return employee_error
        try:
            restriction = SafetyWorkerRestriction.create(
                {
                    "employee_id": employee_id,
                    "title": data.get("title") or "",
                    "restriction_type": data.get("restriction_type") or "",
                    "applies_to_tags": data.get("applies_to_tags") or [],
                    "severity": data.get("severity") or "warning",
                    "details": data.get("details") or "",
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        return Response.created({"restriction": restriction.to_dict()})

    async def list_service_profiles(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        self._ensure_profiles()
        profiles = SafetyServiceProfile.search(self._tenant_filter())
        profiles.sort(key=lambda item: ((item.name or "").lower(), item.id or 0))
        return Response.ok({"count": len(profiles), "results": [profile.to_dict() for profile in profiles]})

    async def create_service_profile(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        data = request.data or {}
        try:
            profile = SafetyServiceProfile.create(
                {
                    "name": (data.get("name") or "").strip(),
                    "service_type_id": _safe_int(data.get("service_type_id"), None),
                    "risk_level": data.get("risk_level") or "medium",
                    "mandatory_documents": data.get("mandatory_documents") or [],
                    "mandatory_ppe": data.get("mandatory_ppe") or [],
                    "mandatory_checklists": data.get("mandatory_checklists") or [],
                    "recommended_talks": data.get("recommended_talks") or [],
                    "active": _normalize_bool(data.get("active"), True),
                    "company_id": self._company_id(),
                }
            )
            return Response.created(profile.to_dict())
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def update_service_profile(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        profile, error = self._profile_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        for field in (
            "name",
            "risk_level",
            "mandatory_documents",
            "mandatory_ppe",
            "mandatory_checklists",
            "recommended_talks",
        ):
            if field in data:
                setattr(profile, field, data[field])
        if "service_type_id" in data:
            profile.service_type_id = _safe_int(data.get("service_type_id"), None)
        if "active" in data:
            profile.active = _normalize_bool(data.get("active"), True)
        try:
            profile.save()
            return Response.ok(profile.to_dict())
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def delete_service_profile(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        profile, error = self._profile_or_404(request.params.get("id"))
        if error:
            return error
        profile.active = False
        profile.save()
        return Response.ok({"message": f"Profile '{profile.name}' archived"})

    async def list_folders(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folders = SafetyFolder.search(self._tenant_filter())
        search = (request.get_param("search", "") or "").strip().lower()
        traffic_light = (request.get_param("traffic_light", "") or "").strip().lower()
        status = (request.get_param("status", "") or "").strip().lower()

        payloads = [self._folder_payload(folder) for folder in folders]
        if search:
            filtered: List[Dict[str, Any]] = []
            for item in payloads:
                haystack = " ".join(
                    [
                        str(item.get("project_code") or ""),
                        str(item.get("lead_title") or ""),
                        str(item.get("customer_name") or ""),
                        str(item.get("service_profile_name") or ""),
                        str(item.get("client_site_name") or ""),
                        str(item.get("client_area_name") or ""),
                    ]
                ).lower()
                if search in haystack:
                    filtered.append(item)
            payloads = filtered
        if traffic_light:
            payloads = [item for item in payloads if item.get("traffic_light") == traffic_light]
        if status:
            payloads = [item for item in payloads if item.get("status") == status]

        payloads.sort(
            key=lambda item: (
                str(item.get("planned_start_date") or "9999-12-31"),
                str(item.get("lead_title") or "").lower(),
            )
        )
        return Response.ok({"count": len(payloads), "results": payloads})

    async def create_folder(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err

        lead_id = _safe_int(request.get_data("lead_id"))
        if not lead_id:
            return Response.bad_request("lead_id is required")
        lead, error = self._lead_or_404(lead_id)
        if error:
            return error

        existing = SafetyFolder.search([("lead_id", "=", lead.id), ("company_id", "=", lead.company_id)])
        if existing:
            return Response.bad_request("This lead already has a safety folder")

        self._ensure_miper_catalog()
        profile_id = _safe_int(request.get_data("service_profile_id"), None)
        profile = None
        if profile_id:
            profile, error = self._profile_or_404(profile_id)
            if error:
                return error
        else:
            profile = self._find_profile_for_lead(lead)

        procedure_id = _safe_int(request.get_data("procedure_id"), None)
        procedure = None
        if procedure_id:
            procedure, error = self._procedure_or_404(procedure_id)
            if error:
                return error
        else:
            procedure = self._find_procedure_for_profile(profile)
        procedure_ids = _normalize_int_list(request.get_data("procedure_ids") or [])
        if not procedure_ids and procedure:
            procedure_ids = [procedure.id]
        procedures, procedure_error = self._validate_procedure_ids(procedure_ids)
        if procedure_error:
            return procedure_error
        procedure = procedures[0] if procedures else procedure

        job_profile_ids = _normalize_int_list(request.get_data("job_profile_ids") or [])
        _, job_profile_error = self._validate_job_profile_ids(job_profile_ids)
        if job_profile_error:
            return job_profile_error

        equipment_block_ids = _normalize_int_list(request.get_data("equipment_block_ids") or [])
        _, equipment_error = self._validate_equipment_block_ids(equipment_block_ids)
        if equipment_error:
            return equipment_error

        assigned_employee_ids = _normalize_int_list(request.get_data("assigned_employee_ids") or [])
        employee_error = self._validate_employee_assignments(assigned_employee_ids)
        if employee_error:
            return employee_error
        site_id = _safe_int(request.get_data("client_site_id"), None)
        area_id = _safe_int(request.get_data("client_area_id"), None)
        area_ids = _normalize_int_list(request.get_data("client_area_ids") or [])
        if not area_ids and area_id:
            area_ids = [area_id]
        site = None
        area = None
        if site_id:
            site, site_error = self._site_or_404(site_id)
            if site_error:
                return site_error
                if lead.customer_id and site.customer_id != lead.customer_id:
                    return Response.bad_request("Selected client site does not belong to the lead customer")
        if area_id:
            area, area_error = self._area_or_404(area_id)
            if area_error:
                return area_error
            if site and area.site_id != site.id:
                return Response.bad_request("Selected area does not belong to the chosen site")
            if not site:
                site = SafetyClientSite.find_by_id(area.site_id)
                site_id = site.id if site else None
        if site_id and area_ids:
            area_ids = [
                area_candidate_id
                for area_candidate_id in area_ids
                if (SafetyClientArea.find_by_id(area_candidate_id) and SafetyClientArea.find_by_id(area_candidate_id).site_id == site_id)
            ]
        areas, area_ids_error = self._validate_client_area_ids(area_ids, lead=lead, site_id=site_id)
        if area_ids_error:
            return area_ids_error
        if areas and not site_id:
            first_site = SafetyClientSite.find_by_id(areas[0].site_id)
            site_id = first_site.id if first_site else None
            site = first_site
        area_id = areas[0].id if areas else area_id

        try:
            folder = SafetyFolder.create(
                {
                    "lead_id": lead.id,
                    "service_profile_id": profile.id if profile else None,
                    "procedure_id": procedure.id if procedure else None,
                    "procedure_ids": [item.id for item in procedures] if procedures else procedure_ids,
                    "job_profile_ids": job_profile_ids,
                    "client_site_id": site_id,
                    "client_area_id": area_id,
                    "client_area_ids": [item.id for item in areas] if areas else area_ids,
                    "equipment_block_ids": equipment_block_ids,
                    "status": "draft",
                    "readiness_pct": 0.0,
                    "traffic_light": "red",
                    "planned_start_date": request.get_data("planned_start_date") or "",
                    "assigned_employee_ids": assigned_employee_ids,
                    "notes": request.get_data("notes") or "",
                    "miper_scope_notes": request.get_data("miper_scope_notes") or "",
                    "company_id": self._company_id(),
                }
            )
            self._generate_documents_for_folder(folder)
            payload = self._folder_payload(folder)
            self._log_on_lead(folder.lead_id, "Safety Folder Created", "Se creo carpeta de seguridad para la oportunidad")
            return Response.created(payload)
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def get_folder(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        return Response.ok(self._folder_payload(folder))

    async def update_folder(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error

        data = request.data or {}
        should_regenerate_matrix = False
        if "planned_start_date" in data:
            folder.planned_start_date = data.get("planned_start_date") or ""
        if "notes" in data:
            folder.notes = data.get("notes") or ""
        if "miper_scope_notes" in data:
            folder.miper_scope_notes = data.get("miper_scope_notes") or ""
        if "assigned_employee_ids" in data:
            employee_ids = _normalize_int_list(data.get("assigned_employee_ids"))
            employee_error = self._validate_employee_assignments(employee_ids)
            if employee_error:
                return employee_error
            folder.assigned_employee_ids = employee_ids
            should_regenerate_matrix = True
        if "client_site_id" in data:
            site_id = _safe_int(data.get("client_site_id"), None)
            if site_id:
                site, site_error = self._site_or_404(site_id)
                if site_error:
                    return site_error
                folder.client_site_id = site.id
                folder.client_area_ids = [
                    area_id
                    for area_id in self._folder_client_area_ids(folder)
                    if (SafetyClientArea.find_by_id(area_id) and SafetyClientArea.find_by_id(area_id).site_id == site.id)
                ]
                folder.client_area_id = folder.client_area_ids[0] if folder.client_area_ids else None
            else:
                folder.client_site_id = None
                folder.client_area_id = None
                folder.client_area_ids = []
            should_regenerate_matrix = True
        if "client_area_id" in data:
            area_id = _safe_int(data.get("client_area_id"), None)
            if area_id:
                area, area_error = self._area_or_404(area_id)
                if area_error:
                    return area_error
                if folder.client_site_id and area.site_id != folder.client_site_id:
                    return Response.bad_request("Selected area does not belong to the chosen site")
                folder.client_area_id = area.id
                if not folder.client_site_id:
                    folder.client_site_id = area.site_id
            else:
                folder.client_area_id = None
            should_regenerate_matrix = True
        if "client_area_ids" in data:
            area_ids = _normalize_int_list(data.get("client_area_ids"))
            try:
                from modules.crm.module_crm import Lead

                lead = Lead.find_by_id(folder.lead_id)
            except Exception:
                lead = None
            areas, area_ids_error = self._validate_client_area_ids(
                area_ids,
                lead=lead,
                site_id=_safe_int(folder.client_site_id, None),
            )
            if area_ids_error:
                return area_ids_error
            folder.client_area_ids = [area.id for area in areas]
            folder.client_area_id = folder.client_area_ids[0] if folder.client_area_ids else None
            if areas and not folder.client_site_id:
                folder.client_site_id = areas[0].site_id
            should_regenerate_matrix = True
        if "service_profile_id" in data:
            profile_id = _safe_int(data.get("service_profile_id"), None)
            if profile_id:
                profile, profile_error = self._profile_or_404(profile_id)
                if profile_error:
                    return profile_error
                folder.service_profile_id = profile.id
                if not folder.procedure_id:
                    procedure = self._find_procedure_for_profile(profile)
                    if procedure:
                        folder.procedure_id = procedure.id
                self._generate_documents_for_folder(folder)
            else:
                folder.service_profile_id = None
            should_regenerate_matrix = True
        if "procedure_id" in data:
            procedure_id = _safe_int(data.get("procedure_id"), None)
            if procedure_id:
                procedure, procedure_error = self._procedure_or_404(procedure_id)
                if procedure_error:
                    return procedure_error
                folder.procedure_id = procedure.id
            else:
                folder.procedure_id = None
            self._generate_documents_for_folder(folder)
            should_regenerate_matrix = True
        if "procedure_ids" in data:
            procedure_ids = _normalize_int_list(data.get("procedure_ids"))
            procedures, procedure_ids_error = self._validate_procedure_ids(procedure_ids)
            if procedure_ids_error:
                return procedure_ids_error
            folder.procedure_ids = [procedure.id for procedure in procedures]
            folder.procedure_id = folder.procedure_ids[0] if folder.procedure_ids else None
            self._generate_documents_for_folder(folder)
            should_regenerate_matrix = True
        if "job_profile_ids" in data:
            job_profile_ids = _normalize_int_list(data.get("job_profile_ids"))
            _, job_profile_error = self._validate_job_profile_ids(job_profile_ids)
            if job_profile_error:
                return job_profile_error
            folder.job_profile_ids = job_profile_ids
            should_regenerate_matrix = True
        if "equipment_block_ids" in data:
            equipment_block_ids = _normalize_int_list(data.get("equipment_block_ids"))
            _, equipment_error = self._validate_equipment_block_ids(equipment_block_ids)
            if equipment_error:
                return equipment_error
            folder.equipment_block_ids = equipment_block_ids
            should_regenerate_matrix = True

        try:
            self._sync_folder_scope(folder)
            folder.save()
            summary = self._refresh_folder_metrics(folder)
        except ValidationError as exc:
            return Response.bad_request(str(exc))

        if "status" in data:
            new_status = str(data.get("status") or "").strip()
            if new_status not in FOLDER_STATUSES:
                return Response.bad_request(f"status must be one of: {', '.join(FOLDER_STATUSES)}")
            if new_status in ("ready", "in_progress") and summary["critical_blockers"]:
                return Response.bad_request(
                    "Cannot move folder to ready or in_progress while critical blockers remain"
                )
            folder.status = new_status
            if new_status == "ready":
                folder.approver_user_id = self.env.user.id if self.env.user else None
                folder.approved_at = utc_now()
            elif new_status != "closed":
                folder.approver_user_id = None
                folder.approved_at = None
            folder.save()
            summary = self._refresh_folder_metrics(folder)

        if should_regenerate_matrix:
            self._generate_matrix_for_folder(folder)
            summary = self._refresh_folder_metrics(folder)

        self._log_on_lead(folder.lead_id, "Safety Folder Updated", "Se actualizaron datos de la carpeta de seguridad")
        return Response.ok({**folder.to_dict(), "summary": summary})

    async def delete_folder(self, request: Request) -> Response:
        err = self._require_admin()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error

        for document in self._documents_for_folder(folder.id):
            document.delete()
        matrix = self._matrix_for_folder(folder.id)
        if matrix:
            matrix.delete()
        for delivery in self._deliveries_for_folder(folder.id):
            delivery.delete()
        for talk in self._talks_for_folder(folder.id):
            talk.delete()
        for checklist in self._checklists_for_folder(folder.id):
            checklist.delete()
        folder.delete()
        self._log_on_lead(folder.lead_id, "Safety Folder Deleted", "Se elimino la carpeta de seguridad")
        return Response.ok({"message": "Safety folder deleted"})

    async def folder_dossier(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error

        matrix = self._matrix_for_folder(folder.id)
        payload = self._folder_payload(folder)
        return Response.ok(
            {
                "folder": payload,
                "documents": [document.to_dict() for document in self._documents_for_folder(folder.id)],
                "risk_matrix": matrix.to_dict() if matrix else None,
                "irl_records": [record.to_dict() for record in self._irl_records_for_folder(folder.id)],
                "ppe_deliveries": [delivery.to_dict() for delivery in self._deliveries_for_folder(folder.id)],
                "talks": [talk.to_dict() for talk in self._talks_for_folder(folder.id)],
                "checklists": [checklist.to_dict() for checklist in self._checklists_for_folder(folder.id)],
                "lookups": self._lookups_payload(),
                "libraries": {
                    "checklists": CHECKLIST_LIBRARY,
                    "talks": TALK_LIBRARY,
                },
            }
        )

    async def generate_documents(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        stats = self._generate_documents_for_folder(folder)
        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "Safety Documents Generated", "Se regeneraron documentos base de seguridad")
        return Response.ok({"folder": payload, "generation": stats})

    async def generate_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        matrix, generation_summary = self._generate_matrix_for_folder(folder)
        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "MIPER Generated", "Se genero o actualizo la matriz automatizada")
        return Response.ok(
            {
                "folder": payload,
                "risk_matrix": matrix.to_dict(),
                "generation_summary": generation_summary,
            }
        )

    async def list_documents(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        documents = [document.to_dict() for document in self._documents_for_folder(folder.id)]
        return Response.ok({"count": len(documents), "results": documents})

    async def create_document(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        try:
            document = SafetyFolderDocument.create(
                {
                    "folder_id": folder.id,
                    "code": (data.get("code") or data.get("title") or f"doc_{folder.id}").strip().lower().replace(" ", "_"),
                    "title": (data.get("title") or "").strip(),
                    "document_type": data.get("document_type") or "other",
                    "status": data.get("status") or "draft",
                    "version": _safe_int(data.get("version"), 1) or 1,
                    "is_critical": _normalize_bool(data.get("is_critical"), False),
                    "owner_user_id": _safe_int(data.get("owner_user_id"), self.env.user.id if self.env.user else None),
                    "due_date": data.get("due_date") or "",
                    "content": data.get("content") or "",
                    "company_id": folder.company_id,
                }
            )
            payload = self._folder_payload(folder)
            self._log_on_lead(folder.lead_id, "Safety Document Added", f"Documento agregado: {document.title}")
            return Response.created({"document": document.to_dict(), "folder": payload})
        except ValidationError as exc:
            return Response.bad_request(str(exc))

    async def update_document(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        document, error = self._document_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        for field in ("code", "title", "document_type", "status", "due_date", "content"):
            if field in data:
                setattr(document, field, data.get(field))
        if "version" in data:
            document.version = _safe_int(data.get("version"), 1) or 1
        if "is_critical" in data:
            document.is_critical = _normalize_bool(data.get("is_critical"), False)
        if "owner_user_id" in data:
            document.owner_user_id = _safe_int(data.get("owner_user_id"), None)
        try:
            document.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        folder = SafetyFolder.find_by_id(document.folder_id)
        if folder:
            payload = self._folder_payload(folder)
            self._log_on_lead(folder.lead_id, "Safety Document Updated", f"Documento actualizado: {document.title}")
            return Response.ok({"document": document.to_dict(), "folder": payload})
        return Response.ok({"document": document.to_dict()})

    async def delete_document(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        document, error = self._document_or_404(request.params.get("id"))
        if error:
            return error
        folder = SafetyFolder.find_by_id(document.folder_id)
        title = document.title
        document.delete()
        if folder:
            payload = self._folder_payload(folder)
            self._log_on_lead(folder.lead_id, "Safety Document Deleted", f"Documento eliminado: {title}")
            return Response.ok({"message": "Document deleted", "folder": payload})
        return Response.ok({"message": "Document deleted"})

    async def list_irl_records(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        records = [record.to_dict() for record in self._irl_records_for_folder(folder.id)]
        return Response.ok({"count": len(records), "results": records})

    async def generate_irl_record(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        employee_id = _safe_int(data.get("employee_id"), None)
        employee_payload = self._employee_payload(employee_id)
        if not employee_payload and not str(data.get("position_title") or "").strip():
            return Response.bad_request("Selecciona un trabajador o informa un cargo para generar el IRL")

        draft = self._build_irl_draft_payload(folder, employee_payload, data)
        existing_versions = [
            int(record.version or 0)
            for record in self._irl_records_for_folder(folder.id)
            if (
                (_safe_int(record.employee_id, None) == _safe_int(draft.get("employee_id"), None))
                and ((record.position_title or "").strip().lower() == str(draft.get("position_title") or "").strip().lower())
            )
        ]
        draft["version"] = (max(existing_versions) if existing_versions else 0) + 1

        try:
            record = SafetyIRLRecord.create(draft)
        except ValidationError as exc:
            return Response.bad_request(str(exc))

        self._log_on_lead(folder.lead_id, "IRL Generated", f"Se genero IRL: {record.title}")
        return Response.created({"irl_record": record.to_dict(), "folder": self._folder_payload(folder)})

    async def update_irl_record(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        irl, error = self._irl_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}

        if "employee_id" in data:
            irl.employee_id = _safe_int(data.get("employee_id"), None)
        if "version" in data:
            irl.version = _safe_int(data.get("version"), 1) or 1
        for field in (
            "title",
            "status",
            "worker_name",
            "worker_identifier",
            "position_title",
            "place_name",
            "activity_name",
            "activity_period",
            "modality",
            "duration_hours",
            "executor_name",
            "relator_background",
            "target_group",
            "workspace_features",
            "environmental_conditions",
            "order_cleanliness",
            "machines_tools",
            "observations",
            "intro_text",
            "theme_color",
        ):
            if field in data:
                setattr(irl, field, data.get(field))
        if "service_functions" in data:
            irl.service_functions = data.get("service_functions") or []
        if "risk_items" in data:
            irl.risk_items = data.get("risk_items") or []
        if "complement_materials" in data:
            irl.complement_materials = data.get("complement_materials") or []

        try:
            irl.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))

        folder = SafetyFolder.find_by_id(irl.folder_id)
        if folder:
            self._log_on_lead(folder.lead_id, "IRL Updated", f"Se actualizo IRL: {irl.title}")
        return Response.ok({"irl_record": irl.to_dict(), "folder": self._folder_payload(folder) if folder else None})

    async def delete_irl_record(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        irl, error = self._irl_or_404(request.params.get("id"))
        if error:
            return error
        folder = SafetyFolder.find_by_id(irl.folder_id)
        title = irl.title
        irl.delete()
        if folder:
            self._log_on_lead(folder.lead_id, "IRL Deleted", f"Se elimino IRL: {title}")
        return Response.ok({"message": "IRL deleted", "folder": self._folder_payload(folder) if folder else None})

    async def export_irl_pdf(self, request: Request):
        err = self._require_access()
        if err:
            return err
        irl, error = self._irl_or_404(request.params.get("id"))
        if error:
            return error
        from fastapi.responses import StreamingResponse

        content = self._build_irl_pdf(irl)
        filename = self._irl_filename(irl, "pdf")
        return StreamingResponse(
            BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    async def upsert_risk_matrix(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        matrix = self._matrix_for_folder(folder.id)
        try:
            if matrix:
                if "title" in data:
                    matrix.title = data.get("title") or matrix.title
                if "status" in data:
                    matrix.status = data.get("status") or matrix.status
                if "version" in data:
                    matrix.version = _safe_int(data.get("version"), 1) or 1
                if "rows" in data:
                    matrix.rows = data.get("rows") or []
                if "generation_summary" in data and isinstance(data.get("generation_summary"), dict):
                    matrix.generation_summary = data.get("generation_summary") or {}
                if matrix.status == "approved":
                    matrix.reviewed_by = self.env.user.id if self.env.user else None
                    matrix.reviewed_at = utc_now()
                matrix.save()
            else:
                matrix = SafetyRiskMatrix.create(
                    {
                        "folder_id": folder.id,
                        "title": data.get("title") or "Matriz de riesgo",
                        "status": data.get("status") or "draft",
                        "version": _safe_int(data.get("version"), 1) or 1,
                        "rows": data.get("rows") or [],
                        "generation_summary": data.get("generation_summary") or {},
                        "reviewed_by": self.env.user.id if (data.get("status") == "approved" and self.env.user) else None,
                        "reviewed_at": utc_now() if data.get("status") == "approved" else None,
                        "company_id": folder.company_id,
                    }
                )
        except ValidationError as exc:
            return Response.bad_request(str(exc))

        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "Risk Matrix Updated", "Se actualizo la matriz de riesgo")
        return Response.ok({"risk_matrix": matrix.to_dict(), "folder": payload})

    def _generate_delivery_documents(
        self,
        folder: SafetyFolder,
        delivery: SafetyPPEDelivery,
        template_ids: List[int],
        request: Optional[Request],
    ) -> Dict[str, Any]:
        if not template_ids:
            return {"generated_documents": [], "summary": {"documents_generated": 0}}
        document_center = self.core.module_registry.get_module("document_center")
        if not document_center or not hasattr(document_center, "generate_worker_documents_internal"):
            raise ValidationError("Document center module is not available")

        try:
            from modules.crm.module_crm import Lead

            lead = Lead.find_by_id(folder.lead_id) if folder.lead_id else None
        except Exception:
            lead = None
        folder_payload = folder.to_dict()
        payload = {
            "employee_id": delivery.employee_id,
            "template_ids": template_ids,
            "source_module": "safety",
            "source_record_id": delivery.id,
            "source_label": f"EPP {folder_payload.get('project_code') or folder.id}",
            "target_module": "safety",
            "target_record_id": folder.id,
            "customer_id": lead.customer_id if lead else folder_payload.get("customer_id"),
            "lead_id": folder.lead_id,
            "safety_folder_id": folder.id,
            "document_date": delivery.delivery_date,
            "issued_on": delivery.delivery_date,
            "delivery_date": delivery.delivery_date,
            "detail_items": delivery.items or [],
            "notes": delivery.notes or "",
            "extra_context": {
                "ppe_delivery_id": delivery.id,
                "ppe_status": delivery.status or "",
                "ppe_notes": delivery.notes or "",
                "delivery_items": ", ".join(delivery.items or []),
                "delivery_items_multiline": "\n".join(delivery.items or []),
            },
        }
        return document_center.generate_worker_documents_internal(payload, request=request)

    async def list_ppe_deliveries(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        deliveries = [delivery.to_dict() for delivery in self._deliveries_for_folder(folder.id)]
        return Response.ok({"count": len(deliveries), "results": deliveries})

    async def create_ppe_delivery(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        employee_id = _safe_int(request.get_data("employee_id"))
        if not employee_id:
            return Response.bad_request("employee_id is required")
        employee_error = self._validate_employee_assignments([employee_id])
        if employee_error:
            return employee_error
        try:
            delivery = SafetyPPEDelivery.create(
                {
                    "folder_id": folder.id,
                    "employee_id": employee_id,
                    "delivery_date": request.get_data("delivery_date") or utc_today_iso(),
                    "status": request.get_data("status") or "delivered",
                    "items": request.get_data("items") or [],
                    "notes": request.get_data("notes") or "",
                    "company_id": folder.company_id,
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        template_ids = _normalize_int_list(
            request.get_data("document_template_ids") or request.get_data("document_template_id")
        )
        generated_documents: List[Dict[str, Any]] = []
        generation_error = ""
        if template_ids:
            try:
                generation_result = self._generate_delivery_documents(
                    folder,
                    delivery,
                    template_ids,
                    request,
                )
                generated_documents = generation_result.get("generated_documents") or []
            except Exception as exc:
                generation_error = str(exc)
                self.logger.warning(f"PPE delivery document generation skipped: {exc}")
        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "PPE Delivery Registered", "Se registro entrega de EPP")
        return Response.created(
            {
                "delivery": delivery.to_dict(),
                "folder": payload,
                "generated_documents": generated_documents,
                "document_generation_error": generation_error,
            }
        )

    async def update_ppe_delivery(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        delivery, error = self._delivery_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        if "employee_id" in data:
            employee_id = _safe_int(data.get("employee_id"))
            employee_error = self._validate_employee_assignments([employee_id] if employee_id else [])
            if employee_error:
                return employee_error
            delivery.employee_id = employee_id
        for field in ("delivery_date", "status", "items", "notes"):
            if field in data:
                setattr(delivery, field, data.get(field))
        try:
            delivery.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        folder = SafetyFolder.find_by_id(delivery.folder_id)
        payload = self._folder_payload(folder) if folder else None
        template_ids = _normalize_int_list(
            data.get("document_template_ids") or data.get("document_template_id")
        )
        generated_documents: List[Dict[str, Any]] = []
        generation_error = ""
        if folder and template_ids:
            try:
                generation_result = self._generate_delivery_documents(
                    folder,
                    delivery,
                    template_ids,
                    request,
                )
                generated_documents = generation_result.get("generated_documents") or []
            except Exception as exc:
                generation_error = str(exc)
                self.logger.warning(f"PPE delivery document regeneration skipped: {exc}")
        if folder:
            self._log_on_lead(folder.lead_id, "PPE Delivery Updated", "Se actualizo una entrega de EPP")
        return Response.ok(
            {
                "delivery": delivery.to_dict(),
                "folder": payload,
                "generated_documents": generated_documents,
                "document_generation_error": generation_error,
            }
        )

    async def delete_ppe_delivery(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        delivery, error = self._delivery_or_404(request.params.get("id"))
        if error:
            return error
        folder = SafetyFolder.find_by_id(delivery.folder_id)
        delivery.delete()
        payload = self._folder_payload(folder) if folder else None
        if folder:
            self._log_on_lead(folder.lead_id, "PPE Delivery Deleted", "Se elimino un registro de entrega de EPP")
        return Response.ok({"message": "PPE delivery deleted", "folder": payload})

    async def list_talks(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        talks = [talk.to_dict() for talk in self._talks_for_folder(folder.id)]
        return Response.ok({"count": len(talks), "results": talks})

    async def create_talk(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        attendee_ids = _normalize_int_list(request.get_data("attendee_ids") or [])
        employee_error = self._validate_employee_assignments(attendee_ids)
        if employee_error:
            return employee_error
        try:
            talk = SafetyTalk.create(
                {
                    "folder_id": folder.id,
                    "talk_date": request.get_data("talk_date") or utc_today_iso(),
                    "topic": request.get_data("topic") or "",
                    "speaker_user_id": _safe_int(request.get_data("speaker_user_id"), self.env.user.id if self.env.user else None),
                    "attendee_ids": attendee_ids,
                    "notes": request.get_data("notes") or "",
                    "company_id": folder.company_id,
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "Safety Talk Registered", f"Se registro charla: {talk.topic}")
        return Response.created({"talk": talk.to_dict(), "folder": payload})

    async def update_talk(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        talk, error = self._talk_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        if "attendee_ids" in data:
            attendee_ids = _normalize_int_list(data.get("attendee_ids") or [])
            employee_error = self._validate_employee_assignments(attendee_ids)
            if employee_error:
                return employee_error
            talk.attendee_ids = attendee_ids
        for field in ("talk_date", "topic", "notes"):
            if field in data:
                setattr(talk, field, data.get(field))
        if "speaker_user_id" in data:
            talk.speaker_user_id = _safe_int(data.get("speaker_user_id"), None)
        try:
            talk.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        folder = SafetyFolder.find_by_id(talk.folder_id)
        payload = self._folder_payload(folder) if folder else None
        if folder:
            self._log_on_lead(folder.lead_id, "Safety Talk Updated", f"Se actualizo charla: {talk.topic}")
        return Response.ok({"talk": talk.to_dict(), "folder": payload})

    async def delete_talk(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        talk, error = self._talk_or_404(request.params.get("id"))
        if error:
            return error
        folder = SafetyFolder.find_by_id(talk.folder_id)
        talk.delete()
        payload = self._folder_payload(folder) if folder else None
        if folder:
            self._log_on_lead(folder.lead_id, "Safety Talk Deleted", "Se elimino una charla de seguridad")
        return Response.ok({"message": "Talk deleted", "folder": payload})

    async def list_checklists(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        checklists = [checklist.to_dict() for checklist in self._checklists_for_folder(folder.id)]
        return Response.ok({"count": len(checklists), "results": checklists})

    async def create_checklist(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        folder, error = self._folder_or_404(request.params.get("id"))
        if error:
            return error
        try:
            checklist = SafetyChecklistRun.create(
                {
                    "folder_id": folder.id,
                    "checklist_name": request.get_data("checklist_name") or "",
                    "checklist_type": request.get_data("checklist_type") or "",
                    "executed_at": request.get_data("executed_at") or utc_today_iso(),
                    "executed_by": _safe_int(request.get_data("executed_by"), self.env.user.id if self.env.user else None),
                    "result": request.get_data("result") or "pending",
                    "items": request.get_data("items") or [],
                    "findings": request.get_data("findings") or "",
                    "requires_action": _normalize_bool(request.get_data("requires_action"), False),
                    "company_id": folder.company_id,
                }
            )
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        payload = self._folder_payload(folder)
        self._log_on_lead(folder.lead_id, "Safety Checklist Registered", f"Se registro checklist: {checklist.checklist_name}")
        return Response.created({"checklist": checklist.to_dict(), "folder": payload})

    async def update_checklist(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        checklist, error = self._checklist_or_404(request.params.get("id"))
        if error:
            return error
        data = request.data or {}
        for field in ("checklist_name", "checklist_type", "executed_at", "result", "items", "findings"):
            if field in data:
                setattr(checklist, field, data.get(field))
        if "executed_by" in data:
            checklist.executed_by = _safe_int(data.get("executed_by"), None)
        if "requires_action" in data:
            checklist.requires_action = _normalize_bool(data.get("requires_action"), False)
        try:
            checklist.save()
        except ValidationError as exc:
            return Response.bad_request(str(exc))
        folder = SafetyFolder.find_by_id(checklist.folder_id)
        payload = self._folder_payload(folder) if folder else None
        if folder:
            self._log_on_lead(folder.lead_id, "Safety Checklist Updated", f"Se actualizo checklist: {checklist.checklist_name}")
        return Response.ok({"checklist": checklist.to_dict(), "folder": payload})

    async def delete_checklist(self, request: Request) -> Response:
        err = self._require_access()
        if err:
            return err
        checklist, error = self._checklist_or_404(request.params.get("id"))
        if error:
            return error
        folder = SafetyFolder.find_by_id(checklist.folder_id)
        checklist.delete()
        payload = self._folder_payload(folder) if folder else None
        if folder:
            self._log_on_lead(folder.lead_id, "Safety Checklist Deleted", "Se elimino un checklist")
        return Response.ok({"message": "Checklist deleted", "folder": payload})
